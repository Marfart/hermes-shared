#!/usr/bin/env python3
"""
Build enriched JSON from BLIIoT_Customer_Leads_v2.xlsx → Hermes WhatsApp pipeline format.

Usage:
    python build_enriched_for_whatsapp.py [XLSX_PATH] [OUTPUT_DIR]

Output:
    iiot_search_enriched_<YYYY>-<MM>-<DD>.json (with whatsappNumbers, inferredNeeds, recommendedProducts)

This bridges the BLIIoT lead mining output (Excel) into the Hermes CDP WhatsApp pipeline.
"""
import openpyxl, json, sys, os
from datetime import date

def infer_needs(desc, need, category):
    t = (desc + ' ' + need + ' ' + category).lower()
    needs = []
    if any(k in t for k in ['scada', 'plc', 'dcs', 'protocol']):
        needs.append('PLC connectivity and industrial protocol conversion')
    if any(k in t for k in ['remote monitoring', 'remote']):
        needs.append('remote monitoring gateway deployment')
    if any(k in t for k in ['automation', 'integrator']):
        needs.append('automation gateway and edge integration')
    if any(k in t for k in ['iiot', 'iot', 'industry 4.0']):
        needs.append('IIoT platform connectivity and field data access')
    if any(k in t for k in ['energy', 'utility', 'bems', 'building']):
        needs.append('energy monitoring and building/utility management')
    if any(k in t for k in ['edge', 'ai']):
        needs.append('industrial edge computing hardware')
    if any(k in t for k in ['cloud']):
        needs.append('cloud-connected industrial data transmission')
    if not needs:
        needs.append('IIoT platform connectivity and field data access')
    return needs

def recommend_products(desc, need, category):
    t = (desc + ' ' + need + ' ' + category).lower()
    prods = [
        'BLIIOT gateways/routers (R40, BA/BE/BL gateway series)',
        'ARMxy edge computers/controllers'
    ]
    if any(k in t for k in ['monitor', 'data', 'acquisition', 'io']):
        prods.append('Remote IO / RTU / data acquisition modules')
    return prods

def classify_style(desc, need, category):
    """Classify lead into message style bucket."""
    t = (desc + ' ' + need + ' ' + category).lower()
    if any(k in t for k in ['plc', 'scada', 'dcs', 'protocol']):
        return 'scada_plc'
    if any(k in t for k in ['energy', 'utility', 'bems', 'building']):
        return 'energy_monitoring'
    if any(k in t for k in ['integrator', 'integration', 'system']):
        return 'system_integrator'
    return 'generic_iiot'

def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else \
        'C:/Users/Admin/Desktop/Working/BLIIoT_Customer_Leads_v2.xlsx'
    out_dir = sys.argv[2] if len(sys.argv) > 2 else \
        'C:/Users/Admin/AppData/Local/hermes/memories/buyer-development'

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['BLIIoT High-Potential Leads']

    enriched = []
    excluded = []

    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if not name:
            continue
        desc = ws.cell(r, 3).value or ''
        need = ws.cell(r, 4).value or ''
        phone_raw = ws.cell(r, 5).value or ''
        phone_str = str(phone_raw).replace(' ', '').replace('-', '')
        website = ws.cell(r, 6).value or ''
        linkedin = ws.cell(r, 7).value or ''
        country = ws.cell(r, 9).value or ''
        category = ws.cell(r, 10).value or ''

        # Skip existing customers
        if 'EXISTING' in category.upper() or 'ZODSAT' in name.upper() or 'POWERTEL' in name.upper():
            excluded.append(name)
            continue

        domain = website.replace('https://', '').replace('http://', '').replace('www.', '').rstrip('/') if website else ''

        item = {
            'name': name,
            'country': country,
            'domain': domain,
            'desc': desc,
            'emailCount': 0,
            'reasons': ['公司业务与 IIoT Integration 匹配;'],
            'social': [],
            'normalizedDomain': domain,
            'officialEmails': [],
            'linkedinUrls': [linkedin] if linkedin else [],
            'whatsappUrls': [f'https://wa.me/{phone_str}'],
            'whatsappNumbers': [phone_str],
            'fetchedUrls': [website] if website else [],
            'fitScore': 100,
            'inferredNeeds': infer_needs(desc, need, category),
            'recommendedProducts': recommend_products(desc, need, category),
        }
        enriched.append(item)

    today = date.today().isoformat()
    out_path = os.path.join(out_dir, f'iiot_search_enriched_{today}.json')
    os.makedirs(out_dir, exist_ok=True)
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(enriched, f, indent=2, ensure_ascii=False)

    print(f'Enriched {len(enriched)} leads → {out_path}')
    print(f'Excluded {len(excluded)} existing customers: {excluded}')

if __name__ == '__main__':
    main()