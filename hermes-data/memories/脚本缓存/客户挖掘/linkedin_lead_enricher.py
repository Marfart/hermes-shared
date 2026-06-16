#!/usr/bin/env python3
"""
LinkedIn Lead Enricher — 客户信息增强管道
输入：lead_source.xlsx（DDG X-Ray产出）
流程：
  1. 清洗公司名（从标题/描述中提取真实公司名）
  2. 对每家公司搜官网 → 搜CEO/决策人 → 搜邮箱/电话
  3. 输出 enriched_leads.xlsx

用法：python linkedin_lead_enricher.py
"""

import re, json, time, logging, csv, os, sys
import urllib.request, urllib.parse
from dataclasses import dataclass, field, asdict
from datetime import datetime
from typing import Optional, List, Dict
from pathlib import Path
from collections import defaultdict

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("enricher")

# ──────────────────────────
# Config
# ──────────────────────────

INPUT_FILE = r"C:\Users\Admin\Desktop\Working\linkedin_xray_0604_1517.xlsx"
OUTPUT_DIR = Path(r"C:\Users\Admin\Desktop\Working")

# ──────────────────────────
# Data Model
# ──────────────────────────

@dataclass
class EnrichedLead:
    # Original
    full_name: str
    job_title: str
    company_raw: str  # original messy company
    country: str
    linkedin_url: str
    description: str
    
    # Cleaned
    company_clean: str = ""
    
    # Enriched fields
    company_website: str = ""
    company_phone: str = ""
    company_email: str = ""
    ceo_name: str = ""
    ceo_email: str = ""
    ceo_phone: str = ""
    decision_maker_name: str = ""
    decision_maker_title: str = ""
    decision_maker_email: str = ""
    alternative_contacts: str = ""  # JSON array of {name, title}
    
    # Metadata
    search_queries_used: str = ""
    enrichment_status: str = "pending"

# ──────────────────────────
# Company Name Cleaner
# ──────────────────────────

def extract_company_from_parts(name: str, job: str, company_raw: str, desc: str, url: str) -> str:
    """Extract clean company name from all available signals."""
    
    candidates = []
    
    # 1. Try to extract from description - often has "Experience: Company" pattern
    m = re.search(r'Experience[:\s]+((?:[A-Z][A-Za-z0-9&.]+(?:\s[A-Z][A-Za-z0-9&.]+){0,4}))', desc)
    if m:
        candidates.append(m.group(1).strip())
    
    # 2. Parse job title for "at company" pattern
    m = re.search(r'(?:at|@)\s+([A-Z][A-Za-z0-9&.\s]{2,40}?)(?:[,]|\s+[|]|\s*$)', job)
    if m:
        candidates.append(m.group(1).strip())
    
    # 3. Raw company field - but clean it
    # Remove LinkedIn noise like trailing names/URLs
    cleaned = re.sub(r'\s*\| LinkedIn.*', '', company_raw)
    cleaned = re.sub(r'\s*LinkedIn.*', '', cleaned)
    cleaned = re.sub(r'\s*\.\s*$', '', cleaned)
    # Remove "..." at end
    cleaned = re.sub(r'\s*\.{3,}\s*$', '', cleaned)
    if cleaned and len(cleaned) > 2:
        candidates.append(cleaned.strip())
    
    # 4. Try to extract from url pattern: za.linkedin.com/in/name-company-
    # Sometimes the company name is in the slug
    slug_match = re.search(r'/in/[^/]+-(\w+(?:-\w+){0,3})-?\d+', url)
    
    # 5. Check if title itself has company: "CEO at X" type
    m = re.search(r'(?:CEO|Director|Manager|Head|President|Lead|Owner)\s+(?:at|of|@)\s+([A-Z][A-Za-z0-9&.\s]{2,40}?)(?:[,]|\s*$)', job)
    if m:
        candidates.append(m.group(1).strip())
    
    # Score candidates - prefer shortest that looks like a real company
    best = ""
    for c in candidates:
        # Must be at least 3 chars, no more than 50
        if len(c) < 3 or len(c) > 50:
            continue
        # Must start with capital letter or number
        if not re.match(r'[A-Z0-9]', c):
            continue
        # Prefer shorter = cleaner
        if not best or len(c) < len(best):
            best = c
    
    if best:
        return best
    
    # Fallback: use raw but heavily cleaned
    return cleaned if cleaned else "Unknown"


# ──────────────────────────
# Web Search Functions
# ──────────────────────────

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
]


def web_search_raw(query: str, max_results: int = 5) -> List[dict]:
    """Search using hermes_tools.web_search"""
    try:
        from hermes_tools import web_search
        r = web_search(query=query, limit=max_results)
        if isinstance(r, dict) and 'data' in r:
            return r['data'].get('web', [])
    except Exception as e:
        log.warning(f"  ✗ Search error: {e}")
    return []


def find_company_website(company: str) -> Optional[str]:
    """Find company website."""
    results = web_search_raw(f'"{company}" official website', max_results=5)
    for r in results:
        url = r.get('url', '')
        title = r.get('title', '').lower()
        # Skip social media, Wikipedia, job sites
        skip = ['linkedin.com', 'facebook.com', 'twitter.com', 'instagram.com',
                'wikipedia.org', 'indeed.com', 'glassdoor.com', 'crunchbase.com']
        if any(s in url.lower() for s in skip):
            continue
        # Looks like a real company site
        if company.lower()[:4] in url.lower() or company.lower()[:4] in title:
            return url
    # More generic fallback
    for r in results:
        url = r.get('url', '')
        skip = ['linkedin.com', 'facebook.com', 'twitter.com']
        if not any(s in url.lower() for s in skip):
            return url
    return None


def find_ceo_or_decision_maker(company: str, known_names: List[str]) -> Dict:
    """Search for CEO/decision maker info using web search."""
    result = {
        'ceo_name': '',
        'ceo_email': '',
        'decision_maker': '',
        'decision_maker_title': '',
        'decision_maker_email': '',
        'company_phone': '',
        'contact_email': '',
    }
    
    # Search 1: CEO of company
    queries = [
        f'"{company}" CEO contact email',
        f'"{company}" managing director "{company}"',
        f'"{company}" "{company}" email address',
    ]
    
    for q in queries[:2]:
        results = web_search_raw(q, max_results=5)
        for r in results:
            title = r.get('title', '')
            snippet = r.get('description', '')
            text = f"{title} {snippet}"
            
            # Extract CEO name
            m = re.search(r'(?:CEO|Chief Executive|Managing Director|President)[:\s]+([A-Z][a-z]+(?:\s[A-Z][a-z]+){1,2})', text)
            if m and not result['ceo_name']:
                result['ceo_name'] = m.group(1).strip()
            
            # Extract email
            m = re.search(r'([\w.+-]+@[\w-]+\.[\w.-]+)', text)
            if m and not result['contact_email']:
                email = m.group(1)
                # Skip personal emails
                if 'gmail' not in email and 'yahoo' not in email:
                    result['contact_email'] = email
            
            # Extract phone
            m = re.search(r'(\+?\d{1,3}[-\s]?\(?\d{1,4}\)?[-\s]?\d{1,4}[-\s]?\d{1,4}[-\s]?\d{1,4})', text)
            if m and not result['company_phone']:
                result['company_phone'] = m.group(1).strip()
        
        time.sleep(1.5)
    
    # Search 2: Try to find known person's email
    for known in known_names[:2]:
        if known:
            q = f'"{known}" "{company}" email contact'
            results = web_search_raw(q, max_results=3)
            for r in results:
                snippet = r.get('description', '')
                m = re.search(r'([\w.+-]+@[\w-]+\.[\w.-]+)', snippet)
                if m:
                    email = m.group(1)
                    if 'gmail' not in email and 'yahoo' not in email:
                        if not result['decision_maker_email']:
                            result['decision_maker'] = known
                            result['decision_maker_email'] = email
                        break
            time.sleep(1.5)
    
    # Search 3: Generic company contact
    q = f'"{company}" contact email info phone'
    results = web_search_raw(q, max_results=5)
    for r in results:
        snippet = r.get('description', '')
        if not result['contact_email']:
            m = re.search(r'([\w.+-]+@[\w-]+\.[\w.-]+)', snippet)
            if m:
                email = m.group(1)
                if 'gmail' not in email and 'yahoo' not in email:
                    result['contact_email'] = email
        if not result['company_phone']:
            m = re.search(r'(\+?\d{1,3}[-\s]?\(?\d{1,4}\)?[-\s]?\d{1,4}[-\s]?\d{1,4}[-\s]?\d{1,4})', snippet)
            if m:
                result['company_phone'] = m.group(1).strip()
    
    return result


def try_website_scrape(website: str) -> Dict:
    """Try to extract contact info from company website."""
    result = {'email': '', 'phone': ''}
    if not website:
        return result
    
    try:
        req = urllib.request.Request(website, headers={
            'User-Agent': USER_AGENTS[0], 'Accept': 'text/html'
        })
        with urllib.request.urlopen(req, timeout=10) as resp:
            html = resp.read().decode('utf-8', errors='replace')[:50000]
        
        # Look for emails
        emails = re.findall(r'[\w.+-]+@[\w-]+\.[\w.-]+', html)
        for e in emails:
            if 'gmail' not in e and 'yahoo' not in e:
                if not result['email']:
                    result['email'] = e
        
        # Look for phone
        phones = re.findall(r'(\+?27|+?234|+?254|+?233|+?1|+?44)[-\s]?\(?\d{1,4}\)?[-\s]?\d{1,4}[-\s]?\d{1,4}[-\s]?\d{1,4}', html)
        if phones:
            result['phone'] = phones[0]
        
        # Also try contact page
        for sub in ['contact', 'contact-us', 'about', 'about-us']:
            try:
                contact_url = f"{website.rstrip('/')}/{sub}"
                req2 = urllib.request.Request(contact_url, headers={
                    'User-Agent': USER_AGENTS[1], 'Accept': 'text/html'
                })
                with urllib.request.urlopen(req2, timeout=8) as resp2:
                    html2 = resp2.read().decode('utf-8', errors='replace')[:30000]
                
                emails2 = re.findall(r'[\w.+-]+@[\w-]+\.[\w.-]+', html2)
                for e in emails2:
                    if 'gmail' not in e and 'yahoo' not in e:
                        if not result['email']:
                            result['email'] = e
            except:
                pass
    
    except Exception as e:
        log.debug(f"  Website scrape failed: {e}")
    
    return result


# ──────────────────────────
# Main Pipeline
# ──────────────────────────

def load_leads():
    """Load leads from Excel."""
    import openpyxl
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    
    leads = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row[1]:
            continue
        name = str(row[1] or '').strip()
        job = str(row[2] or '').strip()
        company_raw = str(row[3] or '').strip()
        country = str(row[4] or '').strip()
        desc = str(row[7] or '').strip() if len(row) > 7 else ''
        url = str(row[5] or '').strip() if len(row) > 5 else ''
        
        if not name:
            continue
        
        lead = EnrichedLead(
            full_name=name,
            job_title=job,
            company_raw=company_raw,
            country=country,
            linkedin_url=url,
            description=desc,
        )
        
        # Clean company name
        lead.company_clean = extract_company_from_parts(name, job, company_raw, desc, url)
        leads.append(lead)
    
    log.info(f"Loaded {len(leads)} leads from {INPUT_FILE}")
    return leads


def enrich_lead(lead: EnrichedLead) -> EnrichedLead:
    """Full enrichment pipeline for one lead."""
    company = lead.company_clean
    
    # Step 1: Find company website
    log.info(f"  Searching website for {company}...")
    website = find_company_website(company)
    if website:
        lead.company_website = website
    
    # Step 2: Try to scrape website for contact info
    if website:
        log.info(f"  Scraping {website}...")
        scrape = try_website_scrape(website)
        if scrape['email']:
            lead.company_email = scrape['email']
        if scrape['phone']:
            lead.company_phone = scrape['phone']
    
    # Step 3: Search for CEO/decision maker
    log.info(f"  Searching for contacts at {company}...")
    known_names = [lead.full_name]
    contacts = find_ceo_or_decision_maker(company, known_names)
    
    if contacts['ceo_name']:
        lead.ceo_name = contacts['ceo_name']
    if contacts['ceo_email']:
        lead.ceo_email = contacts['ceo_email']
    if contacts['contact_email'] and not lead.company_email:
        lead.company_email = contacts['contact_email']
    if contacts['company_phone']:
        lead.company_phone = contacts['company_phone']
    if contacts['decision_maker']:
        lead.decision_maker_name = contacts['decision_maker']
    if contacts['decision_maker_email']:
        lead.decision_maker_email = contacts['decision_maker_email']
    if contacts['decision_maker_title']:
        lead.decision_maker_title = contacts['decision_maker_title']
    
    # Record queries used
    queries_used = [
        f'"{company}" CEO contact email',
        f'"{company}" contact email info',
    ]
    if known_names[0]:
        queries_used.append(f'"{known_names[0]}" "{company}" email')
    lead.search_queries_used = ' | '.join(queries_used[:2])
    
    lead.enrichment_status = "completed"
    return lead


def export_to_excel(leads: List[EnrichedLead], output_path: str):
    """Export enriched leads to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Enriched LinkedIn Leads"
        
        headers = [
            "Full Name", "Job Title", "Company (Cleaned)", "Company (Raw)",
            "Country", "LinkedIn URL",
            "Company Website", "Company Phone", "Company Email",
            "CEO Name", "CEO Email", "Decision Maker", "Decision Maker Title",
            "Decision Maker Email", "Description"
        ]
        
        hfont = Font(bold=True, color="FFFFFF", size=10)
        hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        thin = Side(style='thin', color='B4C6E7')
        border = Border(bottom=thin)
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        email_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        for row_idx, lead in enumerate(leads, 2):
            vals = [
                lead.full_name, lead.job_title,
                lead.company_clean, lead.company_raw,
                lead.country, lead.linkedin_url,
                lead.company_website, lead.company_phone or '', lead.company_email or '',
                lead.ceo_name or '', lead.ceo_email or '',
                lead.decision_maker_name or '', lead.decision_maker_title or '',
                lead.decision_maker_email or '', lead.description[:200] if lead.description else ''
            ]
            
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                # Highlight rows that have email
                has_email = bool(lead.company_email or lead.ceo_email or lead.decision_maker_email)
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
                
                if col == 6:  # URL column
                    cell.font = Font(color="0563C1", underline="single", size=9)
                elif col in (7, 8, 9) and val:  # Contact info highlight
                    cell.fill = email_fill
        
        # Column widths
        widths = [22, 30, 25, 25, 15, 45, 30, 18, 30, 22, 30, 22, 25, 30, 50]
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        
        wb.save(output_path)
        log.info(f"✅ Saved {len(leads)} enriched leads to {output_path}")
        return True
        
    except ImportError as e:
        log.warning(f"openpyxl not available: {e}")
        return False


def main():
    log.info("🚀 LinkedIn Lead Enricher Pipeline")
    log.info(f"   Input: {INPUT_FILE}")
    log.info("")
    
    # Load leads
    leads = load_leads()
    
    # Deduplicate by company (prioritize decision makers)
    # Actually, keep all leads individually - each person is a separate entry
    log.info(f"\n{'='*60}")
    log.info(f"Starting enrichment for {len(leads)} leads...")
    log.info(f"{'='*60}")
    
    enriched = []
    for i, lead in enumerate(leads, 1):
        log.info(f"[{i}/{len(leads)}] {lead.full_name} @ {lead.company_clean}")
        try:
            result = enrich_lead(lead)
            enriched.append(result)
        except Exception as e:
            log.warning(f"  ✗ Enrichment failed: {e}")
            lead.enrichment_status = "failed"
            enriched.append(lead)
        
        # Rate limiting
        if i < len(leads):
            time.sleep(0.5)
    
    # Summary stats
    with_email = sum(1 for l in enriched if l.company_email or l.ceo_email or l.decision_maker_email)
    with_website = sum(1 for l in enriched if l.company_website)
    with_phone = sum(1 for l in enriched if l.company_phone)
    with_ceo = sum(1 for l in enriched if l.ceo_name)
    
    log.info(f"\n{'='*60}")
    log.info(f"📊 Enrichment Summary")
    log.info(f"{'='*60}")
    log.info(f"   Total leads:        {len(enriched)}")
    log.info(f"   With website:       {with_website}")
    log.info(f"   With email:         {with_email}")
    log.info(f"   With phone:         {with_phone}")
    log.info(f"   With CEO found:     {with_ceo}")
    
    # Export
    ts = datetime.now().strftime("%m%d_%H%M")
    output_path = OUTPUT_DIR / f"enriched_leads_{ts}.xlsx"
    
    if export_to_excel(enriched, str(output_path)):
        log.info(f"\n📁 Output: {output_path}")
    else:
        # CSV fallback
        csv_path = OUTPUT_DIR / f"enriched_leads_{ts}.csv"
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f)
            w.writerow(["Name","Job","Company Clean","Company Raw","Country",
                       "LinkedIn","Website","Phone","Email",
                       "CEO","CEO Email","Decision Maker","DM Email"])
            for l in enriched:
                w.writerow([l.full_name, l.job_title, l.company_clean, l.company_raw,
                          l.country, l.linkedin_url, l.company_website,
                          l.company_phone, l.company_email,
                          l.ceo_name, l.ceo_email,
                          l.decision_maker_name, l.decision_maker_email])
        log.info(f"\n📁 CSV Output: {csv_path}")


if __name__ == "__main__":
    main()