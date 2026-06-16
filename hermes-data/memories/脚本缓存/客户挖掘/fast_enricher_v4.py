#!/usr/bin/env python3
"""
Fast Enricher v4 — 只用 web_search（已验证可用），每公司1个复合query
核心优化：只搜1次/公司，同时搜官网+邮箱+CEO，减少等待
缓存已找到的，避免重复

用法: python fast_enricher_v4.py
"""

import openpyxl, re, json, sqlite3, os, time
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT = r"C:\Users\Admin\Desktop\Working\linkedin_xray_0604_1517.xlsx"
OUTPUT_DIR = Path(r"C:\Users\Admin\Desktop\Working")

# Previously found data (no need to re-search)
CACHED = {
    "Omniflex":               {"website":"https://www.omniflex.com", "email":"sales@omniflex.com", "phone":"", "ceo":""},
    "AUTOMATED SYSTEM WORKS":{"website":"https://www.aswgroup.co.za","email":"info@aswgroup.co.za","phone":"","ceo":""},
    "Chilltemp Automation Technologies":{"website":"http://chilltemp.co.za","email":"jaco@chilltemp.co.za","phone":"+27 63 778 0442","ceo":"Wilco Lubbe"},
    "Rand Water":             {"website":"https://www.randwater.co.za","email":"","phone":"","ceo":""},
    "Schneider Electric":     {"website":"https://www.se.com","email":"","phone":"","ceo":""},
    "Rockwell Automation":    {"website":"https://www.rockwellautomation.com","email":"","phone":"","ceo":""},
    "Anglo American":         {"website":"https://www.angloamerican.com","email":"","phone":"","ceo":""},
    "Chevron":                {"website":"https://www.chevron.com","email":"","phone":"","ceo":""},
    "Alstom":                 {"website":"https://www.alstom.com","email":"","phone":"","ceo":""},
    "KHS":                    {"website":"https://www.khs.com","email":"","phone":"","ceo":""},
    "Jendamark Automation":   {"website":"https://jendamark.io","email":"","phone":"","ceo":""},
    "A2Z Automation":         {"website":"https://a2zautomations.com","email":"","phone":"","ceo":""},
}

# Companies that need web search
NEED_SEARCH = [
    "Liquid Automation Systems",
    "Letek Industrial Automation",
    "UltraSec Fencing and Automation",
    "Aveng Manufacturing",
    "Maximize Systems",
    "Honour World",
    "LynxTrace",
    "INTECH Automation",
    "Quad Automation",
]

def search_web(query):
    """Use hermes_tools.web_search."""
    from hermes_tools import web_search
    r = web_search(query=query, limit=5)
    if isinstance(r, dict) and 'data' in r:
        return r['data'].get('web', [])
    return []

def extract(text):
    """Extract website, email, phone, CEO from text."""
    we = ""; em = ""; ph = ""; ce = ""
    m = re.search(r'https?://(?:www\.)?([A-Za-z0-9][A-Za-z0-9.-]+\.[A-Za-z]{2,})(?:/[^\s"\'<>)]*)?', text)
    if m:
        u = m.group(0).rstrip(')}')
        skip = ['linkedin','facebook','twitter','instagram','wikipedia','youtube','github','pinterest']
        if not any(s in u.lower() for s in skip):
            we = u
    es = re.findall(r'([\w.+-]+@(?:[^.]+\.)+(?:com|co\.za|co\.ke|co\.ng|org|net|io))', text)
    for e in es:
        if not any(p in e.lower() for p in ['gmail','yahoo','hotmail','outlook','icloud']):
            em = e; break
    ps = re.findall(r'(\+27[\s-]?\d[\d\s-]{6,12}|\+234[\s-]?\d[\d\s-]{6,12}|\+254[\s-]?\d[\d\s-]{6,12})', text)
    if ps: ph = ps[0]
    cm = re.search(r'(?:CEO|Managing Director|Founder|President)\s*[:\-–\s]{1,5}\s*([A-Z][a-z]+(?:\s[A-Z][a-z]+){1,2})', text)
    if cm: ce = cm.group(1).strip()
    return we, em, ph, ce

# Merge all data
ALL = dict(CACHED)
for co in NEED_SEARCH:
    ALL[co] = {"website":"", "email":"", "phone":"", "ceo":""}

# Load people from original file
wb = openpyxl.load_workbook(INPUT)
ws = wb.active

GOOD_NAMES = {x.lower(): x for x in ALL}

def get_company(name, job, raw, desc, url):
    text = f"{desc} {job}".lower()
    for k, v in GOOD_NAMES.items():
        if k in text: return v
    m = re.search(r'Experience[:\s]+([A-Z][A-Za-z0-9&.]+(?:\s[A-Z][A-Za-z0-9&.]+){0,3})', desc)
    if m:
        c = m.group(1).strip()
        if c.lower() in GOOD_NAMES: return GOOD_NAMES[c.lower()]
    c = re.sub(r'\s*[|]\s*LinkedIn.*', '', raw).strip()
    c = re.sub(r'\s*\.{3,}\s*$', '', c).strip()
    if c.lower() in GOOD_NAMES: return GOOD_NAMES[c.lower()]
    return ""

# Search remaining companies in parallel via execute_code
print("Companies with data already: ", sum(1 for v in ALL.values() if v['email'] or v['website']))
print("Companies needing search: ", [co for co in ALL if not ALL[co]['website']])
print()
print("Launching searches...")
print()

# One search per remaining company
remaining = [co for co in ALL if not ALL[co]['website'] and not ALL[co]['email']]
for co in remaining:
    print(f"  Searching {co}...", end=" ", flush=True)
    try:
        results = search_web(f'"{co}" official website email contact')
        text = json.dumps(results)
        we, em, ph, ce = extract(text)
        ALL[co]['website'] = we
        ALL[co]['email'] = em
        ALL[co]['phone'] = ph
        ALL[co]['ceo'] = ce
        status = "✅" if em else ("🌐" if we else "🔍")
        print(f"{status} email={em or '—'}")
    except Exception as e:
        print(f"✗ {e}")
    time.sleep(1.2)

print()

# Build Excel
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Enriched Leads"

headers = ["Company","Contact Name","Job Title","Country",
           "Company Website","Company Email","Company Phone",
           "LinkedIn URL","Status"]

hfont = Font(bold=True, color="FFFFFF", size=10)
hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
alt1 = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
thin = Border(bottom=Side(style='thin', color='00CCCCCC'))

for c, h in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = hfont; cell.fill = hfill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

by_co = defaultdict(list)
row = 2
stats = {'total': 0, 'email': 0, 'website': 0}

for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[1]: continue
    name = str(r[1] or '').strip()
    job = str(r[2] or '').strip()
    raw = str(r[3] or '').strip()
    country = str(r[4] or '').strip()
    url = str(r[5] or '').strip() if len(r)>5 else ''
    desc = str(r[7] or '').strip() if len(r)>7 else ''
    
    co = get_company(name, job, raw, desc, url)
    ec = ALL.get(co, {})
    
    has_email = bool(ec.get('email'))
    status = "✅ Email" if has_email else ("🌐 Website" if ec.get('website') else "🔍 Manual")
    
    vals = [co, name, job, country,
            ec.get('website','') or '', ec.get('email','') or '', ec.get('phone','') or '',
            url, status]
    
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=row, column=c, value=v)
        cell.border = thin; cell.alignment = Alignment(wrap_text=True, vertical="top")
        if row % 2 == 0: cell.fill = alt1
        if c == 6 and v: cell.fill = green
        if c == 5 and v: cell.font = Font(color="0563C1", underline="single", size=9)
        if c == 7 and v: cell.fill = green
    
    stats['total'] += 1
    if has_email: stats['email'] += 1
    if ec.get('website'): stats['website'] += 1
    by_co[co].append(name)
    row += 1

for i, w in enumerate([28,24,35,15,32,32,22,50,15], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f"A1:I{row-1}"

# Summary sheet
ws3 = wb2.create_sheet("Company Summary")
for c, h in enumerate(["Company","Website","Email","Phone","CEO","Status","Contacts"], 1):
    cell = ws3.cell(row=1, column=c, value=h)
    cell.font = hfont; cell.fill = hfill

sr = 2
for co in sorted(ALL):
    ec = ALL[co]
    st = "✅" if ec['email'] else ("🌐" if ec['website'] else "🔍")
    for c, v in enumerate([co, ec['website'], ec['email'], ec['phone'], ec['ceo'], st, len(by_co.get(co,[]))], 1):
        ws3.cell(row=sr, column=c, value=v)
        if c == 3 and v: ws3.cell(row=sr, column=c).fill = green
    sr += 1

for i, w in enumerate([35,35,35,22,22,10,10], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ts = datetime.now().strftime("%m%d_%H%M")
out = OUTPUT_DIR / f"enriched_leads_{ts}.xlsx"
wb2.save(str(out))

print(f"\n{'='*60}")
print(f"📊 FINAL REPORT")
print(f"{'='*60}")
print(f"   Total people:      {stats['total']}")
print(f"   Companies:         {len(ALL)}")
print(f"   ✅ With email:     {stats['email']}")
print(f"   🌐 With website:   {stats['website']}")
print(f"\n📁 Output: {out}")
print()
for co in sorted(ALL):
    ec = ALL[co]
    if ec['email']:
        print(f"   ✅ {co:35s} → {ec['email']}")
    elif ec['website']:
        print(f"   🌐 {co:35s} → {ec['website'][:40]}")
    else:
        print(f"   🔍 {co:35s} → no contact info")