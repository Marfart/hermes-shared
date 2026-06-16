#!/usr/bin/env python3
import openpyxl, re
from datetime import datetime
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT = r"C:\Users\Admin\Desktop\Working\linkedin_xray_0604_1517.xlsx"
OUTPUT = Path(r"C:\Users\Admin\Desktop\Working")

enrich_data = {
    "Omniflex": {"website": "https://www.omniflex.com", "email": "sales@omniflex.com", "phone": "", "ceo": ""},
    "Rand Water": {"website": "https://www.randwater.co.za", "email": "", "phone": "", "ceo": ""},
    "Schneider Electric": {"website": "https://www.se.com", "email": "", "phone": "", "ceo": ""},
    "AUTOMATED SYSTEM WORKS": {"website": "https://www.aswgroup.co.za", "email": "info@aswgroup.co.za", "phone": "", "ceo": ""},
    "Chilltemp Automation Technologies": {"website": "http://chilltemp.co.za", "email": "jaco@chilltemp.co.za", "phone": "+27 63 778 0442", "ceo": "Wilco Lubbe"},
    "Jendamark Automation": {"website": "https://jendamark.io", "email": "", "phone": "", "ceo": ""},
    "A2Z Automation": {"website": "https://a2zautomations.com", "email": "", "phone": "", "ceo": ""},
    "Anglo American": {"website": "https://www.angloamerican.com", "email": "", "phone": "", "ceo": ""},
    "Quad Automation": {"website": "", "email": "", "phone": "", "ceo": ""},
    "Alstom": {"website": "https://www.alstom.com", "email": "", "phone": "", "ceo": ""},
    "Rockwell Automation": {"website": "https://www.rockwellautomation.com", "email": "", "phone": "", "ceo": ""},
    "Chevron": {"website": "https://www.chevron.com", "email": "", "phone": "", "ceo": ""},
    "KHS": {"website": "https://www.khs.com", "email": "", "phone": "", "ceo": ""},
    "Liquid Automation Systems": {"website": "", "email": "", "phone": "", "ceo": ""},
    "Letek Industrial Automation": {"website": "", "email": "", "phone": "", "ceo": ""},
    "UltraSec Fencing and Automation": {"website": "", "email": "", "phone": "", "ceo": ""},
    "Aveng Manufacturing": {"website": "", "email": "", "phone": "", "ceo": ""},
    "Maximize Systems": {"website": "", "email": "", "phone": "", "ceo": ""},
    "Honour World": {"website": "", "email": "", "phone": "", "ceo": ""},
    "LynxTrace": {"website": "", "email": "", "phone": "", "ceo": ""},
    "INTECH Automation": {"website": "", "email": "", "phone": "", "ceo": ""},
}

def get_company(name, job, raw, desc, url):
    text = f"{desc} {job}".lower()
    for co in enrich_data:
        if co.lower() in text:
            return co
    m = re.search(r'Experience[:\s]+([A-Z][A-Za-z0-9&.]+(?:\s[A-Z][A-Za-z0-9&.]+){0,3})', desc)
    if m:
        c = m.group(1).strip()
        for co in enrich_data:
            if co.lower() in c.lower():
                return co
    c = re.sub(r'\s*[|]\s*LinkedIn.*', '', raw).strip()
    c = re.sub(r'\s*LinkedIn.*', '', c).strip()
    c = re.sub(r'\s*\.{3,}\s*$', '', c).strip()
    if len(c) > 2 and c not in ('Automation','Vice President','Remote','Energy','','-','None','Junior Electrical Control','Ease','Dongf'):
        return c
    return ""

wb = openpyxl.load_workbook(INPUT)
ws = wb.active

wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Enriched Leads"

headers = ["Company","Contact Name","Job Title","Country",
           "Company Website","Company Email","Company Phone",
           "LinkedIn URL","Status"]

hfont = Font(bold=True, color="FFFFFF", size=10)
hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
alt1 = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
thin = Border(bottom=Side(style='thin', color='00CCCCCC'))

for c, h in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = hfont; cell.fill = hfill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

row = 2
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[1]: continue
    name = str(r[1] or '').strip()
    job = str(r[2] or '').strip()
    raw = str(r[3] or '').strip()
    country = str(r[4] or '').strip()
    url = str(r[5] or '').strip() if len(r)>5 else ''
    desc = str(r[7] or '').strip() if len(r)>7 else ''
    
    co = get_company(name, job, raw, desc, url)
    ec = enrich_data.get(co, {})
    has_email = bool(ec.get('email'))
    status = "✅ Email found" if has_email else ("🌐 Website found" if ec.get('website') else "🔍 Manual research")
    
    vals = [co, name, job, country,
            ec.get('website','') or '', ec.get('email','') or '', ec.get('phone','') or '',
            url, status]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=row, column=c, value=v)
        cell.border = thin
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if row % 2 == 0: cell.fill = alt1
        if c == 6 and v: cell.fill = green
        if c == 5 and v: cell.font = Font(color="0563C1", underline="single", size=9)
        if c == 7 and v: cell.fill = green
    row += 1

for i, w in enumerate([28,24,35,15,32,35,25,50,22], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f"A1:I{row-1}"

# Sheet 2: Company summary
ws3 = wb2.create_sheet("Company Summary")
for c, h in enumerate(["Company","Website","Email","Phone","CEO","Status"], 1):
    cell = ws3.cell(row=1, column=c, value=h)
    cell.font = hfont; cell.fill = hfill
sr = 2
for co, ec in sorted(enrich_data.items()):
    s = "✅ Email" if ec['email'] else ("🌐 Website" if ec['website'] else "🔍 Manual")
    vals = [co, ec['website'], ec['email'], ec['phone'], ec['ceo'], s]
    for c, v in enumerate(vals, 1):
        ws3.cell(row=sr, column=c, value=v)
        if c == 3 and v: ws3.cell(row=sr, column=c).fill = green
    sr += 1
ws3.column_dimensions['A'].width = 35; ws3.column_dimensions['B'].width = 35
ws3.column_dimensions['C'].width = 35; ws3.column_dimensions['D'].width = 25
ws3.column_dimensions['E'].width = 20; ws3.column_dimensions['F'].width = 18

ts = datetime.now().strftime("%m%d_%H%M")
out = OUTPUT / f"enriched_leads_{ts}.xlsx"
wb2.save(str(out))

total = row - 2
we = sum(1 for r in ws2.iter_rows(min_row=2, max_row=row-1, values_only=True) if r[5])
ww = sum(1 for r in ws2.iter_rows(min_row=2, max_row=row-1, values_only=True) if r[4])

print(f"✅ Done! | {total} people, {len(enrich_data)} companies")
print(f"   Email found: {we} | Website found: {ww}")
print(f"📁 {out}")
print()
for co, ec in sorted(enrich_data.items()):
    if ec.get('email'):
        print(f"   ✅ {co:35s} → {ec['email']}")
for co, ec in sorted(enrich_data.items()):
    if ec.get('website') and not ec.get('email'):
        print(f"   🌐 {co:35s} → {ec['website'][:40]}")
for co in sorted(enrich_data):
    ec = enrich_data[co]
    if not ec.get('website') and not ec.get('email'):
        print(f"   🔍 {co:35s} → no contact info found")