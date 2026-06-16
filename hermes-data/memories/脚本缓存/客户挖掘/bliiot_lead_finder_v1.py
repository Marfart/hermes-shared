#!/usr/bin/env python
"""BLIIoT High-Potential Customer Finder - Google Maps + Web Search Edition
Route A - 完全安全，零封号风险
Target: 工业自动化系统集成商、IoT公司、PLC工程公司
Priority: 有WhatsApp的客户优先收集
"""
from hermes_tools import web_search, web_extract, terminal, write_file, read_file
import json, re, time, os

class B2BLeadFinder:
    def __init__(self):
        self.leads = []
        self.processed_urls = set()
    
    def search_companies(self, query, limit=10):
        """Search for companies via web search"""
        print(f"\n🔍 Searching: {query}")
        result = web_search(query=query, limit=limit)
        if result and 'data' in result and 'web' in result['data']:
            return result['data']['web']
        return []
    
    def extract_emails_from_text(self, text):
        """Extract email addresses from text"""
        pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
        return list(set(re.findall(pattern, text)))
    
    def extract_phones_from_text(self, text):
        """Extract phone numbers from text (international format)"""
        # +XX XXXX XXXX or similar formats
        pattern = r'(?:\+?\d{1,3}[\s-]?)?\(?\d{1,4}\)?[\s-]?\d{1,4}[\s-]?\d{1,4}[\s-]?\d{1,4}'
        phones = re.findall(pattern, text)
        # Filter to reasonable phone numbers
        valid = []
        for p in phones:
            digits = re.sub(r'\D', '', p)
            if 7 <= len(digits) <= 15:
                valid.append(p.strip())
        return valid
    
    def check_whatsapp(self, text):
        """Check if text mentions WhatsApp"""
        whatsapp_indicators = [
            r'whatsapp', r'wa\.me', r'wa/(\d+)', r'whatsapp\.com',
            r'chat\.whatsapp', r'api\.whatsapp', r'wa\.link'
        ]
        for pattern in whatsapp_indicators:
            if re.search(pattern, text, re.IGNORECASE):
                return True
        return False
    
    def analyze_company_for_bliiot(self, name, description, website):
        """Score how well a company matches BLIIoT products"""
        desc_lower = (description + ' ' + name).lower()
        
        # BLIIoT product match keywords
        matches = {
            'PLC/SCADA': ['plc', 'scada', 'programmable logic', 'automation control'],
            'IoT/IIoT': ['iot', 'iiot', 'internet of things', 'industrial iot', 'mqtt'],
            'Gateway/Edge': ['gateway', 'edge computing', 'edge gateway', 'protocol conversion'],
            'Remote Monitoring': ['remote monitoring', 'telemetry', 'rtu', 'remote terminal'],
            'Energy/Power': ['energy', 'power', 'substation', 'smart grid', 'solar', 'renewable'],
            'Building Auto': ['building automation', 'hvac', 'bas', 'bms', 'bacnet'],
            'System Integrator': ['system integrator', 'integration', 'automation solutions'],
            'Industrial Network': ['industrial ethernet', 'modbus', 'opc ua', 'profibus', 'profinet'],
        }
        
        found_categories = []
        for category, keywords in matches.items():
            for kw in keywords:
                if kw in desc_lower:
                    found_categories.append(category)
                    break
        
        # Score: 0-10
        score = len(found_categories) * 1.5
        score = min(score, 10)
        
        return {
            'score': round(score, 1),
            'categories': found_categories,
            'is_high_potential': score >= 4
        }
    
    def crawl_website_for_contacts(self, url):
        """Crawl a company website to find contact info"""
        if not url or url in self.processed_urls:
            return {}
        self.processed_urls.add(url)
        
        print(f"  🌐 Crawling: {url}")
        contacts = {'emails': [], 'phones': [], 'has_whatsapp': False, 'whatsapp_numbers': []}
        
        try:
            result = web_extract(urls=[url])
            if result and 'results' in result:
                for r in result['results']:
                    content = r.get('content', '') or ''
                    # Find emails
                    emails = self.extract_emails_from_text(content)
                    contacts['emails'].extend(emails)
                    
                    # Find phones
                    phones = self.extract_phones_from_text(content)
                    contacts['phones'].extend(phones)
                    
                    # Check WhatsApp
                    if self.check_whatsapp(content):
                        contacts['has_whatsapp'] = True
                        wa_patterns = [
                            r'(?:whatsapp[:\s]*?)(\+\d{1,3}[\d\s\-]{7,15})',
                            r'wa\.me/(\d+)',
                            r'wa/(\d+)',
                        ]
                        for pat in wa_patterns:
                            wa_matches = re.findall(pat, content, re.IGNORECASE)
                            contacts['whatsapp_numbers'].extend(wa_matches)
        except Exception as e:
            print(f"  ⚠️ Error crawling: {e}")
        
        # Also try Contact Us page
        contact_urls = [
            url.rstrip('/') + '/contact',
            url.rstrip('/') + '/contact-us',
            url.rstrip('/') + '/about',
            url.rstrip('/') + '/about-us',
        ]
        for cu in contact_urls:
            try:
                result2 = web_extract(urls=[cu])
                if result2 and 'results' in result2:
                    for r in result2['results']:
                        content = r.get('content', '') or ''
                        emails = self.extract_emails_from_text(content)
                        contacts['emails'].extend(emails)
                        phones = self.extract_phones_from_text(content)
                        contacts['phones'].extend(phones)
                        if self.check_whatsapp(content):
                            contacts['has_whatsapp'] = True
            except:
                pass
        
        # Deduplicate
        contacts['emails'] = list(set(contacts['emails']))
        contacts['phones'] = list(set(contacts['phones']))
        contacts['whatsapp_numbers'] = list(set(contacts['whatsapp_numbers']))
        
        return contacts
    
    def add_lead(self, name, website, phone, description, source_query):
        """Add a lead to the collection"""
        lead = {
            'company': name,
            'website': website or '',
            'phone': phone or '',
            'description': description[:200] if description else '',
            'source': source_query,
        }
        
        # Analyze match
        analysis = self.analyze_company_for_bliiot(name, description or '', website or '')
        lead['match_score'] = analysis['score']
        lead['match_categories'] = ', '.join(analysis['categories'])
        lead['high_potential'] = '✅ YES' if analysis['is_high_potential'] else '❌'
        
        # Crawl website
        if website:
            contacts = self.crawl_website_for_contacts(website)
            lead['emails'] = ', '.join(contacts['emails'][:5]) if contacts['emails'] else ''
            lead['phones_found'] = ', '.join(contacts['phones'][:3]) if contacts['phones'] else ''
            lead['has_whatsapp'] = '✅' if contacts['has_whatsapp'] else ''
            lead['whatsapp_numbers'] = ', '.join(contacts['whatsapp_numbers'][:3]) if contacts['whatsapp_numbers'] else ''
        else:
            lead['emails'] = ''
            lead['phones_found'] = ''
            lead['has_whatsapp'] = ''
            lead['whatsapp_numbers'] = ''
        
        self.leads.append(lead)
        print(f"  {'✅' if analysis['is_high_potential'] else '  '} {name} (Score: {analysis['score']})")
        
        # Show WhatsApp priority
        if lead['has_whatsapp']:
            print(f"    📱 HAS WHATSAPP! {lead['whatsapp_numbers']}")
        if lead['emails']:
            print(f"    📧 {lead['emails']}")
        
        time.sleep(0.5)  # Be polite
    
    def output_excel(self, filename):
        """Output collected leads to Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "BLIIoT High-Potential Leads"
        
        # Headers
        headers = [
            'Company', 'Website', 'Phone (Google Maps)', 'Email(s) Found', 
            'Phone(s) Found', 'Has WhatsApp', 'WhatsApp Number(s)',
            'Match Score', 'Matched Categories', 'High Potential', 
            'Description', 'Source'
        ]
        
        # Style
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        high_potential_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Sort: WhatsApp first, then high score
        sorted_leads = sorted(self.leads, key=lambda x: (
            -int(x['has_whatsapp'] == '✅'),
            -(x['match_score'] if x['high_potential'] == '✅ YES' else 0)
        ))
        
        for row, lead in enumerate(sorted_leads, 2):
            ws.cell(row=row, column=1, value=lead['company'])
            ws.cell(row=row, column=2, value=lead['website'])
            ws.cell(row=row, column=3, value=lead['phone'])
            ws.cell(row=row, column=4, value=lead['emails'])
            ws.cell(row=row, column=5, value=lead['phones_found'])
            ws.cell(row=row, column=6, value=lead['has_whatsapp'])
            ws.cell(row=row, column=7, value=lead['whatsapp_numbers'])
            ws.cell(row=row, column=8, value=lead['match_score'])
            ws.cell(row=row, column=9, value=lead['match_categories'])
            ws.cell(row=row, column=10, value=lead['high_potential'])
            ws.cell(row=row, column=11, value=lead['description'][:100])
            ws.cell(row=row, column=12, value=lead['source'])
            
            # Highlight high potential
            if lead['high_potential'] == '✅ YES':
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = high_potential_fill
        
        # Auto-width
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)
        
        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2.cell(row=1, column=1, value=f"Total Leads Found: {len(self.leads)}")
        ws2.cell(row=2, column=1, value=f"High Potential: {sum(1 for l in self.leads if l['high_potential'] == '✅ YES')}")
        ws2.cell(row=3, column=1, value=f"With WhatsApp: {sum(1 for l in self.leads if l['has_whatsapp'] == '✅')}")
        ws2.cell(row=4, column=1, value=f"With Email: {sum(1 for l in self.leads if l['emails'])}")
        
        wb.save(filename)
        print(f"\n📊 Excel saved: {filename}")
        return filename


def main():
    finder = B2BLeadFinder()
    
    # Phase 1: Search for companies by keyword + region
    print("=" * 60)
    print("🔍 PHASE 1: Searching for high-potential companies")
    print("=" * 60)
    
    # Target regions with high BLIIoT product fit
    search_queries = [
        # Africa - SmartGrid market (verified!)
        ("industrial automation system integrator South Africa", "South Africa"),
        ("IoT solutions company South Africa", "South Africa"),
        ("PLC SCADA system integrator South Africa", "South Africa"),
        
        # Energy/Power sector
        ("solar energy monitoring systems South Africa", "South Africa"),
        ("power substation automation South Africa", "South Africa"),
        
        # Target more countries
        ("industrial automation Nigeria", "Nigeria"),
        ("PLC automation Kenya", "Kenya"),
        ("system integrator Zimbabwe", "Zimbabwe"),
    ]
    
    for query, region in search_queries:
        results = finder.search_companies(query, limit=8)
        for r in results:
            name = r.get('title', '')
            desc = r.get('description', '')
            url = r.get('url', '')
            
            # Skip obvious non-companies
            skip_words = ['youtube', 'wikipedia', 'facebook', 'twitter', 'linkedin', 'amazon', 'news']
            if any(s in (url + name).lower() for s in skip_words):
                continue
            
            finder.add_lead(name, url, '', desc, query)
        
        print()
    
    # Phase 2: Output results
    print("\n" + "=" * 60)
    print("📊 PHASE 2: Generating Excel Report")
    print("=" * 60)
    
    output_path = r"C:\Users\Admin\Desktop\Working\BLIIoT_High_Potential_Leads.xlsx"
    finder.output_excel(output_path)
    
    # Summary
    print("\n" + "=" * 60)
    print("📋 FINAL SUMMARY")
    print("=" * 60)
    print(f"  Total companies found: {len(finder.leads)}")
    print(f"  High potential: {sum(1 for l in finder.leads if l['high_potential'] == '✅ YES')}")
    print(f"  With WhatsApp: {sum(1 for l in finder.leads if l['has_whatsapp'] == '✅')}")
    print(f"  With Email(s): {sum(1 for l in finder.leads if l['emails'])}")
    
    # Print top leads
    print("\n🔥 TOP HIGH-POTENTIAL LEADS (WhatsApp first):")
    sorted_for_display = sorted(finder.leads, key=lambda x: (
        -int(x['has_whatsapp'] == '✅'),
        -(x['match_score'] if x['high_potential'] == '✅ YES' else 0)
    ))
    for l in sorted_for_display[:15]:
        wa_flag = '📱 WhatsApp!' if l['has_whatsapp'] else ''
        print(f"  {wa_flag} {l['company']} | Score: {l['match_score']} | {l['emails'] or 'No email'} | {l['phones_found'] or l['phone'] or 'No phone'}")


if __name__ == '__main__':
    main()
