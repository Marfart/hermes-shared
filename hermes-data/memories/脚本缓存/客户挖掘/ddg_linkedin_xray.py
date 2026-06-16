#!/usr/bin/env python3
"""
DDG LinkedIn X-Ray Scraper
通过 DuckDuckGo 搜索引掣批量搜 LinkedIn 公开资料，不登录、不走API、零费用。

用法：python ddg_linkedin_xray.py
输出：ddg_linkedin_leads.xlsx
"""

import csv
import json
import logging
import re
import sys
import time
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("ddg_xray")

# ──────────────────────────
# Data Model
# ──────────────────────────

@dataclass
class LinkedInLead:
    full_name: str
    job_title: str
    company: Optional[str] = None
    location: Optional[str] = None
    connections: Optional[str] = None
    headline: Optional[str] = None
    profile_url: Optional[str] = None
    source_query: str = ""
    email_found: bool = False

# ──────────────────────────
# Search Engine Adapter
# ──────────────────────────

class DDGSearchEngine:
    """DuckDuckGo search using hermes_tools.web_search."""

    @staticmethod
    def search(query: str, max_results: int = 10) -> List[dict]:
        """Search and return results with title, url, description."""
        from hermes_tools import web_search
        result = web_search(query=query, limit=max_results)
        if isinstance(result, dict) and 'data' in result:
            return result['data'].get('web', [])
        return []


# ──────────────────────────
# Result Parser
# ──────────────────────────

def is_linkedin_profile(url: str) -> bool:
    """Check if URL is a LinkedIn personal profile (not company/jobs)."""
    return bool(re.search(r'linkedin\.com/in/', url)) and 'authwall' not in url


def parse_title(title: str) -> tuple:
    """
    Parse Bing/DDG result title into name, job, company.
    Formats:
      "John Doe - CEO at Acme Corp | LinkedIn"
      "John Doe - Software Engineer - Acme Corp | LinkedIn"
      "John Doe | LinkedIn"
    """
    # Remove trailing | LinkedIn / | 领英
    title = re.sub(r'\s*[|]\s*(LinkedIn|领英)\s*$', '', title).strip()
    parts = [p.strip() for p in title.split(' - ')]
    
    name = parts[0] if len(parts) >= 1 else title
    job = parts[1] if len(parts) >= 2 else ""
    company = ""
    
    if len(parts) >= 3:
        # Name - Title - Company
        job = parts[1]
        company = parts[2]
    
    # Handle "Mr./Mrs./Dr." prefixes
    name = re.sub(r'^(Mr\.|Mrs\.|Ms\.|Dr\.|Prof\.)\s+', '', name)
    
    # Split "Title at Company" pattern in job field
    at_match = re.match(r'(.+?)\s+at\s+(.+)', job)
    if at_match:
        job = at_match.group(1).strip()
        if not company:
            company = at_match.group(2).strip()
    
    return name, job, company


def extract_email(text: str) -> Optional[str]:
    """Extract email from text if present."""
    match = re.search(r'[\w.+-]+@[\w-]+\.[\w.-]+', text)
    return match.group(0) if match else None


def extract_location(snippet: str) -> Optional[str]:
    """Extract location from snippet text."""
    # Common patterns in LI snippet descriptions
    patterns = [
        # "Location: City, Country"
        r'Location[:\s]+([^.\n]+)',
        # "City, Province, Country" patterns for Africa
        r'([A-Z][a-z]+(?:\s[A-Z][a-z]+)*,\s*(?:Gauteng|Western Cape|KwaZulu-Natal|Eastern Cape|Limpopo|Mpumalanga|North West|Free State|Northern Cape),\s*(?:South Africa|Namibia|Zimbabwe|Zambia|Kenya|Nigeria|Ghana|Ethiopia|Tanzania|Uganda|Mozambique|Angola|Botswana|Malawi|Rwanda))',
        # "City, Country"
        r'([A-Z][a-z]+(?:\s[A-Z][a-z]+)*,\s*(?:South Africa|Nigeria|Kenya|Ghana|Ethiopia|Tanzania|Uganda|Namibia|Zimbabwe|Zambia|Botswana|Mozambique|Angola|Rwanda|Morocco|Egypt|Algeria|Tunisia))',
        # "Located in City" 
        r'(?:Located|Based|Working)\s+(?:in|at)\s+([A-Z][a-z]+[^.\n]*)',
    ]
    for pattern in patterns:
        match = re.search(pattern, snippet)
        if match:
            return match.group(1).strip()
    return None


def extract_connections(snippet: str) -> Optional[str]:
    """Extract connections/followers count."""
    match = re.search(r'(\d+[KkMmBb]?\+?\s*(?:connections?|followers?))', snippet)
    if match:
        return match.group(1).strip()
    match = re.search(r'(\d+)\+?\s*connections?', snippet)
    if match:
        return match.group(0).strip()
    return None


def parse_snippet_description(snippet, title, name):
    """Extract additional details from the snippet/description."""
    info = {
        'location': None,
        'connections': None,
        'email': None,
        'headline': None,
    }
    
    info['location'] = extract_location(snippet)
    info['connections'] = extract_connections(snippet)
    info['email'] = extract_email(snippet)
    
    # The headline is the snippet minus boilerplate
    headline = snippet
    # Remove known LI boilerplate
    boilerplate = [
        r'View mutual connections[^.]*\.',
        r'\d+ connections on LinkedIn',
        r'\d+[Kk]?\s*followers?\s*\d+[Kk]?\s*connections?',
        r'Experience[:\s]+[^.\n]*',
        r'Education[:\s]+[^.\n]*',
        r'Location[:\s]+[^.\n]*',
    ]
    for pat in boilerplate:
        headline = re.sub(pat, '', headline)
    headline = re.sub(r'\s+', ' ', headline).strip(' .,;')
    
    if headline and headline != name:
        info['headline'] = headline[:200] if len(headline) > 200 else headline
    
    return info


def result_to_lead(result: dict, query: str) -> Optional[LinkedInLead]:
    """Convert a search result dict to a LinkedInLead."""
    url = result.get('url', '')
    title = result.get('title', '')
    snippet = result.get('description', '')
    
    if not is_linkedin_profile(url):
        return None
    
    name, job, company = parse_title(title)
    if not name or name == title:
        return None
    
    # Clean URL
    clean_url = url.split('?')[0] if '?' in url else url
    if 'authwall' in clean_url:
        return None
    
    # Parse details from snippet
    details = parse_snippet_description(snippet, title, name)
    
    # If no company extracted from title, try snippet
    if not company:
        # "at Company" pattern in snippet
        at_match2 = re.search(r'(?:at|@)\s+([A-Z][A-Za-z0-9\s&.]+?)(?:[,.]|\s+[|]|\s*$)', snippet[:200])
        if at_match2:
            company = at_match2.group(1).strip()
    
    # Extract location from title if available
    if not details['location']:
        # Sometimes locations appear in title e.g. "John Doe - Location | LinkedIn"
        location_match = re.search(r'-\s*([A-Z][a-z]+.*(?:South Africa|Kenya|Nigeria|Africa))\s*[|]', result.get('title',''))
        if location_match:
            details['location'] = location_match.group(1).strip()
    
    return LinkedInLead(
        full_name=name,
        job_title=job,
        company=company,
        location=details['location'],
        connections=details['connections'],
        headline=details.get('headline'),
        profile_url=clean_url,
        source_query=query,
        email_found=bool(details.get('email')),
    )


# ──────────────────────────
# Query Templates
# ──────────────────────────

QUERY_TEMPLATES = {
    "africa_system_integrators": [
        'site:linkedin.com/in/ "system integrator" "South Africa"',
        'site:linkedin.com/in/ "system integrator" Africa industrial automation',
        'site:linkedin.com/in/ "system integration" "industrial" "Africa"',
    ],
    "africa_plc_scada": [
        'site:linkedin.com/in/ "PLC" "SCADA" "South Africa" engineer',
        'site:linkedin.com/in/ "PLC" "SCADA" Africa automation',
        'site:linkedin.com/in/ "SCADA engineer" Africa',
        'site:linkedin.com/in/ "automation engineer" "Africa" PLC',
    ],
    "africa_iiot": [
        'site:linkedin.com/in/ "IIoT" Africa engineer',
        'site:linkedin.com/in/ "industrial IoT" Africa',
        'site:linkedin.com/in/ "remote monitoring" Africa SCADA',
        'site:linkedin.com/in/ "IoT" "industrial" "Africa" engineer',
    ],
    "africa_energy": [
        'site:linkedin.com/in/ "solar" "Africa" engineer automation',
        'site:linkedin.com/in/ "energy monitoring" Africa IIoT',
        'site:linkedin.com/in/ "power" "SCADA" "Africa"',
        'site:linkedin.com/in/ "renewable energy" "Africa" automation',
    ],
    "country_specific": [
        'site:linkedin.com/in/ "industrial automation" Nigeria',
        'site:linkedin.com/in/ "automation" Kenya engineer',
        'site:linkedin.com/in/ "control systems" Zimbabwe Zambia',
        'site:linkedin.com/in/ "mining" "automation" South Africa',
        'site:linkedin.com/in/ "water" "SCADA" "South Africa"',
        'site:linkedin.com/in/ "oil and gas" "Africa" automation',
    ],
    "bliiot_targets": [
        'site:linkedin.com/in/ "gateway" "industrial" "Africa"',
        'site:linkedin.com/in/ "RTU" "SCADA" Africa',
        'site:linkedin.com/in/ "edge computing" "industrial" "Africa"',
        'site:linkedin.com/in/ "PLC" "SCADA" "manufacturing" "Africa"',
        'site:linkedin.com/in/ "control systems" "Africa" engineer',
        'site:linkedin.com/in/ "automation" "solutions" "Africa" director',
        'site:linkedin.com/in/ "telemetry" "Africa" industrial',
        'site:linkedin.com/in/ "instrumentation" "Africa" engineer',
    ],
}


# ──────────────────────────
# Main Scraper
# ──────────────────────────

def run_scrape(queries: List[str], max_total: int = 200) -> List[LinkedInLead]:
    """Run multiple X-Ray queries and collect unique leads."""
    all_leads = {}
    seen_urls = set()
    
    engine = DDGSearchEngine()
    
    for i, query in enumerate(queries, 1):
        log.info(f"[{i}/{len(queries)}] Searching: {query[:80]}")
        
        try:
            results = engine.search(query, max_results=10)
        except Exception as e:
            log.warning(f"  ✗ Search failed: {e}")
            continue
        
        lead_count = 0
        for result in results:
            url = result.get('url', '')
            if url in seen_urls:
                continue
            seen_urls.add(url)
            
            lead = result_to_lead(result, query)
            if lead:
                dedup_key = (lead.full_name.lower(), lead.company or '')
                if dedup_key not in all_leads:
                    all_leads[dedup_key] = lead
                    lead_count += 1
        
        log.info(f"  → {lead_count} new leads (total: {len(all_leads)})")
        
        # Polite delay between queries
        if i < len(queries):
            time.sleep(1.5)
        
        if len(all_leads) >= max_total:
            log.info(f"Reached max_total={max_total}, stopping")
            break
    
    return list(all_leads.values())[:max_total]


def export_excel(leads: List[LinkedInLead], output_path: str):
    """Export to Excel (fallback to CSV)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LinkedIn Leads"
        
        headers = [
            "Full Name", "Job Title", "Company", "Location",
            "Connections", "Profile URL", "Headline", "Source Query"
        ]
        hfont = Font(bold=True, color="FFFFFF")
        hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center")
        
        for row, lead in enumerate(leads, 2):
            ws.cell(row=row, column=1, value=lead.full_name)
            ws.cell(row=row, column=2, value=lead.job_title or "")
            ws.cell(row=row, column=3, value=lead.company or "")
            ws.cell(row=row, column=4, value=lead.location or "")
            ws.cell(row=row, column=5, value=lead.connections or "")
            ws.cell(row=row, column=6, value=lead.profile_url or "")
            ws.cell(row=row, column=7, value=lead.headline or "")
            ws.cell(row=row, column=8, value=lead.source_query or "")
        
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
        
        wb.save(output_path)
        log.info(f"✅ Saved {len(leads)} leads to {output_path}")
        
    except ImportError:
        csv_path = output_path.replace(".xlsx", ".csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=[
                "full_name","job_title","company","location",
                "connections","profile_url","headline","source_query"
            ])
            w.writeheader()
            for lead in leads:
                w.writerow(asdict(lead))
        log.info(f"✅ Saved {len(leads)} leads to {csv_path} (CSV fallback)")


def print_summary(leads: List[LinkedInLead]):
    """Print a nice summary."""
    log.info(f"\n{'='*60}")
    log.info(f"📊 TOTAL: {len(leads)} unique LinkedIn leads")
    log.info(f"{'='*60}")
    
    # By location
    locs = {}
    for l in leads:
        loc = l.location or "Unknown"
        locs[loc] = locs.get(loc, 0) + 1
    log.info("\n📍 By Location:")
    for loc, count in sorted(locs.items(), key=lambda x: -x[1])[:10]:
        log.info(f"   {loc:40s} {count}")
    
    # By company top
    cos = {}
    for l in leads:
        if l.company:
            cos[l.company] = cos.get(l.company, 0) + 1
    log.info("\n🏢 Top Companies:")
    for co, count in sorted(cos.items(), key=lambda x: -x[1])[:15]:
        log.info(f"   {co:45s} {count}")
    
    # By job area
    jobs_area = {"SCADA/PLC/Controls": 0, "System Integrator": 0, "IIoT/IoT": 0, "Energy/Solar": 0, "Other": 0}
    for l in leads:
        t = (l.job_title or "").lower()
        if 'scada' in t or 'plc' in t or 'control' in t: jobs_area["SCADA/PLC/Controls"] += 1
        elif 'system integrator' in t or 'integration' in t: jobs_area["System Integrator"] += 1
        elif 'iiot' in t or 'iot' in t: jobs_area["IIoT/IoT"] += 1
        elif 'solar' in t or 'energy' in t or 'renewable' in t: jobs_area["Energy/Solar"] += 1
        else: jobs_area["Other"] += 1
    
    log.info("\n💼 By Job Area:")
    for area, count in sorted(jobs_area.items(), key=lambda x: -x[1]):
        log.info(f"   {area:30s} {count}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="DDG LinkedIn X-Ray Scraper (No Login)")
    parser.add_argument("--template", "-t", choices=list(QUERY_TEMPLATES.keys()),
                        default="bliiot_targets", help="Query template set")
    parser.add_argument("--query", "-q", help="Custom single query")
    parser.add_argument("--limit", "-l", type=int, default=150,
                        help="Max leads (default: 150)")
    parser.add_argument("--output", "-o", default=None,
                        help="Output file (default: auto-name)")
    parser.add_argument("--verbose", "-v", action="store_true")
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Determine queries
    queries = []
    if args.query:
        queries.append(args.query)
    else:
        queries = QUERY_TEMPLATES.get(args.template, QUERY_TEMPLATES["bliiot_targets"])
    
    # Output path
    output = args.output
    if not output:
        ts = datetime.now().strftime("%m%d_%H%M")
        output = f"ddg_linkedin_leads_{ts}.xlsx"
    
    log.info(f"🚀 DDG LinkedIn X-Ray Scraper")
    log.info(f"   Template: {args.template}")
    log.info(f"   Queries: {len(queries)}")
    log.info(f"   Max leads: {args.limit}")
    log.info(f"   Output: {output}")
    log.info("")
    
    leads = run_scrape(queries, max_total=args.limit)
    export_excel(leads, output)
    print_summary(leads)
    
    log.info(f"\n📁 Output file: {output}")