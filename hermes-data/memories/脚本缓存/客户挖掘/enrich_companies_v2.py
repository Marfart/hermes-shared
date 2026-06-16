#!/usr/bin/env python3
"""
LinkedIn Lead Enricher — 用 DDGS 库搜索，更快更稳
输入: Desktop/Working/linkedin_xray_0604_1517.xlsx
输出: Desktop/Working/enriched_leads_FINAL.xlsx
"""

import openpyxl, re, json, time, csv
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══ Search engine ═══

def search_web(query):
    """Use ddgs library for DuckDuckGo search."""
    from ddgs import DDGS
    try:
        s = DDGS()
        results = list(s.text(query, max_results=5))
        return results
    except Exception as e:
        return [{'title': f'ERR: {e}', 'href': '', 'body': ''}]


# ═══ Company name cleaning ═══

GOOD_COMPANIES = {
    'Omniflex', 'Schneider Electric', 'Rand Water', 'Anglo American',
    'Chevron', 'Alstom', 'KHS', 'Hitachi', 'Jendamark Automation',
    'A2Z Automation', 'Rockwell Automation', 'Phoenix Contact',
    'Siemens', 'Emerson', 'Yokogawa', 'Honeywell', 'ABB', 'Endress+Hauser',
    'Automated System Works', 'UltraSec Fencing and Automation',
    'Chilltemp Automation Technologies', 'Letek Industrial Automation',
    'Liquid Automation Systems', 'Quad Automation', 'Aveng Manufacturing',
    'Maximize Systems', 'Honour World', 'LynxTrace', 'INTECH Automation',
    'Ease Automation', 'Dongfang', '4B Systems',
}

def get_company(name, job, raw, desc, url):
    """ID the company from all available signals."""
    text = f"{desc} {job} {url}".lower()
    
    # Check known companies first
    for co in GOOD_COMPANIES:
        if co.lower() in text:
            return co
    
    # Pattern matching
    patterns = [
        (r'(?:at|@|with|for)\s+(Omniflex|Schneider\s+Electric|Rand\s+Water|Anglo\s+American|Chevron|Alstom)\b', 1),
        (r'Experience[:\s]+([A-Z][A-Za-z0-9&.]+(?:\s[A-Z][A-Za-z0-9&.]+){0,3})', 1),
        (r'(?:at|@)\s+([A-Z][A-Za-z0-9&.\s]{2,30}?)(?:[,.]|\s+[|]|\s*$)', 1),
    ]
    for pat, group in patterns:
        m = re.search(pat, desc + ' ' + job)
        if m:
            c = m.group(group).strip()
            if len(c) > 2 and c not in {'Automation','Vice President','Remote','Energy','Ease','Dongf'}:
                return c
    
    # Clean raw
    c = re.sub(r'\s*[|]\s*LinkedIn.*', '', raw)
    c = re.sub(r'\s*LinkedIn.*', '', c)
    c = re.sub(r'\s*\.{3,}\s*$', '', c).strip()
    if len(c) > 2 and c not in {'Automation','Vice President','Remote','Energy','','-','None'}:
        return c
    
    return ''


# ═══ Main ═══

INPUT = r"C:\Users\Admin\Desktop\Working\linkedin_xray_0604_1517.xlsx"
OUTPUT = Path(r"C:\Users\Admin\Desktop\Working")

print("🚀 Loading leads...")
wb = openpyxl.load_workbook(INPUT)
ws = wb.active

people = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[1]: continue
    n = str(row[1] or '').strip()
    j = str(row[2] or '').strip()
    r = str(row[3] or '').strip()
    c = str(row[4] or '').strip()
    u = str(row[5] or '').strip() if len(row)>5 else ''
    d = str(row[7] or '').strip() if len(row)>7 else ''
    co = get_company(n, j, r, d, u)
    people.append({'name':n,'job':j,'company':co,'country':c,'url':u,'desc':d[:200]})

by_co = defaultdict(list)
for p in people:
    by_co[p['company']].append(p)

companies = [(co,ps) for co,ps in by_co.items() if co]
unknown = by_co.get('', [])

print(f"   Total: {len(people)} people")
print(f"   Companies: {len(companies)}")
print(f"   Unknown: {len(unknown)}")
print()

# ═══ Enrich each company ═══

print("🚀 Searching for contacts...")
results = []
UNKNOWN_PLACEHOLDER = ''

for idx, (company, ps) in enumerate(companies, 1):
    names = [p['name'] for p in ps[:3]]
    print(f"\n[{idx}/{len(companies)}] {company}")
    names_str = ', '.join(n for n in names)
    print(f"   People: {names_str}")
    
    # Search 1: website
    results1 = search_web(f'"{company}" official website')
    site_text = json.dumps(results1, ensure_ascii=False)
    
    website = ''
    url_m = re.search(r'https?://(?:www\.)?([A-Za-z0-9][A-Za-z0-9.-]+\.[A-Za-z]{2,})(?:/[^\s"\'<>)]*)?', site_text)
    if url_m:
        full = url_m.group(0)
        skip = ['linkedin','facebook','twitter','instagram','wikipedia','youtube','github','pinterest']
        if not any(s in full.lower() for s in skip):
            website = full.rstrip(')').rstrip('}')
    
    time.sleep(1.5)
    
    # Search 2: CEO + email + phone
    results2 = search_web(f'"{company}" email contact')
    contact_text = json.dumps(results2, ensure_ascii=False)
    
    ceo = ''
    cm = re.search(r'(?:CEO|Chief Executive|Managing Director|Founder|President)\s*[:\-–\s]*\s*([A-Z][a-z]+(?:\s[A-Z][a-z]+){1,2})', contact_text)
    if cm: ceo = cm.group(1).strip()
    
    email = ''
    emails = re.findall(r'([\w.+-]+@(?:[^.]+\.)+(?:com|co\.za|co\.ke|co\.ng|org|net))', contact_text)
    for e in emails:
        if not any(p in e.lower() for p in ['gmail.com','yahoo.com','hotmail.com','outlook.com']):
            email = e
            break
    
    phone = ''
    phones = re.findall(r'(\+27[\s-]?\d[\d\s-]{6,12}|\+234[\s-]?\d[\d\s-]{6,12}|\+254[\s-]?\d[\d\s-]{6,12})', contact_text)
    if phones: phone = phones[0]
    
    time.sleep(1.5)
    
    # Search 3: person-specific (try first contact)
    person_email = ''
    if names and not email:
        results3 = search_web(f'"{names[0]}" "{company}"')
        p_text = json.dumps(results3, ensure_ascii=False)
        pe = re.findall(r'([\w.+-]+@(?:[^.]+\.)+(?:com|co\.za|co\.ke|co\.ng|org|net))', p_text)
        for e in pe:
            if not any(p in e.lower() for p in ['gmail','yahoo','hotmail','outlook']):
                person_email = e
                break
        time.sleep(1.5)
    
    best_email = email or person_email
    
    results.append({
        'company': company,
        'website': website,
        'email': best_email,
        'phone': phone,
        'ceo': ceo,
        'people': ps,
    })
    
    print(f"   🌐 {website[:35] if website else '—'}")
    print(f"   📧 {best_email or '—'}")
    print(f"   📞 {phone or '—'}")
    print(f"   👤 {ceo or '—'}")

# ═══ Export ═══

print("\n🚀 Exporting...")
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Enriched Leads"

headers = ["Company","Contact Name","Job Title","Country",
           "Website","Email","Phone","CEO/Decision Maker","LinkedIn URL","Status"]

hfont = Font(bold=True, color="FFFFFF", size=10)
hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
alt1 = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
thin = Border(bottom=Side(style='thin', color='CCC'))

for c, h in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = hfont; cell.fill = hfill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

row = 2
ai = 0
for ec in results:
    ai += 1
    for p in ec['people']:
        has = bool(ec['email'])
        status = "✅ Email" if has else ("🌐 Website" if ec['website'] else "🔍 Manual")
        vals = [ec['company'], p['name'], p['job'], p['country'],
                ec['website'] or '', ec['email'] or '', ec.get('phone','') or '',
                ec['ceo'] or '', p['url'], status]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=c, value=v)
            cell.border = thin; cell.alignment = Alignment(wrap_text=True, vertical="top")
            if ai % 2 == 0: cell.fill = alt1
            if c == 6 and v: cell.fill = green  # Email green
            if c == 5 and v: cell.font = Font(color="0563C1", underline="single", size=9)
        row += 1

# Unknown
if unknown:
    ws2.cell(row=row, column=1, value="UNKNOWN COMPANY").font = Font(bold=True, color="FF0000")
    row += 1
    for p in unknown:
        ws2.cell(row=row, column=2, value=p['name'])
        ws2.cell(row=row, column=3, value=p['job'])
        ws2.cell(row=row, column=4, value=p['country'])
        ws2.cell(row=row, column=9, value=p['url'])
        ws2.cell(row=row, column=10, value="🔍 Unknown company")
        row += 1

for i, w in enumerate([28,24,35,15,32,32,22,28,45,18], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f"A1:J{row-1}"

ts = datetime.now().strftime("%m%d_%H%M")
out = OUTPUT / f"enriched_leads_{ts}.xlsx"
wb2.save(str(out))

# ═══ Stats ═══

with_email = sum(1 for ec in results if ec['email'])
with_website = sum(1 for ec in results if ec['website'])
with_ceo = sum(1 for ec in results if ec['ceo'])
with_phone = sum(1 for ec in results if ec['phone'])

print(f"\n{'='*60}")
print(f"📊 ENRICHMENT SUMMARY")
print(f"{'='*60}")
print(f"   Companies:        {len(results)}")
print(f"   Total people:     {sum(len(ec['people']) for ec in results)}")
print(f"   With email:       {with_email}")
print(f"   With website:     {with_website}")
print(f"   With CEO:         {with_ceo}")
print(f"   With phone:       {with_phone}")
print(f"\n📁 Output: {out}")
print(f"\n📧 Email summary:")
for ec in results:
    if ec['email']:
        print(f"   ✅ {ec['company']:30s} → {ec['email']}")
    elif ec['website']:
        print(f"   🌐 {ec['company']:30s} → {ec['website'][:40]}")
    else:
        print(f"   🔍 {ec['company']:30s} → nothing found")