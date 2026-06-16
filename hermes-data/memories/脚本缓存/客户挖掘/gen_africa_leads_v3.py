#!/usr/bin/env python3
"""Generate BLIIoT Africa Customer Leads Excel v3"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Africa System Integrators"

# Styles
header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
whatsapp_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')

headers = [
    '#', 'Country', 'Company Name', 'Specialty', 'Email', 'Phone',
    'WhatsApp', 'Website', 'Address', 'Product Match', 'Priority', 'Notes'
]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

leads = [
    # === South Africa ===
    [1, 'South Africa', 'Advansys (Pty) Ltd',
     'PLC, SCADA, MES, IIoT, Industrial Automation, Batch Solutions, Control Systems',
     'information@advansys.co.za', '+27 11 595 8450',
     'Yes', 'https://advansys.co.za',
     'The Pivot, Block C, 2nd Floor, Montecasino Blvd, Fourways 2088, Johannesburg',
     'STARS_5 - PLC/SCADA/IIoT all match BLIIOT portfolio',
     'HIGH - Best fit',
     'Contact: David Greaves (david.greaves@advansys.co.za). Food/Beverage, Mining, Cement, Utilities sectors.'],

    [2, 'South Africa', 'INTEG System Integrators',
     'PLC, SCADA, HMI, MES, Batching, Control Systems Engineering',
     'info@integ.co.za', '+27 21 855 3020',
     'Yes', 'https://integ.co.za',
     'Unit 203/204 Paardevlei Rising 2, 10 Gardner Williams Ave, Somerset West',
     'STARS_5 - PLC/SCADA/HMI perfect match',
     'HIGH - Best fit',
     'Award-winning (Best Partner 2025). Food/Bev, Pharma, Water, Mining. Level 2 B-BBEE.'],

    [3, 'South Africa', 'SAM (Systems Automation & Management)',
     'PLC, SCADA, Fieldbus, Industrial IT, MES, BMS, Motion Control',
     'info@sam.co.za', '+27 11 803 0570',
     'Yes', 'https://www.sam.co.za',
     '13 Rembrandt St, Petervale, Sandton 2191, Gauteng',
     'STARS_5 - Full IIoT/SCADA fit',
     'HIGH - Best fit',
     '40+ years. Offices: Gauteng, KZN, Northern Province, Denmark. ISO 9001:2015. MD: Claudio Agostinetto.'],

    [4, 'South Africa', 'AM Systems Integration',
     'PLC, SCADA, HMI, Wonderware, Industry 4.0, Turnkey Automation',
     'info@amsi.co.za', '+27 11 914 1115',
     'Yes', 'https://www.amsi.co.za',
     'Clearwater Office Park, Atlas Rd, Boksburg, Gauteng 1459',
     'STARS_4 - SCADA/HMI/IIoT good fit',
     'HIGH',
     'Wonderware registered integrator. Offices: Gauteng, Mpumalanga, NW, Botswana. Contact: Michael Rautenbach.'],

    [5, 'South Africa', 'T.M. Engineering (Pty) Ltd',
     'PLC, HMI, SCADA, Industrial Automation, Electronic Repairs',
     'sales@tmeng.co.za', '+27 11 791 1562',
     'Yes', 'https://tmengineering.co.za',
     'Randburg, Gauteng, South Africa',
     'STARS_4 - PLC/SCADA good match',
     'MEDIUM',
     '38 years. Preferred installer of PLCs. After-hours: 083 375 7670.'],

    [6, 'South Africa', 'Calibrant (Pty) Ltd',
     'Industrial Automation, VFD, PLC, HMI/SCADA, OT Security, Systems Integration',
     'Contact via website', 'See website',
     'Yes (WhatsApp on site)', 'https://www.calibrant.co.za',
     'South Africa (nationwide)',
     'STARS_4 - Automation+Security+Integration',
     'MEDIUM',
     'Sophos Silver Partner. Manufacturing, Logistics, Port/Marine, Property sectors. Strong OT security.'],

    [7, 'South Africa', 'Mokhandi Projects',
     'SCADA (Ignition & N3uron), Industrial Automation, Renewable Energy',
     'N/A', 'N/A',
     'Yes', 'https://mokhandi.co.za',
     'South Africa (Indermark area)',
     'STARS_4 - SCADA/Energy perfect for R40/RTU',
     'MEDIUM',
     '2-10 employees. MD: Moshe Mabina. N3uron & Ignition specialist. Renewable energy focus.'],

    [8, 'South Africa', 'Wecon PLC South Africa',
     'PLC, HMI, SCADA, IIoT, VSD, Motion Control, Cloud SCADA',
     'Info@weconplc.co.za', 'N/A',
     'Yes', 'https://weconplc.co.za',
     '97 Jacobson Dr, Lynnwood Ridge, Pretoria 0040',
     'STARS_4 - Distributor/Integrator, IIoT focus',
     'MEDIUM',
     '30+ years. Also: support@weconplc.co.za. PLC fault finding, panel design services.'],

    [9, 'South Africa', 'Nkokhi Group',
     'SCADA, PLC, Defence/Military Automation, Secure Communication',
     'N/A', 'N/A',
     'Yes', 'https://www.nkokhigroup.co.za',
     'South Africa',
     'STARS_3 - Niche defence focus',
     'LOW',
     'Defence/ruggedized SCADA. Very specific niche but could use remote monitoring.'],

    # === Nigeria ===
    [10, 'Nigeria', 'Gil Automations Ltd',
     'PLC, SCADA, Instrumentation, Turbomachinery, Electrical, Automation',
     'info@gilautomation.com', '+234 1 736 0905',
     'Yes', 'https://www.gilautomation.com',
     '10 Gbolahan Lawal Close, off Ashabi Cole St, Alausa, Ikeja, Lagos',
     'STARS_5 - Full industrial automation + Siemens partner',
     'HIGH - Best fit',
     '195 employees. Siemens VAR. Valmet partner. IMI distributor. Oil&Gas, Power sectors.'],

    [11, 'Nigeria', 'Fronthill Controls (Fronthill Ltd)',
     'Building Automation (KNX), BMS, Energy Management, Power Quality',
     'contact@fronthillcontrols.com', '+234 703 423 9259',
     'Yes', 'https://www.fronthillcontrols.com',
     'Lagos, Nigeria',
     'STARS_3 - Building Automation, BLIIOT gateways match',
     'MEDIUM',
     'KNX system integrator. Smart buildings, hotels, data centres. Also: sales@fronthillcontrols.com'],

    [12, 'Nigeria', 'Vacker Nigeria',
     'Automation, Instrumentation, Electrical, MEP, Energy Conservation',
     'contact.ng@vackerglobal.com', 'N/A',
     'Yes', 'https://en-ng.vackerglobal.com',
     'Nigeria (Lagos area)',
     'STARS_3 - General automation, could need IIoT solutions',
     'MEDIUM',
     'Part of Vacker Global group. Measuring instruments, automation, electrical equipment.'],

    # === Kenya / East Africa ===
    [13, 'Kenya', 'IET Africa',
     'PLC, SCADA, Industrial Automation, Drives Solutions',
     'N/A', 'N/A',
     'Yes', 'https://www.ietafrica.com',
     'Nairobi, Kenya',
     'STARS_4 - PLC/SCADA/Drives match for BLIIOT',
     'HIGH',
     'East Africa automation solutions. PLC programming, SCADA integration. Growing market entry point.'],

    [14, 'Kenya', 'AVEVA Select East Africa',
     'SCADA, Industrial Software, System Integration Partner Network',
     'N/A', 'N/A',
     'Yes', 'N/A',
     'East Africa (Kenya based)',
     'STARS_3 - SCADA ecosystem, partner network',
     'LOW',
     'AVEVA distributor for East Africa. Can connect to local SI partners.'],

    [15, 'Multi', 'SMA Altenso (SA/Germany/Australia)',
     'Energy Storage, BESS, Renewable Integration, System Integration',
     'N/A', 'N/A',
     'Yes', 'N/A',
     'South Africa + Germany + Australia',
     'STARS_4 - Energy/IoT monitoring match for R40/RTU',
     'MEDIUM',
     'International SI with SA presence. Large-scale BESS and renewable energy projects.'],
]

country_fills = {
    'South Africa': PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid'),
    'Nigeria': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    'Kenya': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    'Multi': PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid'),
}

priority_fills = {
    'HIGH - Best fit': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'HIGH': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'MEDIUM': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    'LOW': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
}

for row_idx, lead in enumerate(leads, 2):
    for col_idx, val in enumerate(lead, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = wrap_align
        if col_idx == 2 and val in country_fills:
            cell.fill = country_fills[val]
        if col_idx == 7:
            cell.fill = whatsapp_fill
            cell.alignment = Alignment(horizontal='center', vertical='top')
        if col_idx == 11 and val in priority_fills:
            cell.fill = priority_fills[val]
            cell.font = Font(bold=True)
        if col_idx == 10 and 'STARS_5' in str(val):
            cell.font = Font(bold=True, color='006100')

col_widths = [4, 14, 28, 42, 38, 22, 16, 30, 45, 40, 16, 55]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:L{len(leads)+1}'

# === Summary Sheet ===
ws2 = wb.create_sheet('Summary')
ws2['A1'] = 'BLIIoT Africa Customer Leads Summary'
ws2['A1'].font = Font(bold=True, size=14)
ws2.merge_cells('A1:D1')

ws2['A3'] = 'Total Leads'; ws2['B3'] = len(leads)
ws2['A4'] = 'South Africa'; ws2['B4'] = sum(1 for l in leads if l[1]=='South Africa')
ws2['A5'] = 'Nigeria'; ws2['B5'] = sum(1 for l in leads if l[1]=='Nigeria')
ws2['A6'] = 'Kenya/East Africa'; ws2['B6'] = sum(1 for l in leads if l[1]=='Kenya')
ws2['A7'] = 'Multi-country'; ws2['B7'] = sum(1 for l in leads if l[1]=='Multi')

ws2['C3'] = 'HIGH Priority'; ws2['D3'] = sum(1 for l in leads if 'HIGH' in str(l[11]))
ws2['C4'] = 'MEDIUM Priority'; ws2['D4'] = sum(1 for l in leads if 'MEDIUM' in str(l[11]))
ws2['C5'] = 'LOW Priority'; ws2['D5'] = sum(1 for l in leads if 'LOW' in str(l[11]))

ws2['A9'] = 'With Email'; ws2['B9'] = sum(1 for l in leads if l[4] not in ('N/A', 'Contact via website'))
ws2['A10'] = 'With Phone'; ws2['B10'] = sum(1 for l in leads if l[5] != 'N/A' and l[5] != 'See website')
ws2['A11'] = 'WhatsApp Enabled'; ws2['B11'] = sum(1 for l in leads if l[6] == 'Yes' or 'WhatsApp' in str(l[6]))

for cell in ws2['A3':'D11']:
    for c in cell:
        if c.value:
            c.border = thin_border

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 12

# Save
output_path = 'C:/Users/Admin/Desktop/Working/BLIIoT_Customer_Leads_v3.xlsx'
wb.save(output_path)

sa = sum(1 for l in leads if l[1]=='South Africa')
ng = sum(1 for l in leads if l[1]=='Nigeria')
ke = sum(1 for l in leads if l[1]=='Kenya')
mu = sum(1 for l in leads if l[1]=='Multi')
hi = sum(1 for l in leads if 'HIGH' in str(l[11]))
md = sum(1 for l in leads if 'MEDIUM' in str(l[11]))
lo = sum(1 for l in leads if 'LOW' in str(l[11]))
em = sum(1 for l in leads if l[4] not in ('N/A', 'Contact via website'))
wa = sum(1 for l in leads if l[6]=='Yes' or 'WhatsApp' in str(l[6]))

print(f"=== BLIIoT Africa Customer Leads v3 ===")
print(f"Total: {len(leads)} leads")
print(f"  South Africa: {sa}  |  Nigeria: {ng}  |  Kenya: {ke}  |  Multi: {mu}")
print(f"  HIGH: {hi}  |  MEDIUM: {md}  |  LOW: {lo}")
print(f"  With Email: {em}  |  WhatsApp ready: {wa}")
print(f"Saved: {output_path}")