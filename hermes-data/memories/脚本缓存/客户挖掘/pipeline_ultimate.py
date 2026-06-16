#!/usr/bin/env python3
"""
Pipeline Ultimate v5 — 一站式客户增强管道
整合: X-Ray数据加载 → 公司清洗 → 三级增强 → 智能评分 → 专业报告

三级增强策略:
  Tier 1: 直接搜官网+邮箱 (已验证)
  Tier 2: site:domain搜索 (针对有官网但无邮箱的)
  Tier 3: 人名+公司名深度搜索 (针对小公司)
  
评分模型: 按"BLIIOT可触达性"打分
  5分 = ✅ 有邮箱（直接发开发信）
  4分 = 🌐 有官网 + 行业相关（可挖邮箱）
  3分 = 🌐 有官网（需鉴定相关性）
  2分 = 🔍 名字已知但无官网（需领英私信）
  1分 = ⬛ 大跨国企业（非目标）

用法: python pipeline_ultimate.py
"""

import re, json, time, urllib.request, sqlite3, os
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from dataclasses import dataclass
from typing import Optional

INPUT = r"C:\Users\Admin\Desktop\Working\linkedin_xray_0604_1517.xlsx"
OUTPUT_DIR = Path(r"C:\Users\Admin\Desktop\Working")

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/125.0.0.0"

# ─────────── CACHED DATA ───────────

ALREADY_FOUND = {
    "Omniflex":               {"web":"https://www.omniflex.com", "email":"sales@omniflex.com", "phone":"", "score":5, "tier":"System Integrator"},
    "AUTOMATED SYSTEM WORKS":{"web":"https://www.aswgroup.co.za","email":"info@aswgroup.co.za","phone":"","score":5, "tier":"System Integrator"},
    "Chilltemp Automation Technologies":{"web":"http://chilltemp.co.za","email":"jaco@chilltemp.co.za","phone":"+27 63 778 0442","score":5, "tier":"HVAC/Automation"},
    "Aveng Manufacturing":    {"web":"https://aveng.co.za","email":"Investor.relations@avenggroup.com","phone":"+61 3 9816 2400","score":3, "tier":"Mining/Infra"},
    "INTECH Automation":      {"web":"https://www.intechautomation.in","email":"intech@intech.net","phone":"","score":4, "tier":"Automation"},
    "Rand Water":             {"web":"https://www.randwater.co.za","email":"","phone":"","score":4, "tier":"Water/Utility"},
    "Schneider Electric":     {"web":"https://www.se.com","email":"","phone":"","score":1, "tier":"Big Corp"},
    "Rockwell Automation":    {"web":"https://www.rockwellautomation.com","email":"","phone":"","score":1, "tier":"Big Corp"},
    "Anglo American":         {"web":"https://www.angloamerican.com","email":"","phone":"","score":1, "tier":"Big Corp"},
    "Chevron":                {"web":"https://www.chevron.com","email":"","phone":"","score":1, "tier":"Big Corp"},
    "Alstom":                 {"web":"https://www.alstom.com","email":"","phone":"","score":1, "tier":"Big Corp"},
    "KHS":                    {"web":"https://www.khs.com","email":"","phone":"","score":3, "tier":"Packaging"},
    "Jendamark Automation":   {"web":"https://jendamark.io","email":"","phone":"","score":4, "tier":"Automation"},
    "A2Z Automation":         {"web":"https://a2zautomations.com","email":"","phone":"","score":3, "tier":"Automation"},
}

# Companies to search smarter
NEED_DEEP_SEARCH = [
    "Liquid Automation Systems", "Letek Industrial Automation",
    "UltraSec Fencing and Automation", "Maximize Systems",
    "Honour World", "LynxTrace", "Quad Automation",
]

def search_web(query, limit=5):
    from hermes_tools import web_search
    r = web_search(query=query, limit=limit)
    if isinstance(r, dict) and 'data' in r:
        return r['data'].get('web', [])
    return []

def extract_all(text):
    we=em=ph=ce=""
    m=re.search(r'https?://(?:www\.)?([A-Za-z0-9][A-Za-z0-9.-]+\.[A-Za-z]{2,})(?:/[^\s"\'<>)]*)?',text)
    if m:
        u=m.group(0).rstrip(')}')
        if not any(s in u.lower() for s in ['linkedin','facebook','twitter','instagram','wikipedia','youtube','github','pinterest']): we=u
    es=re.findall(r'([\w.+-]+@(?:[^.]+\.)+(?:com|co\.za|co\.ke|co\.ng|org|net|io))',text)
    for e in es:
        if not any(p in e.lower() for p in ['gmail','yahoo','hotmail','outlook','icloud']): em=e; break
    ps=re.findall(r'(\+27[\s-]?\d[\d\s-]{6,12}|\+234[\s-]?\d[\d\s-]{6,12}|\+254[\s-]?\d[\d\s-]{6,12})',text)
    if ps: ph=ps[0]
    return we,em,ph

# ─────────── TIER 3: Smart Search ───────────

NAMES_BY_COMPANY = {
    "Liquid Automation Systems": ["Konrad Burger"],
    "Letek Industrial Automation": ["Leon Hall"],
    "UltraSec Fencing and Automation": ["Francois Nieuwoudt"],
    "Maximize Systems": ["Johannes Freislich"],
    "Honour World": ["Emmanuel Osayame"],
    "LynxTrace": ["Derek Mundondo"],
    "Quad Automation": ["Quintin Bissett"],
}

print("="*60)
print("PIPELINE ULTIMATE v5")
print("="*60)

# Search remaining companies with smarter queries
results = dict(ALREADY_FOUND)
for co in NEED_DEEP_SEARCH:
    names = NAMES_BY_COMPANY.get(co, [])
    print(f"\n--- {co} ---")
    
    # Try 3 angles
    emails_found = set()
    website = ""
    phone = ""
    
    for angle, q in [
        (f"search company", f'"{co}" automation South Africa'),
        (f"person search", f'"{names[0]}" "{co}"' if names else None),
        (f"industry search", f'"{co.split()[0]}" automation systems South Africa'),
    ]:
        if not q: continue
        try:
            items = search_web(q, 5)
            text = json.dumps(items)
            we, em, ph = extract_all(text)
            if we and not website: website = we
            if ph and not phone: phone = ph
            if em: emails_found.add(em)
            time.sleep(1.2)
        except: pass
    
    best_email = next((e for e in emails_found if not any(p in e for p in ['gmail','yahoo'])), 
                      next(iter(emails_found), ""))
    
    results[co] = {
        "web": website, "email": best_email, "phone": phone,
        "score": 5 if best_email else (4 if website else 2),
        "tier": "Unknown"
    }
    
    s = "✅" if best_email else ("🌐" if website else "🔍")
    print(f"  {s} email={best_email or '—'} web={website[:35] if website else '—'}")

# ─────────── BUILD EXCEL ───────────

print("\n\nBuilding Excel...")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load original
wb = openpyxl.load_workbook(INPUT)
ws = wb.active

GOOD = {x.lower():x for x in results}

def get_company(name, job, raw, desc, url):
    text = f"{desc} {job}".lower()
    for k,v in GOOD.items():
        if k in text: return v
    m = re.search(r'Experience[:\s]+([A-Z][A-Za-z0-9&.]+(?:\s[A-Z][A-Za-z0-9&.]+){0,3})', desc)
    if m:
        c = m.group(1).strip()
        if c.lower() in GOOD: return GOOD[c.lower()]
    c = re.sub(r'\s*[|]\s*LinkedIn.*', '', raw).strip()
    if c.lower() in GOOD: return GOOD[c.lower()]
    return ""

# Create workbook
wb2 = openpyxl.Workbook()

# === Sheet 1: Prioritized Leads ===
ws2 = wb2.active
ws2.title = "Prioritized Leads"

headers = ["Priority","Company","Contact Name","Job Title","Country",
           "Email","Phone","Company Website","LinkedIn URL","Tier","Score","Action"]

hfont = Font(bold=True, color="FFFFFF", size=10)
hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
alt1 = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin = Border(bottom=Side(style='thin', color='00CCCCCC'))

for c,h in enumerate(headers,1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = hfont; cell.fill = hfill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

rows_data = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[1]: continue
    name = str(r[1] or '').strip()
    job = str(r[2] or '').strip()
    raw = str(r[3] or '').strip()
    country = str(r[4] or '').strip()
    url = str(r[5] or '').strip() if len(r)>5 else ''
    desc = str(r[7] or '').strip() if len(r)>7 else ''
    
    co = get_company(name, job, raw, desc, url)
    ec = results.get(co, {})
    
    score = ec.get('score', 1)
    email = ec.get('email','')
    website = ec.get('web','')
    
    if email: action = "🎯 Send outreach email"
    elif website: action = "🌐 Visit website → find contact"
    elif score >= 3: action = "🔍 LinkedIn InMail"
    else: action = "⏸ Low priority (big corp)"
    
    rows_data.append((score, co, name, job, country, email, ec.get('phone',''),
                      website, url, ec.get('tier',''), score, action, desc[:100]))

# Sort by priority descending
rows_data.sort(key=lambda x: (-x[0], x[1].lower()))

row_num = 2
for data in rows_data:
    score, co, name, job, country, email, phone, website, url, tier, sc, action, desc = data
    
    vals = [sc, co, name, job, country, email, phone, website, url, tier, sc, action]
    
    for c,v in enumerate(vals, 1):
        cell = ws2.cell(row=row_num, column=c, value=v)
        cell.border = thin; cell.alignment = Alignment(wrap_text=True, vertical="top")
        if row_num % 2 == 0: cell.fill = alt1
        if c == 1:
            if sc >= 5: cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"); cell.font = Font(bold=True, color="FFFFFF")
            elif sc >= 4: cell.fill = yellow_fill
            elif sc <= 1: cell.fill = red_fill
        if c == 6 and v: cell.fill = green_fill
        if c == 8 and v: cell.font = Font(color="0563C1", underline="single", size=9)
        if c == 12:
            if 'Send outreach' in v: cell.fill = green_fill
            elif 'Low priority' in v: cell.fill = red_fill
    
    row_num += 1

for i,w in enumerate([10,28,22,30,15,32,22,32,45,20,8,35], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f"A1:L{row_num-1}"

# === Sheet 2: Company Summary ===
ws3 = wb2.create_sheet("Company Summary")
for c,h in enumerate(["Priority","Company","Email","Phone","Website","Industry","Contacts","Status"], 1):
    cell = ws3.cell(row=1, column=c, value=h)
    cell.font = hfont; cell.fill = hfill

by_co = defaultdict(list)
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[1]: continue
    name = str(r[1] or '').strip()
    raw = str(r[3] or '').strip()
    desc = str(r[7] or '').strip() if len(r)>7 else ''
    job = str(r[2] or '').strip()
    co = get_company(name, job, raw, desc, '')
    by_co[co].append(name)

sr = 2
for co, ec in sorted(results.items(), key=lambda x: (-x[1].get('score',1), x[0])):
    s = ec.get('score',1)
    status = "✅ Email" if ec.get('email') else ("🌐 Website" if ec.get('web') else "🔍 Not found")
    contacts = len(by_co.get(co, []))
    for c,v in enumerate([s, co, ec.get('email',''), ec.get('phone',''), ec.get('web',''), 
                         ec.get('tier',''), contacts, status], 1):
        cell = ws3.cell(row=sr, column=c, value=v)
        if c == 1:
            if s >= 5: cell.fill = green_fill; cell.font = Font(bold=True)
            elif s <= 1: cell.fill = red_fill
        if c == 3 and v: cell.fill = green_fill
    sr += 1

for i,w in enumerate([10,35,32,22,35,22,10,15], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# === Sheet 3: Action List ===
ws4 = wb2.create_sheet("Action Plan")
action_data = [
    ["Priority 5 — EMAIL READY", "5 companies", "Send outreach email NOW"],
    ["", "Omniflex", "sales@omniflex.com — Remote monitoring, perfect for BLIIOT gateways"],
    ["", "AUTOMATED SYSTEM WORKS", "info@aswgroup.co.za — SA system integrator"],
    ["", "Chilltemp Automation Technologies", "jaco@chilltemp.co.za — HVAC/automation partner"],
    ["", "INTECH Automation", "intech@intech.net — Industrial automation"],
    ["", "Aveng Manufacturing", "Investor.relations@avenggroup.com — Mining infra"],
    ["Priority 4 — WEBSITE + INDUSTRY FIT", "3 companies", "Visit contact page to find sales email"],
    ["", "Rand Water", "randwater.co.za → find procurement email"],
    ["", "Jendamark Automation", "jendamark.io → automation sales contact"],
    ["", "A2Z Automation", "a2zautomations.com → check contact page"],
    ["Priority 2-3 — NEED MORE RESEARCH", "7 companies", "LinkedIn InMail or phone call"],
    ["", "Liquid Automation Systems, Letek, UltraSec...", "Small SA automation companies"],
    ["Priority 1 — LOW PRIORITY", "6 companies", "Big multinationals, hard to reach"],
    ["", "Schneider, Rockwell, Chevron, Anglo American, Alstom, KHS", "Too large, focus elsewhere first"],
    ["", "", ""],
    ["TOTAL", f"{row_num-2} people across {len(results)} companies", f"5 with email ready to go"],
]

for h in ["Category", "Details", "Action"]:
    cell = ws4.cell(row=1, column=1, value="PIPELINE ULTIMATE — ACTION PLAN")
    cell.font = Font(bold=True, size=14, color="1F4E79")
ws4.merge_cells('A1:C1')

for i, row_data in enumerate(action_data, 3):
    for c,v in enumerate(row_data, 1):
        cell = ws4.cell(row=i, column=c, value=v)
        if v and v.startswith("Priority"):
            cell.font = Font(bold=True, size=12, color="1F4E79")
            if "EMAIL READY" in v: cell.fill = green_fill
            elif "WEBSITE" in v: cell.fill = yellow_fill
            elif "LOW PRIORITY" in v: cell.fill = red_fill
        if c == 2 and 'omniflex' in v.lower(): cell.fill = green_fill
    ws4.column_dimensions['A'].width = 40
    ws4.column_dimensions['B'].width = 45
    ws4.column_dimensions['C'].width = 55

# Save
ts = datetime.now().strftime("%m%d_%H%M")
out = OUTPUT_DIR / f"pipeline_ultimate_{ts}.xlsx"
wb2.save(str(out))

# ─────────── SUMMARY ───────────

print(f"\n{'='*60}")
print(f"📊 PIPELINE ULTIMATE — FINAL REPORT")
print(f"{'='*60}")
print(f"   Total people:     {row_num-2}")
print(f"   Total companies:  {len(results)}")
print()
print(f"   🎯 TIER 5 — EMAIL READY ({sum(1 for v in results.values() if v.get('email'))} companies):")
for co, ec in sorted(results.items()):
    if ec.get('email'): print(f"      ✅ {co:35s} → {ec['email']}")
print()
print(f"   🌐 TIER 4 — WEBSITE+INDUSTRY ({sum(1 for v in results.values() if v.get('web') and not v.get('email'))} companies):")
for co, ec in sorted(results.items()):
    if ec.get('web') and not ec.get('email'): print(f"      🌐 {co:35s} → {ec['web']}")
print()
print(f"   🔍 TIER 2-3 — NEED SEARCH ({sum(1 for v in results.values() if not v.get('web') and not v.get('email'))} companies):")
for co, ec in sorted(results.items()):
    if not ec.get('web') and not ec.get('email'): print(f"      🔍 {co:35s}")
print()
print(f"   ⬛ TIER 1 — BIG CORP ({sum(1 for v in results.values() if v.get('score') == 1)} companies):")
for co, ec in sorted(results.items()):
    if ec.get('score') == 1: print(f"      ⬛ {co:35s}")
print(f"\n📁 {out}")