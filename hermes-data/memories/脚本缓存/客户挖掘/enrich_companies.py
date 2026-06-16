#!/usr/bin/env python3
"""
LinkedIn Lead Enricher v2 — 客户信息管道
对每家公司搜索：官网 → CEO/决策人 → 邮箱 → 电话
输出：Desktop/Working/enriched_leads_FINAL.xlsx

用法：python enrich_companies.py
"""

import openpyxl, re, json, time, csv, os, sys
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT = r"C:\Users\Admin\Desktop\Working\linkedin_xray_0604_1517.xlsx"
OUTPUT_DIR = Path(r"C:\Users\Admin\Desktop\Working")

# ═══════════════════════════════════════
# STEP 0: Load & clean company names
# ═══════════════════════════════════════

def extract_company(name, job, raw, desc, url):
    candidates = []
    m = re.search(r'Experience[:\s]+((?:[A-Z][A-Za-z0-9&.]+(?:\s[A-Z][A-Za-z0-9&.]+){0,4}))', desc)
    if m: candidates.append(m.group(1).strip())
    m = re.search(r'(?:at|@)\s+([A-Z][A-Za-z0-9&.\s]{2,40}?)(?:[,]|\s*[|]|\s*$)', job)
    if m: candidates.append(m.group(1).strip())
    cleaned = re.sub(r'\s*\|\s*LinkedIn.*', '', raw)
    cleaned = re.sub(r'\s*LinkedIn.*', '', cleaned)
    cleaned = re.sub(r'\s*\.{3,}\s*$', '', cleaned)
    cleaned = re.sub(r'^LinkedIn', '', cleaned)
    cleaned = re.sub(r'\s+$', '', cleaned)
    if cleaned and len(cleaned) > 2: candidates.append(cleaned.strip())
    if candidates:
        best = min(candidates, key=lambda c: len(c))
        return best
    return cleaned if cleaned else ""


def search_bing(query, limit=5):
    """Search using requests via Bing (free, no API key)."""
    import urllib.request, urllib.parse
    url = f"https://www.bing.com/search?q={urllib.parse.quote(query)}"
    req = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml",
        "Accept-Language": "en-US,en;q=0.9",
    })
    try:
        with urllib.request.urlopen(req, timeout=12) as resp:
            html = resp.read().decode("utf-8", errors="replace")
    except: return []
    
    results = []
    # Parse Bing results
    blocks = re.findall(r'<li class="b_algo[^"]*"[^>]*>(.*?)</li>', html, re.DOTALL)
    for block in blocks:
        link_m = re.search(r'<a[^>]*href="(https?://[^"]+)"[^>]*>(.*?)</a>', block, re.DOTALL)
        snippet_m = re.search(r'<p[^>]*>(.*?)</p>', block, re.DOTALL)
        if link_m:
            url_r = link_m.group(1).split('?')[0]
            title = re.sub(r'<[^>]+>', '', link_m.group(2)).strip()
            snippet = re.sub(r'<[^>]+>', '', snippet_m.group(1)).strip() if snippet_m else ''
            results.append({'url': url_r, 'title': title, 'snippet': snippet})
    return results[:limit]


def search_company(company, query_suffix, limit=3):
    """Search for company + suffix, return text results."""
    q = f'"{company}" {query_suffix}'
    results = search_bing(q, limit)
    text = ' '.join(f"{r['title']} {r['snippet']}" for r in results)
    return text


# ═══════════════════════════════════════
# STEP 1: Load & clean
# ═══════════════════════════════════════

print("🚀 Step 1: Loading & cleaning company names...")
wb = openpyxl.load_workbook(INPUT)
ws = wb.active

person_companies = []  # list of {name, job, company, country, url, desc}

for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    if not row[1]: continue
    name = str(row[1] or '').strip()
    job = str(row[2] or '').strip()
    raw = str(row[3] or '').strip()
    country = str(row[4] or '').strip()
    url = str(row[5] or '').strip() if len(row) > 5 else ''
    desc = str(row[7] or '').strip() if len(row) > 7 else ''
    co = extract_company(name, job, raw, desc, url)
    person_companies.append({
        'name': name, 'job': job, 'company': co,
        'country': country, 'url': url, 'desc': desc[:300],
    })

# Group by company
by_co = defaultdict(list)
for p in person_companies:
    by_co[p['company']].append(p)

print(f"   Total people: {len(person_companies)}")
print(f"   Raw companies: {len(by_co)}")

# Filter out junk company names
JUNK = {'', 'Unknown', 'Remote', 'Automation', 'Vice President', 'None', '-',
        'Junior Electrical Control', 'ENERGY', 'Dongf', 'Ease',
        'LinkedIn', '4B Systems and', 'Application Engineer RTU/SCADA at HITACHI',
        'HITACHI ...Ramesh Kumar', 'Global Industrial Automation Leader',
        'Automation Technology Lead Africa', 'Sr. Automation Engineer at Chevron',
        'Engineer Control Systems', 'Industrial Automation',
        'Lead Automation & Solution'}

# List of companies to research
research_queue = []
unknown_people = []

for co, people in sorted(by_co.items()):
    if co in JUNK or len(co) < 3:
        unknown_people.extend(people)
    else:
        research_queue.append((co, people))

# For unknown people, try to extract company from deeper context
print(f"\n   Clean companies: {len(research_queue)}")
print(f"   Unknown people: {len(unknown_people)}")

# Try to rescue unknown by searching each person's name+jobs
rescued = []
still_unknown = []
for p in unknown_people:
    # Try to find company from their full description/URL
    text = f"{p['desc']} {p['url']} {p['job']}"
    
    # Look for company names in known list
    found = False
    for co, _ in research_queue:
        if co.lower() in text.lower():
            rescued.append({**p, 'company': co})
            found = True
            break
    
    if not found:
        # Try description extraction
        m = re.search(r'(?:at|@|with|for)\s+([A-Z][A-Za-z0-9&.\s]{3,30}?)(?:[,.]|\s+[|]|\s*$)', text)
        if m:
            candidate = m.group(1).strip()
            if len(candidate) > 3 and candidate not in JUNK:
                rescued.append({**p, 'company': candidate})
                found = True
    
    if not found:
        still_unknown.append(p)

# Add rescued back to research queue
rescue_companies = defaultdict(list)
for p in rescued:
    rescue_companies[p['company']].append(p)

for co, people in rescue_companies.items():
    research_queue.append((co, people))

print(f"   Rescued: {len(rescued)}")
print(f"   Still unknown: {len(still_unknown)}")
print(f"   Final companies to research: {len(research_queue)}")
print()

# ═══════════════════════════════════════
# STEP 2: Enrich each company
# ═══════════════════════════════════════

print("🚀 Step 2: Enriching companies (website, CEO, email, phone)...")
print()

enriched = []

for idx, (company, people) in enumerate(research_queue, 1):
    names = [p['name'] for p in people[:3]]
    print(f"[{idx}/{len(research_queue)}] {company} ({len(people)} contacts)")
    print(f"   People: {', '.join(names[:3])}{'...' if len(names)>3 else ''}")
    
    # 2a) Find website
    website = ""
    site_text = search_company(company, "official website company", 5)
    link_m = re.search(r'https?://(?:www\.)?([A-Za-z0-9][A-Za-z0-9.-]+[A-Za-z0-9]\.[A-Za-z]{2,})[^\s"\'<>]*', site_text)
    if link_m:
        url_candidate = link_m.group(0)
        skip = ['linkedin.com','facebook.com','twitter.com','instagram.com',
                'wikipedia.org','youtube.com','indeed.com','glassdoor.com',
                'crunchbase.com','pinterest.com','github.com']
        if not any(s in url_candidate.lower() for s in skip):
            website = url_candidate.split(')')[0].split('}')[0]
    
    time.sleep(1.5)
    
    # 2b) Find CEO / Decision Maker
    ceo_name = ""
    ceo_text = search_company(company, "CEO managing director president", 5)
    m = re.search(r'(?:CEO|Chief Executive|Managing Director|President|Founder)\s*(?:[:\-–\s]+)?\s*([A-Z][a-z]+(?:\s[A-Z][a-z]+){1,2})', ceo_text)
    if m: ceo_name = m.group(1).strip()
    
    time.sleep(1.5)
    
    # 2c) Find email
    contact_email = ""
    phone = ""
    
    # Try email search
    email_text = search_company(company, "email contact info address", 5)
    
    # Extract email
    emails = re.findall(r'([\w.+-]+@(?:[^.]+\.)+(?:com|co\.za|co\.ke|co\.ng|org|net|io))', email_text)
    for e in emails:
        if not any(p in e.lower() for p in ['gmail.com','yahoo.com','hotmail.com','outlook.com']):
            contact_email = e
            break
    
    # Extract phone (SA/Nigeria/Kenya format)
    phones = re.findall(r'(\+27[\s-]?\d{1,4}[\s-]?\d{1,4}[\s-]?\d{1,4}|\+234[\s-]?\d{1,4}[\s-]?\d{1,4}[\s-]?\d{1,4}|\+254[\s-]?\d{1,4}[\s-]?\d{1,4}[\s-]?\d{1,4}|\+?27[\s-]?\d{1,4}[\s-]?\d{1,4}[\s-]?\d{1,4})', email_text)
    if phones: phone = phones[0]
    
    time.sleep(1.5)
    
    # 2d) If no email found, try specific person search
    person_email = ""
    if not contact_email and people[0]['name']:
        person_text = search_company(people[0]['name'], f"{company} email", 3)
        emails2 = re.findall(r'([\w.+-]+@(?:[^.]+\.)+(?:com|co\.za|co\.ke|co\.ng|org|net))', person_text)
        for e in emails2:
            if not any(p in e.lower() for p in ['gmail.com','yahoo.com','hotmail.com','outlook.com']):
                person_email = e
                break
        time.sleep(1.5)
    
    # 2e) Try to find contact page via direct website scrape
    scrape_email = ""
    scrape_phone = ""
    if website:
        try:
            import urllib.request
            req = urllib.request.Request(website, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            })
            with urllib.request.urlopen(req, timeout=8) as resp:
                html = resp.read().decode("utf-8", errors="replace")[:100000]
            
            # Look for email in page
            emails3 = re.findall(r'([\w.+-]+@(?:[^.]+\.)+(?:com|co\.za|co\.ke|co\.ng|org|net))', html)
            for e in emails3:
                if not any(p in e.lower() for p in ['gmail.com','yahoo.com','hotmail.com','outlook.com']):
                    scrape_email = e
                    break
            
            # Look for phone in page
            phones3 = re.findall(r'(\+27[\s-]?\d{1,4}[\s-]?\d{1,4}[\s-]?\d{1,4})', html)
            if phones3: scrape_phone = phones3[0]
        except:
            pass
    
    # Combine results
    best_email = scrape_email or contact_email or person_email
    best_phone = scrape_phone or phone
    
    result = {
        'company': company,
        'website': website,
        'ceo': ceo_name,
        'email': best_email,
        'phone': best_phone,
        'people': people,
    }
    
    print(f"   🌐 Website: {website or '—'}")
    print(f"   👤 CEO: {ceo_name or '—'}")
    print(f"   📧 Email: {best_email or '—'}")
    print(f"   📞 Phone: {best_phone or '—'}")
    print()
    
    enriched.append(result)

# Add unknown people at the bottom
if still_unknown:
    enriched.append({
        'company': 'UNKNOWN (need more info)',
        'website': '',
        'ceo': '',
        'email': '',
        'phone': '',
        'people': still_unknown,
    })

# ═══════════════════════════════════════
# STEP 3: Export to Excel
# ═══════════════════════════════════════

print("🚀 Step 3: Exporting to Excel...")

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Enriched Leads"

headers = [
    "Company", "Contact Name", "Job Title", "Country",
    "Company Website", "Company Email", "Company Phone",
    "CEO/Decision Maker", "LinkedIn URL", "Status"
]

hfont = Font(bold=True, color="FFFFFF", size=10)
hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
alt1 = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
alt2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
thin_border = Border(bottom=Side(style='thin', color='CCC'))

for col, h in enumerate(headers, 1):
    cell = ws_out.cell(row=1, column=col, value=h)
    cell.font = hfont
    cell.fill = hfill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = Border(bottom=Side(style='thin', color='1F4E79'))

row_num = 2
stats = {'total': 0, 'with_email': 0, 'with_website': 0, 'with_ceo': 0, 'with_phone': 0}

for ec in enriched:
    company = ec['company']
    has_email = bool(ec['email'])
    has_website = bool(ec['website'])
    has_ceo = bool(ec['ceo'])
    has_phone = bool(ec['phone'])
    
    for p in ec['people']:
        # Status tag
        if has_email:
            status = "✅ Email found"
        elif has_website:
            status = "🌐 Website only"
        else:
            status = "🔍 Need more research"
        
        vals = [
            company, p['name'], p['job'], p['country'],
            ec['website'] or '', ec['email'] or '', ec.get('phone', '') or '',
            ec['ceo'] or '', p['url'], status
        ]
        
        for col, val in enumerate(vals, 1):
            cell = ws_out.cell(row=row_num, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = alt1 if row_num % 2 == 0 else alt2
            
            if col == 6 and val:  # Email column - green
                cell.fill = green
            if col == 5 and val:  # Website
                cell.font = Font(color="0563C1", underline="single", size=9)
            if col in (8, 9) and val:
                cell.fill = green
        
        stats['total'] += 1
        if has_email: stats['with_email'] += 1
        if has_website: stats['with_website'] += 1
        if has_ceo: stats['with_ceo'] += 1
        if has_phone: stats['with_phone'] += 1
        
        row_num += 1

# Column widths
for i, w in enumerate([28, 24, 35, 15, 32, 32, 22, 28, 45, 20], 1):
    ws_out.column_dimensions[get_column_letter(i)].width = w

ws_out.freeze_panes = 'A2'

# Auto-filter
ws_out.auto_filter.ref = f"A1:J{row_num-1}"

# Save
ts = datetime.now().strftime("%m%d_%H%M")
output_path = OUTPUT_DIR / f"enriched_leads_{ts}.xlsx"
wb_out.save(str(output_path))

# ═══════════════════════════════════════
# FINAL REPORT
# ═══════════════════════════════════════

print(f"\n{'='*60}")
print(f"📊 FINAL REPORT")
print(f"{'='*60}")
print(f"   Companies researched:  {len([e for e in enriched if e['company'] != 'UNKNOWN'])}")
print(f"   Total people mapped:   {stats['total']}")
print(f"   With website found:    {stats['with_website']} rows")
print(f"   With email found:      {stats['with_email']} rows")
print(f"   With phone found:      {stats['with_phone']} rows")
print(f"   With CEO found:        {stats['with_ceo']} rows")
print(f"\n📁 Output: {output_path}")
print()

# Quick summary table
print(f"{'Company':30s} | {'Website':25s} | {'Email':25s} | {'CEO':20s}")
print(f"{'-'*30}-+-{'-'*25}-+-{'-'*25}-+-{'-'*20}")
for ec in enriched:
    if ec['company'] == 'UNKNOWN':
        print(f"\n--- {ec['company']}: {len(ec['people'])} people ---")
        continue
    print(f"{ec['company']:30s} | {ec['website'][:25]:25s} | {ec['email'][:25]:25s} | {ec['ceo'][:20]:20s}")