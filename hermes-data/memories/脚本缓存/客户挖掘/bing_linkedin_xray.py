#!/usr/bin/env python3
"""
Bing LinkedIn X-Ray Scraper — 不登录领英，通过Bing搜索批量获取客户资料
原理：LinkedIn封了Google索引，但Bing/Brave仍有缓存
搜索结果直接返回姓名+职位+公司+地点，足够客户初筛

用法:
  python bing_linkedin_xray.py --query "PLC SCADA" --region "South Africa" --limit 50
  python bing_linkedin_xray.py --file queries.txt --limit 30 --output leads.xlsx

输出: leads_export.xlsx (默认) 或指定文件名
"""

import argparse
import json
import logging
import re
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass, field, asdict
from datetime import datetime
from typing import Optional, List
from pathlib import Path

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("bing_xray")

# ────────────────────────────────────────────
# Data model
# ────────────────────────────────────────────

@dataclass
class LinkedInLead:
    full_name: str
    job_title: str
    company: Optional[str] = None
    location: Optional[str] = None
    connections: Optional[str] = None
    headline: Optional[str] = None
    profile_url: Optional[str] = None
    source_query: str = ""
    notes: List[str] = field(default_factory=list)


# ────────────────────────────────────────────
# Bing Search (scrape mode — no API key needed)
# ────────────────────────────────────────────

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
]


def _build_bing_url(query: str, page: int = 0) -> str:
    """Build Bing search URL with LinkedIn site restriction."""
    full_query = f"site:linkedin.com/in/ {query}"
    encoded = urllib.parse.quote(full_query)
    first = page * 10 + 1  # Bing pagination: first=1, 11, 21, ...
    return f"https://www.bing.com/search?q={encoded}&first={first}"


def _parse_bing_results(html: str) -> List[dict]:
    """Parse Bing search result items from HTML."""
    results = []
    # Each result is in <li class="b_algo">... 
    # We extract the title, URL, and snippet
    pattern = r'<li class="b_algo">(.*?)</li>'
    items = re.findall(pattern, html, re.DOTALL)
    
    for item in items:
        title_match = re.search(r'<a[^>]*href="(https://[^"]*linkedin\.com[^"]*)"[^>]*>(.*?)</a>', item, re.DOTALL)
        snippet_match = re.search(r'<p[^>]*>(.*?)</p>', item, re.DOTALL)
        
        if not title_match:
            continue
        
        url = title_match.group(1)
        title_html = title_match.group(2)
        # Clean HTML tags from title
        title = re.sub(r'<[^>]+>', '', title_html).strip()
        snippet = ""
        if snippet_match:
            snippet_html = snippet_match.group(1)
            snippet = re.sub(r'<[^>]+>', '', snippet_html).strip()
        
        # Only keep LinkedIn profile URLs (not authwall, not company pages)
        if '/in/' in url and 'authwall' not in url:
            results.append({"title": title, "url": url, "snippet": snippet})
    
    return results


def _extract_lead_from_result(result: dict, query: str) -> Optional[LinkedInLead]:
    """Extract structured lead data from a Bing search result."""
    title = result.get("title", "")
    snippet = result.get("snippet", "")
    url = result.get("url", "")
    
    # Parse: "Rochelle Vivian Goagoses - System Integrator - Andrada Mining | LinkedIn"
    # Title format: "Name - Job Title - Company | LinkedIn"  or "Name - Job Title | LinkedIn"
    
    # Remove trailing " | LinkedIn" or " | 领英"
    title_clean = re.sub(r'\s*[|]\s*(LinkedIn|领英)\s*$', '', title).strip()
    
    # Split by " - " separator
    parts = [p.strip() for p in title_clean.split(" - ")]
    
    full_name = parts[0] if len(parts) >= 1 else title_clean
    
    job_title = ""
    company = ""
    
    if len(parts) >= 3:
        # Name - Title - Company
        job_title = parts[1]
        company = parts[2]
    elif len(parts) >= 2:
        # Name - Title OR Name - Company
        # If the second part looks like a company (contains "at" or is short), it could be either
        job_title = parts[1]
        # Sometimes it's "Name at Company" in a single part
    
    location = ""
    connections = ""
    
    # Parse snippet for location and connections
    # "City of Johannesburg, Gauteng, South Africa. 3K followers 500+ connections."
    
    # Extract location (first sentence of snippet that looks like a location)
    location_match = re.search(
        r'(Location|Locality|Based in|Located in)?\s*:\s*([A-Za-z\s,]+)',
        snippet
    )
    if location_match:
        location = location_match.group(2).strip()
    
    # Also try: snippet starts with location pattern
    loc_patterns = [
        r'^([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*,\s+[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*,\s+(?:South Africa|Nigeria|Kenya|Zimbabwe|Namibia|Ghana|Egypt|Morocco|Algeria|Angola|Mozambique|Tanzania|Uganda|Ethiopia))',
        r'(City of [^\.]+)',
        r'([A-Z][a-z]+\s+Area,\s+[A-Za-z\s,]+(?:Africa|India|UAE|Dubai))',
    ]
    if not location:
        for pat in loc_patterns:
            m = re.search(pat, snippet)
            if m:
                location = m.group(1).strip()
                break
    
    # Extract connections/followers
    conn_match = re.search(r'(\d+[KkMmBb]?)\s*(?:connections?|followers?)', snippet)
    if conn_match:
        connections = conn_match.group(0).strip()
    
    # If connections not found but snippet has "connections" word
    if not connections:
        conn_match2 = re.search(r'(\d+)\+?\s*connections?', snippet)
        if conn_match2:
            connections = conn_match2.group(0).strip()
    
    # Extract headline from snippet (the part before location/connections)
    headline = snippet
    # Remove location and connection parts from headline
    if location:
        headline = headline.replace(location, "").strip()
    if connections:
        headline = headline.replace(connections, "").strip()
    # Clean up residual punctuation
    headline = re.sub(r'\s*\.\s*\.\s*', '. ', headline).strip()
    # Also resume from the various jobs listed in snippet
    # Remove generic LinkedIn boilerplate
    headline = re.sub(r'Experience?:\s*', '', headline)
    headline = re.sub(r'Education?:\s*[^.]+\.?\s*', '', headline)
    headline = re.sub(r'View mutual connections[^.]*\.', '', headline)
    headline = re.sub(r'\d+ connections on LinkedIn', '', headline)
    headline = headline.strip(' .')
    
    # Clean up URL (remove tracking params)
    clean_url = url.split('?')[0] if '?' in url else url
    
    return LinkedInLead(
        full_name=full_name,
        job_title=job_title,
        company=company,
        location=location,
        connections=connections,
        headline=headline if headline else None,
        profile_url=clean_url,
        source_query=query,
    )


def search_bing_batch(query: str, max_results: int = 50, delay: float = 2.0) -> List[LinkedInLead]:
    """Search Bing for LinkedIn profiles and extract leads."""
    leads = []
    page = 0
    seen_urls = set()
    
    while len(leads) < max_results:
        url = _build_bing_url(query, page)
        log.info(f"Bing page {page+1}: {query}")
        
        req = urllib.request.Request(
            url,
            headers={
                "User-Agent": USER_AGENTS[page % len(USER_AGENTS)],
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.5",
                "Referer": "https://www.bing.com/",
            }
        )
        
        try:
            with urllib.request.urlopen(req, timeout=15) as resp:
                html = resp.read().decode("utf-8", errors="replace")
        except Exception as e:
            log.warning(f"Bing request failed on page {page+1}: {e}")
            break
        
        results = _parse_bing_results(html)
        
        if not results:
            log.info("No more results found")
            break
        
        for r in results:
            if r["url"] in seen_urls:
                continue
            seen_urls.add(r["url"])
            lead = _extract_lead_from_result(r, query)
            if lead:
                leads.append(lead)
                log.info(f"  ✓ {lead.full_name:25s} | {lead.job_title or '—':25s} | {lead.company or '—':20s} | {lead.location or '—'}")
        
        page += 1
        time.sleep(delay)  # Be polite to Bing
        
        if len(leads) >= max_results:
            break
    
    return leads[:max_results]


# ────────────────────────────────────────────
# Export to Excel
# ────────────────────────────────────────────

def export_to_excel(leads: List[LinkedInLead], output_path: str):
    """Export leads to Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LinkedIn Leads"
        
        # Headers
        headers = [
            "Full Name", "Job Title", "Company", "Location",
            "Connections", "Headline", "Profile URL", "Source Query"
        ]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row, lead in enumerate(leads, 2):
            ws.cell(row=row, column=1, value=lead.full_name)
            ws.cell(row=row, column=2, value=lead.job_title or "")
            ws.cell(row=row, column=3, value=lead.company or "")
            ws.cell(row=row, column=4, value=lead.location or "")
            ws.cell(row=row, column=5, value=lead.connections or "")
            ws.cell(row=row, column=6, value=lead.headline or "")
            ws.cell(row=row, column=7, value=lead.profile_url or "")
            ws.cell(row=row, column=8, value=lead.source_query)
        
        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
        
        wb.save(output_path)
        log.info(f"✅ Exported {len(leads)} leads to {output_path}")
        
    except ImportError:
        # Fall back to CSV
        csv_path = output_path.replace(".xlsx", ".csv")
        import csv
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=[
                "full_name", "job_title", "company", "location",
                "connections", "headline", "profile_url", "source_query"
            ])
            writer.writeheader()
            for lead in leads:
                writer.writerow(asdict(lead))
        log.info(f"✅ Exported {len(leads)} leads to {csv_path} (openpyxl not available)")


# ────────────────────────────────────────────
# Main
# ────────────────────────────────────────────

QUERY_TEMPLATES = {
    "scada_plc": [
        '("SCADA" OR "PLC" OR "DCS") AND ("industrial automation" OR "control system")',
        '("SCADA engineer" OR "PLC programmer") AND ("Africa" OR "South Africa")',
    ],
    "system_integrator": [
        '"System Integrator" AND ("industrial automation" OR "IIoT")',
        '"System Integration" AND ("Africa" OR "Kenya" OR "Nigeria")',
    ],
    "energy_monitoring": [
        '"remote monitoring" AND ("energy" OR "solar" OR "power")',
        '"IIoT" AND ("SCADA" OR "remote monitoring") AND Africa',
    ],
    "bliiot_targets": [
        '("gateway" OR "router" OR "RTU") AND "industrial automation" AND Africa',
        '"edge computing" AND ("IIoT" OR "industrial IoT") AND Africa',
        '("PLC" OR "SCADA") AND "South Africa" AND "engineer"',
        '("remote monitoring" OR "telemetry") AND ("water" OR "oil" OR "mining") AND Africa',
        '"system integrator" AND ("South Africa" OR "Nigeria" OR "Kenya")',
        '"industrial automation" AND "Zimbabwe" OR "Zambia" OR "Namibia"',
        '("IoT" OR "IIoT") AND "industrial" AND ("Kenya" OR "Ethiopia" OR "Ghana")',
        '"Automation engineer" AND "Africa" AND ("PLC" OR "SCADA" OR "controls")',
    ],
}


def main():
    parser = argparse.ArgumentParser(description="Bing LinkedIn X-Ray Scraper")
    parser.add_argument("--query", "-q", help="Single search query")
    parser.add_argument("--file", "-f", help="File with queries (one per line)")
    parser.add_argument("--template", "-t", choices=list(QUERY_TEMPLATES.keys()),
                        help=f"Query template: {', '.join(QUERY_TEMPLATES.keys())}")
    parser.add_argument("--limit", "-l", type=int, default=50,
                        help="Max leads to collect (default: 50)")
    parser.add_argument("--delay", type=float, default=2.0,
                        help="Delay between Bing pages in seconds (default: 2.0)")
    parser.add_argument("--output", "-o", default="leads_export.xlsx",
                        help="Output file (default: leads_export.xlsx)")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="Verbose output")
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Build query list
    queries = []
    if args.query:
        queries.append(args.query)
    if args.file:
        with open(args.file) as f:
            queries.extend(line.strip() for line in f if line.strip())
    if args.template:
        queries.extend(QUERY_TEMPLATES[args.template])
    if not queries:
        # Default: use BLIIOT targets
        log.info("No query specified, using BLIIOT target template")
        queries = QUERY_TEMPLATES["bliiot_targets"]
    
    # Deduplicate
    queries = list(dict.fromkeys(queries))
    
    log.info(f"🚀 Bing LinkedIn X-Ray Scraper")
    log.info(f"   Queries: {len(queries)}")
    log.info(f"   Max leads: {args.limit}")
    log.info(f"   Delay: {args.delay}s")
    log.info("")
    
    all_leads = []
    seen_urls = set()
    
    for i, query in enumerate(queries, 1):
        log.info(f"[{i}/{len(queries)}] Searching: {query}")
        leads = search_bing_batch(
            query,
            max_results=max(10, args.limit // len(queries) + 5),
            delay=args.delay,
        )
        
        for lead in leads:
            if lead.profile_url and lead.profile_url not in seen_urls:
                seen_urls.add(lead.profile_url)
                all_leads.append(lead)
        
        log.info(f"   → Got {len(leads)} leads, total unique: {len(all_leads)}")
        
        # Smaller delay between queries
        if i < len(queries):
            time.sleep(args.delay * 0.5)
    
    # Trim to limit
    all_leads = all_leads[:args.limit]
    
    log.info(f"\n{'='*60}")
    log.info(f"📊 Total: {len(all_leads)} unique leads")
    log.info(f"{'='*60}")
    
    # Export
    output_path = args.output
    export_to_excel(all_leads, output_path)
    
    # Summary by location
    locations = {}
    for lead in all_leads:
        loc = lead.location or "Unknown"
        locations[loc] = locations.get(loc, 0) + 1
    log.info("\n📌 By Location:")
    for loc, count in sorted(locations.items(), key=lambda x: -x[1]):
        log.info(f"   {loc:40s} {count}")
    
    # Summary by company
    companies = {}
    for lead in all_leads:
        co = lead.company or "Unknown"
        companies[co] = companies.get(co, 0) + 1
    log.info("\n🏢 By Company:")
    top_companies = sorted(companies.items(), key=lambda x: -x[1])[:15]
    for co, count in top_companies:
        log.info(f"   {co:40s} {count}")


if __name__ == "__main__":
    main()