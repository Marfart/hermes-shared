#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
WhatsApp Web CDP 极速群发脚本 — 通过 Chrome DevTools Protocol 直接 JS 操作
无需 Selenium，一次 JS 注入完成全部操作！
"""
import json, time, os, subprocess, sys, re, urllib.request, socket
from datetime import datetime

import openpyxl

# ========== CONFIG ==========
CHROME_PATH = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
USER_DATA_DIR = r"C:\Users\Admin\AppData\Local\Google\Chrome\User Data"
PROFILE = "Default"
CDP_PORT = 9222
LEADS_FILE = r"C:\Users\Admin\Desktop\Working\BLIIoT_Customer_Leads_v2.xlsx"
LOG_DIR = r"C:\Users\Admin\AppData\Local\hermes\memories\脚本缓存\whatsapp_bot"
STATUS_FILE = os.path.join(LOG_DIR, "progress.json")
LOG_FILE = os.path.join(LOG_DIR, "send_log.json")
MIN_DELAY = 120
MAX_DELAY = 300
MAX_SEND = 30

# ========== TEMPLATES ==========
TEMPLATES = [
    (["system integrator", "automation company"],
     """Hi {company},

I'm reaching out from BLIIoT (bliiot.com), an industrial IoT manufacturer based in Shenzhen, China.

We specialize in products perfect for system integrators:

• ARMxy Industrial ARM Edge Controllers (BL330-BL450, up to 6TOPS AI)
• IOy Remote I/O Modules (6 protocols: Modbus/OPC UA/MQTT/BACnet/IEC104/SNMP)
• BL116 High-Performance IIoT Gateway (10,000 data points in 1.5s)
• EdgePLC BL234 (CODESYS + EtherCAT)

Our products are already trusted in SmartGrid projects across Africa. As a system integrator, adding BLIIoT hardware to your solutions can reduce costs and expand capabilities.

Would you be interested in our complete product catalog and pricing?

Best regards,
Kali
BLIIoT | Shenzhen Beilai Technology Co., Ltd"""),

    (["building", "hvac", "bacnet"],
     """Hi {company},

I'm from BLIIoT (bliiot.com), an industrial IoT manufacturer with BACnet-certified solutions for building automation. Our BA116 HVAC Edge Gateway and BA190 BACnet IP Remote I/O Module are ideal for smart building projects in {country}.

Would you be interested in our building automation catalog?

Best regards,
Kali
BLIIoT"""),

    (["energy", "solar", "power", "substation", "smartgrid", "grid"],
     """Hi {company},

I'm from BLIIoT (bliiot.com). We manufacture SmartGrid gateways and energy monitoring solutions including BE116 Smart Grid Edge Gateway (IEC104/IEC61850) and ARMxy controllers for EMS/BMS. Our products are trusted in substation monitoring and solar energy management across Africa.

Interested in seeing our energy product range?

Best regards,
Kali
BLIIoT"""),

    (["distributor", "distribution", "supplier"],
     """Hi {company},

I'm from BLIIoT - we're looking for distribution partners in {country}. Our lines include ARMxy controllers, IOy remote I/O modules, IIoT gateways, 4G routers, switches and isolators. We offer competitive pricing and technical support for partners.

Interested in a distribution partnership?

Best regards,
Kali
BLIIoT"""),
]

DEFAULT_TEMPLATE = """Hi {company},

I'm reaching out from BLIIoT (bliiot.com), an industrial IoT manufacturer based in Shenzhen, China.

We specialize in ARMxy edge controllers, remote I/O modules (Modbus/OPC UA/MQTT/BACnet), IoT gateways, and 4G industrial routers - a complete IIoT product line for industrial automation and SmartGrid applications.

Our products are already used in SmartGrid projects across Africa.

Would you be interested in receiving our product catalog and pricing?

Best regards,
Kali
BLIIoT | Shenzhen Beilai Technology Co., Ltd"""

TEMPLATES.append(([""], DEFAULT_TEMPLATE))


def choose_template(category, company, country):
    if not category: category = ""
    cat_lower = category.lower()
    for keywords, template in TEMPLATES:
        for kw in keywords:
            if kw and kw.lower() in cat_lower:
                return template.format(company=company, country=country)
    return DEFAULT_TEMPLATE.format(company=company, country=country)


def load_leads():
    wb = openpyxl.load_workbook(LEADS_FILE, data_only=True)
    ws = wb.active
    col_map = {'name': 2, 'phone': 5, 'country': 9, 'category': 10}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v:
            vs = str(v).lower()
            if 'company' in vs: col_map['name'] = col
            if 'phone' in vs: col_map['phone'] = col
            if 'country' in vs: col_map['country'] = col
            if 'categor' in vs: col_map['category'] = col

    leads = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=col_map['name']).value
        phone = ws.cell(row=row, column=col_map['phone']).value
        if name and phone:
            phone_str = str(phone).strip()
            phone_clean = re.sub(r'[\s\-\(\)]', '', phone_str)
            leads.append({
                "name": str(name).strip(),
                "phone": phone_str,
                "phone_clean": phone_clean,
                "country": str(ws.cell(row=row, column=col_map.get('country', 9)).value or "").strip(),
                "category": str(ws.cell(row=row, column=col_map.get('category', 10)).value or "").strip(),
            })
    wb.close()
    return leads


# ========== CDP FUNCTIONS ==========
def cdp_send(ws_url, method, params=None):
    """Send CDP command and get response"""
    if params is None: params = {}
    import asyncio, websockets
    async def _send():
        async with websockets.connect(ws_url) as ws:
            msg_id = int(time.time() * 1000) % 100000
            req = json.dumps({"id": msg_id, "method": method, "params": params})
            await ws.send(req)
            resp = await ws.recv()
            return json.loads(resp)
    return asyncio.run(_send())


def cdp_call(ws_url, method, params=None):
    """Execute CDP and return result"""
    result = cdp_send(ws_url, method, params)
    if "result" in result:
        return result["result"]
    elif "error" in result:
        raise Exception(f"CDP error: {result['error']}")
    return result


def launch_chrome():
    """Kill existing Chrome and launch with remote debugging"""
    subprocess.run("taskkill /F /IM chrome.exe 2>nul", shell=True)
    time.sleep(2)

    # Remove SingletonLock
    for f in ["SingletonLock", "SingletonSocket", "SingletonCookie"]:
        p = os.path.join(USER_DATA_DIR, f)
        try: os.remove(p)
        except: pass

    proc = subprocess.Popen([
        CHROME_PATH,
        f"--remote-debugging-port={CDP_PORT}",
        f"--user-data-dir={USER_DATA_DIR}",
        f"--profile-directory={PROFILE}",
        "--no-first-run",
        "--disable-blink-features=AutomationControlled",
        "--window-size=1200,800",
        "https://web.whatsapp.com"
    ])
    print(f"🚀 Chrome launched (PID: {proc.pid})")
    print("⏳ Waiting for CDP port...")

    # Wait for CDP
    for i in range(30):
        time.sleep(2)
        try:
            r = urllib.request.urlopen(f"http://127.0.0.1:{CDP_PORT}/json", timeout=3)
            tabs = json.loads(r.read().decode())
            print(f"✅ CDP ready! Tabs: {len(tabs)}")
            return tabs
        except:
            if i % 5 == 0:
                print(f"  ⏳ Waiting... ({i*2+2}s)")
    raise Exception("Chrome CDP port not available after 60s")


def get_tab(tabs):
    """Find WhatsApp tab"""
    for t in tabs:
        if "whatsapp" in t.get("url", "").lower() or "whatsapp" in t.get("title", "").lower():
            return t
    # Otherwise just use the first tab pointing to web.whatsapp.com
    for t in tabs:
        if "web.whatsapp" in t.get("url", ""):
            return t
    return tabs[0] if tabs else None


def wait_login(ws_url):
    """Wait for WhatsApp Web to be logged in"""
    print("⏳ Checking WhatsApp login status...")
    for i in range(90):
        try:
            result = cdp_call(ws_url, "Runtime.evaluate", {
                "expression": """
                    (() => {
                        const sidebar = document.querySelector('#side');
                        if (sidebar) {
                            const searchBox = document.querySelector('[contenteditable="true"]');
                            if (searchBox) return 'logged_in';
                        }
                        const qr = document.querySelector('canvas');
                        if (qr) return 'qr_needed';
                        return 'loading';
                    })()
                """,
                "returnByValue": True
            })
            status = result.get("result", {}).get("value", "loading")
            if status == "logged_in":
                print("✅ WhatsApp Web ready!")
                return True
            elif status == "qr_needed":
                if i == 0:
                    print("📱 QR Code detected! Please scan with your phone...")
            if i % 10 == 0 and i > 0:
                print(f"  ⏳ Still waiting ({i}s)...")
        except:
            pass
        time.sleep(1)
    print("❌ Login timeout")
    return False


def send_whatsapp(ws_url, phone_clean, message):
    """Send a WhatsApp message via JS"""
    safe_msg = json.dumps(message)
    js = f"""
    (async () => {{
        const phone = "{phone_clean}";
        const msg = {safe_msg};

        // Click new chat button
        const newChatBtns = document.querySelectorAll('[data-testid="chat-new"]');
        if (newChatBtns.length) newChatBtns[0].click();
        else {{
            const btns = document.querySelectorAll('[title="新聊天"], [title="New chat"]');
            if (btns.length) btns[0].click();
        }}
        await new Promise(r => setTimeout(r, 1000));

        // Find search input
        const searchInputs = document.querySelectorAll('[contenteditable="true"]');
        let searchBox = null;
        for (const inp of searchInputs) {{
            if (inp.parentElement && (inp.parentElement.getAttribute('data-testid') === 'chat-list-search' ||
                inp.getAttribute('data-tab') === '3')) {{
                searchBox = inp;
                break;
            }}
        }}
        if (!searchBox) {{
            // Fallback - find the first visible contenteditable
            for (const inp of searchInputs) {{
                if (inp.offsetParent !== null) {{
                    searchBox = inp;
                    break;
                }}
            }}
        }}
        if (!searchBox) return {{status: 'error', msg: 'No search box found'}};

        // Clear and type phone
        searchBox.focus();
        searchBox.innerHTML = '';
        document.execCommand('insertText', false, phone);
        await new Promise(r => setTimeout(r, 2000));

        // Click contact in results
        const contactItems = document.querySelectorAll('[data-testid="cell-frame-container"], ._ak8q, [role="listitem"]');
        for (const item of contactItems) {{
            if (item.offsetParent !== null) {{
                item.click();
                break;
            }}
        }}
        await new Promise(r => setTimeout(r, 1500));

        // Check if we're in a chat now - find message input
        const msgBoxes = document.querySelectorAll('[contenteditable="true"]');
        let msgBox = null;
        for (const inp of msgBoxes) {{
            if (inp !== searchBox && inp.offsetParent !== null) {{
                msgBox = inp;
                break;
            }}
        }}

        if (!msgBox) {{
            // Maybe contact not found - try to click the search result again
            return {{status: 'not_found', msg: 'Contact not on WhatsApp'}};
        }}

        // Type message
        msgBox.focus();
        msgBox.innerHTML = '';
        const lines = msg.split('\\n');
        for (let i = 0; i < lines.length; i++) {{
            document.execCommand('insertText', false, lines[i]);
            if (i < lines.length - 1) {{
                // Shift+Enter for new line
                document.execCommand('insertLineBreak', false);
            }}
        }}
        await new Promise(r => setTimeout(r, 500));

        // Find and click send button
        const sendBtn = document.querySelector('[data-testid="send"]') ||
                        document.querySelector('[data-icon="send"]') ||
                        document.querySelector('button[data-tab="11"]') ||
                        document.querySelector('span[data-icon="send"]');
        if (sendBtn) {{
            sendBtn.click();
            return {{status: 'sent'}};
        }}

        // Try Enter key as fallback
        msgBox.dispatchEvent(new KeyboardEvent('keydown', {{key: 'Enter', keyCode: 13, which: 13, ctrlKey: false}}));
        return {{status: 'sent_enter'}};
    }})()
    """

    result = cdp_call(ws_url, "Runtime.evaluate", {
        "expression": js,
        "returnByValue": True,
        "awaitPromise": True,
        "timeout": 15000
    })
    return result.get("result", {}).get("value", {})


def load_progress():
    if os.path.exists(STATUS_FILE):
        try:
            with open(STATUS_FILE, 'r') as f:
                return json.load(f)
        except:
            return {"sent_indices": []}
    return {"sent_indices": []}


def save_progress(p):
    with open(STATUS_FILE, 'w') as f:
        json.dump(p, f, ensure_ascii=False, indent=2)


def append_log(entry):
    logs = []
    if os.path.exists(LOG_FILE):
        try:
            with open(LOG_FILE, 'r') as f:
                logs = json.load(f)
        except:
            pass
    logs.append(entry)
    if len(logs) > 200:
        logs = logs[-200:]
    with open(LOG_FILE, 'w') as f:
        json.dump(logs, f, ensure_ascii=False, indent=2)


def main():
    print("=" * 55)
    print("📱 WhatsApp CDP 极速群发 v2.0 ⚡")
    print("=" * 55)

    # Load leads
    leads = load_leads()
    if not leads:
        print("❌ No leads in Excel!")
        return
    print(f"📋 Loaded {len(leads)} leads")

    # Check progress
    progress = load_progress()
    sent = set(progress.get("sent_indices", []))
    remaining = [i for i in range(len(leads)) if i not in sent]
    if not remaining:
        print("✅ All done!")
        return

    to_send = remaining[:MAX_SEND]
    print(f"📊 Sent: {len(sent)} | This session: {len(to_send)} | Remaining: {len(remaining)}")

    # Launch Chrome
    print("\n🔧 Launching Chrome with CDP...")
    tabs = launch_chrome()
    tab = get_tab(tabs)
    if not tab:
        print("❌ No tab found")
        return

    ws_url = tab["webSocketDebuggerUrl"]
    print(f"🔗 Connected to CDP: {tab.get('title', '?')[:40]}")

    # Wait for login
    if not wait_login(ws_url):
        return

    # Loop through leads
    success = 0
    failed = 0
    try:
        for idx in to_send:
            lead = leads[idx]
            msg = choose_template(lead['category'], lead['name'], lead['country'])

            print(f"\n[{success+failed+1}/{len(to_send)}] {lead['name']}")
            print(f"  📞 {lead['phone']} | {lead['country']}")

            result = send_whatsapp(ws_url, lead['phone_clean'], msg)
            status = result.get("status", "unknown")

            if status in ("sent", "sent_enter"):
                success += 1
                sent.add(idx)
                print(f"  ✅ SENT!")
                append_log({"company": lead['name'], "phone": lead['phone'],
                           "sent_at": datetime.now().isoformat(), "status": "success",
                           "country": lead['country']})
            elif status == "not_found":
                print(f"  ⚠️ Not on WhatsApp")
                # Still mark as sent (skip)
                sent.add(idx)
                failed += 1
                append_log({"company": lead['name'], "phone": lead['phone'],
                           "sent_at": datetime.now().isoformat(), "status": "not_found",
                           "country": lead['country']})
            else:
                print(f"  ❌ Error: {result.get('msg', status)}")
                failed += 1
                append_log({"company": lead['name'], "phone": lead['phone'],
                           "sent_at": datetime.now().isoformat(), "status": f"error: {result.get('msg', '')[:50]}",
                           "country": lead['country']})

            progress["sent_indices"] = sorted(list(sent))
            save_progress(progress)

            if idx != to_send[-1]:
                delay = int(time.time() % (MAX_DELAY - MIN_DELAY) + MIN_DELAY)
                print(f"  ⏳ Waiting {delay//60}m{delay%60}s...")
                for s in range(delay, 0, -10):
                    if s % 30 == 0 or s <= 10:
                        print(f"     {s}s...", end='\r')
                    time.sleep(10)
                print()

    except KeyboardInterrupt:
        print("\n⏹ Stopped")

    print(f"\n{'='*55}")
    print(f"📊 Done: ✅ {success} sent | ❌ {failed} failed | 📊 {len(sent)}/{len(leads)} total")
    progress["sent_indices"] = sorted(list(sent))
    progress["completed"] = len(sent) >= len(leads)
    save_progress(progress)
    print("✅ Session saved. Chrome stays open.")
    print("📝 Check the WhatsApp window on your screen!")


if __name__ == "__main__":
    import asyncio
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    main()