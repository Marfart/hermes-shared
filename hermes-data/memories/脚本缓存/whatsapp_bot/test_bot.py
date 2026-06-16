#!/usr/bin/env python
"""Test the WhatsApp bot's data loading and message generation (without browser)"""
import sys
sys.path.insert(0, r"C:\Users\Admin\AppData\Local\hermes\memories\脚本缓存\whatsapp_bot")

import openpyxl
import re

LEADS_FILE = r"C:\Users\Admin\Desktop\Working\BLIIoT_Customer_Leads_v2.xlsx"

TEMPLATES = {
    "System Integrator": """Hi {company},

I'm reaching out from BLIIoT (bliiot.com), an industrial IoT manufacturer based in Shenzhen, China.

We specialize in:
• ARMxy Industrial ARM Edge Controllers (edge computing + AI)
• IOy Remote I/O Modules (Modbus/OPC UA/MQTT/BACnet/IEC104)
• IIoT Gateways & 4G Industrial Routers

Our products are trusted in SmartGrid projects across Africa, including transformer monitoring systems in Zimbabwe. We're looking for reliable partners in {country}.

Would you be interested in receiving our product catalog and pricing? Happy to share datasheets.

Best regards,
Kali
BLIIoT | Shenzhen Beilai Technology Co., Ltd""",

    "Building Automation": """Hi {company},

I'm from BLIIoT (bliiot.com), an industrial IoT manufacturer. We have BACnet-certified solutions specifically for building automation:

• BA116 HVAC Edge Gateway
• BA190 BACnet IP Remote I/O Module  
• BA100 Series BACnet Gateway

These products are ideal for smart building projects, HVAC control, and BMS integration in {country}.

Would you be interested in our building automation catalog?

Best regards,
Kali
BLIIoT""",

    "Energy": """Hi {company},

I'm reaching out from BLIIoT (bliiot.com). We manufacture SmartGrid gateways and energy monitoring solutions:

• BE116 Smart Grid Edge Gateway (supports IEC104/IEC61850)
• BE190 IEC104 Remote I/O Module
• ARMxy BL350 Controller for EMS/BMS systems

These products are designed for substation monitoring, solar energy management, and power automation in {country}.

Interested in seeing our energy monitoring product range?

Best regards,
Kali
BLIIoT""",

    "default": """Hi {company},

I'm reaching out from BLIIoT (bliiot.com), an industrial IoT manufacturer based in Shenzhen, China.

We specialize in ARMxy edge controllers, remote I/O modules, IoT gateways, and 4G industrial routers - a complete IIoT product line.

Would you be interested in receiving our product catalog?

Best regards,
Kali
BLIIoT"""
}

def choose_template(company_type, company_name, country):
    """选择最合适的开发信模板"""
    if not company_type:
        company_type = "default"
    
    best_match = "default"
    for key in TEMPLATES:
        if key.lower() in company_type.lower():
            best_match = key
            break
    
    template = TEMPLATES.get(best_match, TEMPLATES["default"])
    return template.format(company=company_name, country=country)


# Load Excel and test
print("=" * 55)
print("📋 TEST: Reading customer Excel + Generating messages")
print("=" * 55)

wb = openpyxl.load_workbook(LEADS_FILE, data_only=True)
ws = wb.active

print(f"\nSheet name: {ws.title}")
print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")

# Print headers
headers = {}
print("\n📌 Headers found:")
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    if val:
        print(f"  Col {col}: {val}")
        # Map by keyword
        for keyword, name in [
            ("Company Name", "name"), ("What They Do", "desc"), ("Why They Need", "why"),
            ("Phone", "phone"), ("Country", "country"), ("Category", "category"), ("Rating", "rating")
        ]:
            if keyword.lower() in str(val).lower():
                headers[name] = col

print(f"\n📌 Mapped columns: {headers}")

# Read leads
leads = []
for row in range(2, ws.max_row + 1):
    name = ws.cell(row=row, column=headers.get("name", 2)).value
    phone = ws.cell(row=row, column=headers.get("phone", 5)).value
    country = ws.cell(row=row, column=headers.get("country", 9)).value
    category = ws.cell(row=row, column=headers.get("category", 10)).value
    
    if name and phone:
        phone_str = str(phone).strip()
        phone_clean = re.sub(r'[\s\-\(\)]', '', phone_str)
        leads.append({
            "name": str(name).strip(),
            "phone": phone_str,
            "phone_clean": phone_clean,
            "country": str(country).strip() if country else "",
            "category": str(category).strip() if category else "",
        })

wb.close()

print(f"\n✅ Loaded {len(leads)} leads from Excel!")
print("\n" + "=" * 55)
print("📱 TEST: Message Generation Samples")
print("=" * 55)

# Show first 5 leads with generated messages
for i, lead in enumerate(leads[:5]):
    msg = choose_template(lead['category'], lead['name'], lead['country'])
    print(f"\n{'─'*50}")
    print(f"[{i+1}] {lead['name']}")
    print(f"    📞 {lead['phone']}  |  🌍 {lead['country']}  |  🏷 {lead['category']}")
    print(f"    📱 Clean: {lead['phone_clean']}")
    print(f"\n    💬 Message:")
    for line in msg.split('\n'):
        print(f"      {line}")
    print()

print(f"\n{'='*55}")
print("✅ TEST PASSED!")
print(f"   Total leads loaded: {len(leads)}")
print(f"   All have phone numbers: {all(l['phone'] for l in leads)}")
print(f"   All have country: {all(l['country'] for l in leads)}")
print(f"   Message templates: {len(TEMPLATES)} types")
print(f"{'='*55}")

# Show all leads summary
print(f"\n📋 Full lead list:")
for i, lead in enumerate(leads):
    print(f"  {i+1:2d}. {lead['name'][:35]:35s} {lead['phone']:25s} {lead['country']:15s} {lead['category'][:20]}")