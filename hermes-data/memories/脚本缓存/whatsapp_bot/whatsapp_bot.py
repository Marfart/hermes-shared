#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
BLIIoT WhatsApp Auto Outreach Bot v1.0
小马 WhatsApp 自动发开发信工具

功能：
1. 读取客户名单 Excel（公司名、电话、国家、推荐产品）
2. 打开 WhatsApp Web（首次需扫码，之后自动登录）
3. 逐个发送个性化开发信
4. 随机延迟（2-5分钟）防封号
5. 记录发送状态，中断后可继续

使用方法：
  python whatsapp_bot.py          # 开始发送
  python whatsapp_bot.py status   # 查看进度
  python whatsapp_bot.py reset    # 重置进度

⚠️ 每天建议不超过 20-30 条
⚠️ 每条间隔至少 2 分钟
⚠️ 建议先发 5 条测试
"""

import os, sys, json, time, random, re
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# ==================== CONFIG ====================
LEADS_FILE = r"C:\Users\Admin\Desktop\Working\BLIIoT_Customer_Leads_v2.xlsx"
CHROME_USER_DATA = r"C:\Users\Admin\AppData\Local\Google\Chrome\User Data"
CHROME_PROFILE = "Default"
LOG_DIR = r"C:\Users\Admin\AppData\Local\hermes\memories\脚本缓存\whatsapp_bot"
os.makedirs(LOG_DIR, exist_ok=True)
LOG_FILE = os.path.join(LOG_DIR, "send_log.json")
STATUS_FILE = os.path.join(LOG_DIR, "progress.json")
MIN_DELAY = 120
MAX_DELAY = 300
MAX_SEND_PER_SESSION = 30

# ==================== TEMPLATES ====================
# (keywords_list, template_text)

TEMPLATE_RULES = [
    (["Machine Builder", "Machine", "Machinery", "Robot", "AGV"],
"""Hi {company},

I'm from BLIIoT (bliiot.com). We manufacture ARMxy industrial controllers - cost-effective alternatives to traditional PLCs for machine builders.

Key products for machinery:
• ARMxy BL350 (TI AM62x, ideal for motion control)
• EdgePLC BL234 (supports CODESYS + EtherCAT real-time control)
• ARMxy BL450 (RK3588, 6TOPS NPU for AI visual inspection)

Would you be interested in exploring our machine control solutions? Happy to share datasheets.

Best regards,
Kali
BLIIoT"""),

    (["Building Automation", "Home Automation", "Building", "BACnet", "HVAC"],
"""Hi {company},

I represent BLIIoT (bliiot.com), an IoT manufacturer with BACnet-certified solutions for building automation:

• BA116 HVAC Edge Gateway (BACnet/IP, 10,000 data points)
• BA190 BACnet IP Remote I/O Module
• BL191 OPC UA IO Module for smart building integration

These products are ideal for smart building projects, HVAC control, and BMS integration in {country}.

Would you be interested in our building automation catalog?

Best regards,
Kali
BLIIoT"""),

    (["Automation Giant", "Rockwell"],
"""Hi {company},

I'm from BLIIoT (bliiot.com), an industrial IoT manufacturer in Shenzhen.

Our IOy remote I/O modules (OPC UA/MQTT/BACnet) can integrate seamlessly with Rockwell PLCs as remote IO extensions - we complement your ecosystem, not compete with it.

Would you be interested in exploring how BLIIoT products can add value to your customers' solutions?

Best regards,
Kali
BLIIoT"""),

    (["Energy", "Solar", "Power", "Substation", "Renewable", "Grid", "SmartGrid"],
"""Hi {company},

I'm from BLIIoT (bliiot.com). We manufacture SmartGrid gateways and energy monitoring solutions:

• BE116 Smart Grid Edge Gateway (IEC104/IEC61850)
• BE190 IEC104 Remote I/O Module
• ARMxy BL350 Controller (designed for EMS/BMS systems)

These are trusted in SmartGrid substation monitoring, solar energy management, and power automation projects across Africa.

Interested in seeing our energy monitoring product range?

Best regards,
Kali
BLIIoT"""),

    (["Distribution", "Distributor", "Supplier"],
"""Hi {company},

I'm from BLIIoT (bliiot.com) - we're looking for distribution partners in {country}.

Our product line:
• ARMxy Industrial ARM Computers
• IOy Remote I/O Modules (6 protocols: Modbus/OPC UA/MQTT/BACnet/IEC104/SNMP)
• IIoT Gateways & 4G Routers
• Industrial Ethernet Switches
• Signal Isolators

We offer competitive pricing, technical training, and marketing support for partners.

Interested in a distribution partnership?

Best regards,
Kali
BLIIoT"""),

    (["Electrical Engineer", "Engineering Solution", "Control System"],
"""Hi {company},

I'm from BLIIoT (bliiot.com), an industrial IoT manufacturer based in Shenzhen.

Our products are great for electrical engineering projects:
• IOy BL190 series - remote I/O expansion for PLCs (Modbus/OPC UA)
• ARMxy BL330 - cost-effective edge controller (from $100)
• R40 4G Router - dual SIM redundancy for reliable connectivity

Would you like to see how these can support your engineering projects in {country}?

Best regards,
Kali
BLIIoT"""),

    (["Process Automation"],
"""Hi {company},

I'm from BLIIoT (bliiot.com). We manufacture products for process automation:

• BL191 OPC UA IO Module - direct SCADA/MES integration
• IOy series - wide temperature -45~80°C, electrical isolation
• ARMxy BL350 - reliable controller for critical processes

Our products pass EMC/EMI testing and are built for harsh industrial environments.

Would you be interested in our product catalog?

Best regards,
Kali
BLIIoT"""),

    (["Software", "IT", "Tech"],
"""Hi {company},

I'm from BLIIoT (bliiot.com). Our hardware + your software = complete solution for your clients.

Our ARMxy edge controllers run Ubuntu/Docker, support Python/C++ development, and come with BLIoTLink protocol conversion software.

Perfect for software companies looking to add IoT hardware to their solutions without building from scratch.

Would you be interested in an OEM partnership?

Best regards,
Kali
BLIIoT"""),

    (["Equipment Supplier"],
"""Hi {company},

I'm from BLIIoT (bliiot.com). We're looking for equipment suppliers to partner with in {country}.

Our products would complement your existing lines:
• ARMxy Industrial ARM Computers
• IOy Remote I/O Modules
• IIoT Gateways & 4G Routers
• Industrial Ethernet Switches

We offer good margins and technical support for partners.

Interested in adding BLIIoT products to your catalog?

Best regards,
Kali
BLIIoT"""),

    (["System Integrator", "System Integrator / Automation", "Automation Company"],
"""Hi {company},

I'm reaching out from BLIIoT (bliiot.com), an industrial IoT manufacturer based in Shenzhen, China.

We specialize in products perfect for system integrators:

• ARMxy Industrial ARM Edge Controllers (BL330-BL450, up to 6TOPS AI)
• IOy Remote I/O Modules (6 protocols: Modbus/OPC UA/MQTT/BACnet/IEC104/SNMP)
• BL116 High-Performance IIoT Gateway (10,000 data points in 1.5s)
• EdgePLC BL234 (CODESYS + EtherCAT)

Our products are already trusted in SmartGrid projects across Africa. As a system integrator, adding BLIIoT hardware to your solutions can reduce costs and expand capabilities.

Would you be interested in our complete product catalog and pricing?

Best regards,
Kali
BLIIoT | Shenzhen Beilai Technology Co., Ltd"""),

    (["Existing", "EXISTING"],
"""Hi {company},

This is Kali from BLIIoT! Hope everything is going well with our ongoing project.

I wanted to let you know we have expanded our product line since we last worked together:

• New ARMxy BL450/BL440 series with AI NPU (1-6 TOPS)
• IOy series now supports SNMP and IEC104 protocols
• R40A/R40B edge control terminal with DI/DO/AI

Would you be interested in seeing the new products? We'd love to support your next phase.

Best regards,
Kali
BLIIoT"""),
]

DEFAULT_TEMPLATE = """Hi {company},

I'm reaching out from BLIIoT (bliiot.com), an industrial IoT manufacturer based in Shenzhen, China.

We specialize in ARMxy edge controllers, remote I/O modules (Modbus/OPC UA/MQTT/BACnet), IoT gateways, and 4G industrial routers - a complete IIoT product line for industrial automation and SmartGrid applications.

Our products are already used in SmartGrid projects across Africa.

Would you be interested in receiving our product catalog and pricing?

Best regards,
Kali
BLIIoT | Shenzhen Beilai Technology Co., Ltd"""


def choose_template(company_type, company_name, country):
    """Choose best template based on company type keywords"""
    if not company_type:
        company_type = ""
    ct_lower = company_type.lower()
    for keywords, template in TEMPLATE_RULES:
        for kw in keywords:
            if kw.lower() in ct_lower:
                return template.format(company=company_name, country=country)
    return DEFAULT_TEMPLATE.format(company=company_name, country=country)


def load_leads():
    """Load leads from Excel"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(LEADS_FILE, data_only=True)
        ws = wb.active
        
        # Find columns
        col_map = {'name': 2, 'desc': 3, 'why': 4, 'phone': 5, 'country': 9, 'category': 10, 'rating': 11}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                vs = str(val).lower()
                if 'company' in vs: col_map['name'] = col
                if 'phone' in vs: col_map['phone'] = col
                if 'country' in vs: col_map['country'] = col
                if 'categor' in vs: col_map['category'] = col
                if 'rating' in vs or 'score' in vs: col_map['rating'] = col
        
        leads = []
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=col_map['name']).value
            phone = ws.cell(row=row, column=col_map['phone']).value
            if name and phone:
                phone_str = str(phone).strip()
                phone_clean = re.sub(r'[\s\-\(\)]', '', phone_str)
                leads.append({
                    "name": str(name).strip(),
                    "phone": phone_str,
                    "phone_clean": phone_clean,
                    "country": str(ws.cell(row=row, column=col_map.get('country', 9)).value or "").strip(),
                    "category": str(ws.cell(row=row, column=col_map.get('category', 10)).value or "").strip(),
                })
        
        wb.close()
        return leads
    except Exception as e:
        print(f"❌ Excel read error: {e}")
        return []


def load_progress():
    if os.path.exists(STATUS_FILE):
        try:
            with open(STATUS_FILE, 'r') as f:
                return json.load(f)
        except:
            return {"sent_indices": []}
    return {"sent_indices": []}


def save_progress(p):
    with open(STATUS_FILE, 'w') as f:
        json.dump(p, f, ensure_ascii=False, indent=2)


def append_log(entry):
    logs = []
    if os.path.exists(LOG_FILE):
        try:
            with open(LOG_FILE, 'r') as f:
                logs = json.load(f)
        except:
            pass
    logs.append(entry)
    if len(logs) > 200:
        logs = logs[-200:]
    with open(LOG_FILE, 'w') as f:
        json.dump(logs, f, ensure_ascii=False, indent=2)


def send_messages():
    print("=" * 55)
    print("📱 BLIIoT WhatsApp Auto Outreach Bot v1.0")
    print("=" * 55)
    
    leads = load_leads()
    if not leads:
        print("❌ No leads found in Excel!")
        return
    print(f"📋 Loaded {len(leads)} leads")
    
    progress = load_progress()
    sent = set(progress.get("sent_indices", []))
    remaining = [i for i in range(len(leads)) if i not in sent]
    
    if not remaining:
        print("✅ All leads already sent!")
        return
    
    to_send = remaining[:MAX_SEND_PER_SESSION]
    print(f"📊 Already sent: {len(sent)} | This session: {len(to_send)} | Remaining: {len(remaining)}")
    
    # Start Chrome
    print("\n🔧 Launching Chrome...")
    print("📌 First time? Scan QR code on WhatsApp Web")
    print("💡 Session saves - no need to scan again\n")
    
    opts = Options()
    opts.add_argument(f"--user-data-dir={CHROME_USER_DATA}")
    opts.add_argument(f"--profile-directory={CHROME_PROFILE}")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    
    driver = webdriver.Chrome(options=opts)
    driver.get("https://web.whatsapp.com")
    
    # Wait for login
    print("⏳ Waiting for WhatsApp Web...")
    try:
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.XPATH, '//div[@contenteditable="true"][@data-tab="3"]'))
        )
        print("✅ WhatsApp Web ready!\n")
    except TimeoutException:
        print("⏳ QR code scan needed...")
        try:
            WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.XPATH, '//div[@contenteditable="true"][@data-tab="3"]'))
            )
            print("✅ WhatsApp Web ready!\n")
        except TimeoutException:
            print("❌ Login timeout. Please scan QR and retry")
            driver.quit()
            return
    
    success = 0
    failed = 0
    
    try:
        for idx in to_send:
            lead = leads[idx]
            msg = choose_template(lead['category'], lead['name'], lead['country'])
            
            print(f"\n[{success+failed+1}/{len(to_send)}] {lead['name']}")
            print(f"  📞 {lead['phone']} | 🌍 {lead['country']} | 🏷 {lead['category']}")
            print(f"  💬 Preview: {msg[:80]}...")
            
            try:
                # New chat button
                nc = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//div[@title="新聊天"]'))
                )
                nc.click()
                time.sleep(1)
                
                # Search contact
                sb = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//div[@contenteditable="true"][@data-tab="3"]'))
                )
                sb.clear()
                sb.send_keys(lead['phone_clean'])
                time.sleep(2)
                
                try:
                    # Click result
                    contact = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, '//div[@class="_ak8q"]'))
                    )
                    contact.click()
                    time.sleep(1.5)
                    
                    # Type message
                    mb = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, '//div[@contenteditable="true"][@data-tab="10"]'))
                    )
                    mb.click()
                    time.sleep(0.5)
                    for line in msg.split('\n'):
                        mb.send_keys(line)
                        mb.send_keys(Keys.SHIFT + Keys.ENTER)
                        time.sleep(0.1)
                    time.sleep(0.5)
                    
                    # Send
                    btn = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, '//button[@data-tab="11"]'))
                    )
                    btn.click()
                    
                    success += 1
                    sent.add(idx)
                    print(f"  ✅ SENT!")
                    append_log({"company": lead['name'], "phone": lead['phone'],
                                "sent_at": datetime.now().isoformat(), "status": "success",
                                "country": lead['country']})
                    
                except TimeoutException:
                    print(f"  ⚠️ Contact not found on WhatsApp")
                    sent.add(idx)
                    failed += 1
                    append_log({"company": lead['name'], "phone": lead['phone'],
                                "sent_at": datetime.now().isoformat(), "status": "not_found",
                                "country": lead['country']})
                
                progress["sent_indices"] = sorted(list(sent))
                save_progress(progress)
                
                if idx != to_send[-1]:
                    delay = random.randint(MIN_DELAY, MAX_DELAY)
                    print(f"  ⏳ Waiting {delay//60}m{delay%60}s...")
                    for s in range(delay, 0, -10):
                        if s % 30 == 0 or s <= 10:
                            print(f"     {s}s...", end='\r')
                        time.sleep(10)
                    print()
                
            except Exception as e:
                print(f"  ❌ Error: {e}")
                failed += 1
                append_log({"company": lead['name'], "phone": lead['phone'],
                            "sent_at": datetime.now().isoformat(), "status": f"error: {str(e)[:80]}",
                            "country": lead['country']})
    
    except KeyboardInterrupt:
        print("\n\n⏹ Interrupted by user")
    
    finally:
        print(f"\n{'='*55}")
        print(f"📊 Session complete")
        print(f"  ✅ Success: {success}")
        print(f"  ❌ Failed: {failed}")
        print(f"  📊 Total sent: {len(sent)}/{len(leads)}")
        print(f"{'='*55}")
        progress["sent_indices"] = sorted(list(sent))
        progress["completed"] = len(sent) >= len(leads)
        save_progress(progress)
        input("\nPress Enter to close browser...")
        driver.quit()


def show_status():
    if os.path.exists(STATUS_FILE):
        with open(STATUS_FILE, 'r') as f:
            p = json.load(f)
        leads = load_leads()
        sent = len(p.get("sent_indices", []))
        print(f"\n📊 Progress: {sent}/{len(leads)}")
        print(f"   {'✅ Complete!' if p.get('completed') else '⏳ In progress'}")
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, 'r') as f:
                logs = json.load(f)
            ok = sum(1 for l in logs if l['status'] == 'success')
            fail = sum(1 for l in logs if l['status'] != 'success')
            print(f"   ✅ Sent: {ok}  ❌ Failed: {fail}")
            if logs:
                print(f"\n   Last 5:")
                for l in logs[-5:]:
                    em = '✅' if l['status'] == 'success' else '❌'
                    print(f"   {em} {l['company'][:30]:30s} {l['phone']}")
    else:
        print("📊 No records yet")


def reset():
    if os.path.exists(STATUS_FILE):
        c = input("⚠️ Reset all send records? (y/N): ")
        if c.lower() == 'y':
            os.remove(STATUS_FILE)
            print("✅ Reset complete")
        else:
            print("❌ Cancelled")
    else:
        print("📊 No records")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        cmd = sys.argv[1].lower()
        if cmd == "status": show_status()
        elif cmd == "reset": reset()
        elif cmd == "help":
            print("Usage:")
            print("  python whatsapp_bot.py          # Send messages")
            print("  python whatsapp_bot.py status   # View progress")
            print("  python whatsapp_bot.py reset    # Reset progress")
            print("  python whatsapp_bot.py help     # This help")
        else:
            print(f"Unknown: {cmd}")
    else:
        send_messages()