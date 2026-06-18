import json
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

# Load data
with open(r'C:\Users\Admin\AppData\Local\hermes\memories\脚本缓存\富通CRM\all_customers_raw.json', 'r') as f:
    customers = json.load(f)

total = len(customers)
print(f"Loaded {total} customers")

def ts_to_date(ts):
    if not ts or ts == 0: return ''
    try:
        return datetime.fromtimestamp(ts / 1000, tz=timezone(timedelta(hours=8))).strftime('%Y-%m-%d %H:%M')
    except: return ''

def safe(v):
    if v is None: return ''
    if isinstance(v, (int, float)): return v
    s = str(v)
    return ''.join(c for c in s if c.isprintable() or c in '\n\r\t').strip()

def sanitize(val):
    """Remove Excel-illegal characters from any value"""
    if isinstance(val, str):
        return ''.join(c for c in val if c.isprintable() or c in '\n\r\t').strip()
    return val

# Define field groups
FIELD_GROUPS = [
    ("基本信息", [
        ('code', '客户代码'), ('name', '客户名称'), ('shortName', '客户简称'),
        ('displayType', '客户类型'), ('status', '状态'), ('duplicateStatus', '重复状态'),
        ('source', '客户来源'), ('grade', '客户等级'), ('industryType', '行业类型'),
        ('mainProduct', '主营产品'), ('businessType', '业务类型'),
    ]),
    ("联系方式", [
        ('contactName', '联系人姓名'), ('contactEmail', '联系人邮箱'),
        ('contactMobile', '联系人手机号'), ('telephone', '固定电话'),
        ('fax', '传真'), ('webSite', '企业网站'), ('address', '联系地址'),
    ]),
    ("地理位置", [
        ('displayRegion', '国家/地区'), ('countryNo', '国家编号'), ('timeZone', '时区'),
    ]),
    ("社交媒体", [
        ('linkedinAccount', 'LinkedIn账号'), ('faceBookCmpMain', 'Facebook主页'),
        ('twitterCmpMain', 'Twitter主页'), ('instagramCmpMain', 'Instagram主页'),
    ]),
    ("销售信息", [
        ('displaySalesman', '业务员'), ('creator', '创建人'),
        ('managerCustomerName', '上级客户'),
    ]),
    ("时间信息", [
        ('displayCreateTime', '创建日期'), ('displayLastFollowTime', '最近联系时间'),
        ('recentlyFollowTime', '最近跟进时间'), ('quoteFirstDate', '首次报价日期'),
        ('quoteLastDate', '最近报价日期'), ('toHighseasTime', '预计转入公海日期'),
        ('nextRemindTime', '跟进提醒时间'), ('remainingTime', '周期剩余时长'),
    ]),
    ("业务活动", [
        ('activity', '最近活动'), ('activityType', '活动类型'),
        ('flowStep', '跟进阶段'), ('quoteCount', '报价次数'),
    ]),
    ("财务信息", [
        ('orderAmountUsd', '成交金额(USD)'), ('orderAmountCny', '成交金额(CNY)'),
        ('receiptSumAmountUsd', '回款金额(USD)'), ('receiptSumAmountCny', '回款金额(CNY)'),
        ('settleOverdueDay', '收汇逾期天数'),
    ]),
    ("企业信息", [
        ('employeesCount', '员工人数'), ('yearTurnover', '年营业额'),
        ('introduce', '企业介绍'), ('description', '备注'),
    ]),
]

# Flatten field list
ALL_FIELDS = [(key, label) for group_name, fields in FIELD_GROUPS for key, label in fields]

# Stats
grade_dist = Counter(cust.get('grade', '') or '未设置' for cust in customers)
source_dist = Counter(cust.get('source', '') or '未知' for cust in customers)
region_dist = Counter(cust.get('displayRegion', '') or '未知' for cust in customers)
salesman_dist = Counter(cust.get('displaySalesman', '') or '未知' for cust in customers)
status_dist = Counter({0: '活跃', 1: '公海', 4: '成交'}.get(cust.get('status'), '其他') for cust in customers)

# Styling constants
DARK_BLUE = '1E3A5F'
MEDIUM_BLUE = '2B579A'
LIGHT_BLUE = 'D6E4F0'
WHITE = 'FFFFFF'
header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
header_font = Font(name='微软雅黑', bold=True, color=WHITE, size=10)
section_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
section_font = Font(name='微软雅黑', bold=True, color=DARK_BLUE, size=10)
data_font = Font(name='微软雅黑', size=10)
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

wb = Workbook()

# ====================== SHEET 1: DASHBOARD ======================
ws1 = wb.active
ws1.title = '客户总览'
ws1.sheet_properties.tabColor = '1E3A5F'

ws1.merge_cells('A1:H1')
ws1['A1'] = 'BLIIOT 富通CRM 客户数据全量报表'
ws1['A1'].font = Font(name='微软雅黑', bold=True, size=16, color=DARK_BLUE)
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 40

ws1.merge_cells('A2:H2')
ws1['A2'] = f'导出时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}  |  总客户：{total}  |  操作员：bliiot03  |  字段：{len(ALL_FIELDS)}个'
ws1['A2'].font = Font(name='微软雅黑', size=9, color='888888')
ws1['A2'].alignment = Alignment(horizontal='center')

# KPI cards
kpi_data = [
    ('总客户数', total, '1E3A5F'),
    ('活跃客户', status_dist.get('活跃', 0), '27AE60'),
    ('公海客户', status_dist.get('公海', 0), 'E67E22'),
    ('有邮箱客户', sum(1 for c in customers if c.get('contactEmail')), '8E44AD'),
    ('有手机客户', sum(1 for c in customers if c.get('contactMobile')), '2980B9'),
    ('有报价客户', sum(1 for c in customers if (c.get('quoteCount') or 0) > 0), 'C0392B'),
]

row = 4
for i, (label, val, color) in enumerate(kpi_data):
    col = i * 2 + 1
    ws1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    cell = ws1.cell(row=row, column=col, value=val)
    cell.font = Font(name='微软雅黑', bold=True, size=22, color=WHITE)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    ws1.cell(row=row, column=col+1).border = thin_border
    ws1.cell(row=row, column=col+1).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    ws1.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    lc = ws1.cell(row=row+1, column=col, value=label)
    lc.font = Font(name='微软雅黑', size=10, color=WHITE)
    lc.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    lc.alignment = Alignment(horizontal='center', vertical='center')
    lc.border = thin_border
    ws1.cell(row=row+1, column=col+1).border = thin_border
    ws1.cell(row=row+1, column=col+1).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    ws1.row_dimensions[row].height = 50
    ws1.row_dimensions[row+1].height = 25

# Grade distribution table
row = 8
ws1.merge_cells(f'A{row}:D{row}')
ws1[f'A{row}'] = '📊 客户等级分布'
ws1[f'A{row}'].font = section_font
ws1[f'A{row}'].fill = section_fill
for c in range(1,5):
    ws1.cell(row=row, column=c).fill = section_fill
    ws1.cell(row=row, column=c).border = thin_border
row += 1
for i, h in enumerate(['等级', '数量', '占比', ' '], 1):
    cell = ws1.cell(row=row, column=i, value=h)
    cell.font = header_font; cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
for g, cnt in sorted(grade_dist.items(), key=lambda x: -x[1]):
    row += 1
    ws1.cell(row=row, column=1, value=g).font = data_font; ws1.cell(row=row, column=1).border = thin_border
    ws1.cell(row=row, column=2, value=cnt).font = data_font
    ws1.cell(row=row, column=2).alignment = Alignment(horizontal='center'); ws1.cell(row=row, column=2).border = thin_border
    ws1.cell(row=row, column=3, value=f'{cnt/total*100:.1f}%').font = data_font
    ws1.cell(row=row, column=3).alignment = Alignment(horizontal='center'); ws1.cell(row=row, column=3).border = thin_border

# Source distribution
col_s = 6
ws1.merge_cells(start_row=8, start_column=col_s, end_row=8, end_column=col_s+2)
ws1.cell(row=8, column=col_s, value='📊 客户来源分布 Top 15').font = section_font
for c in range(col_s, col_s+3):
    ws1.cell(row=8, column=c).fill = section_fill; ws1.cell(row=8, column=c).border = thin_border
r9 = 9
for i, h in enumerate(['来源', '数量', '占比'], col_s):
    cell = ws1.cell(row=r9, column=i, value=h)
    cell.font = header_font; cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
for s, cnt in sorted(source_dist.items(), key=lambda x: -x[1])[:15]:
    r9 += 1
    ws1.cell(row=r9, column=col_s, value=s).font = data_font; ws1.cell(row=r9, column=col_s).border = thin_border
    ws1.cell(row=r9, column=col_s+1, value=cnt).font = data_font
    ws1.cell(row=r9, column=col_s+1).alignment = Alignment(horizontal='center'); ws1.cell(row=r9, column=col_s+1).border = thin_border
    ws1.cell(row=r9, column=col_s+2, value=f'{cnt/total*100:.1f}%').font = data_font
    ws1.cell(row=r9, column=col_s+2).alignment = Alignment(horizontal='center'); ws1.cell(row=r9, column=col_s+2).border = thin_border

# Country distribution (bottom)
row_ct = r9 + 3
ws1.merge_cells(f'A{row_ct}:H{row_ct}')
ws1[f'A{row_ct}'] = '🌍 国家/地区分布 Top 20'
ws1[f'A{row_ct}'].font = section_font; ws1[f'A{row_ct}'].fill = section_fill
for c in range(1,9): ws1.cell(row=row_ct, column=c).fill = section_fill; ws1.cell(row=row_ct, column=c).border = thin_border
row_ct += 1
for i, h in enumerate(['国家/地区', '数量', '占比', ' '], 1):
    cell = ws1.cell(row=row_ct, column=i, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
for reg, cnt in sorted(region_dist.items(), key=lambda x: -x[1])[:20]:
    row_ct += 1
    ws1.cell(row=row_ct, column=1, value=reg).font = data_font; ws1.cell(row=row_ct, column=1).border = thin_border
    ws1.cell(row=row_ct, column=2, value=cnt).font = data_font
    ws1.cell(row=row_ct, column=2).alignment = Alignment(horizontal='center'); ws1.cell(row=row_ct, column=2).border = thin_border
    ws1.cell(row=row_ct, column=3, value=f'{cnt/total*100:.1f}%').font = data_font
    ws1.cell(row=row_ct, column=3).alignment = Alignment(horizontal='center'); ws1.cell(row=row_ct, column=3).border = thin_border

# Sheet 1 column widths
for c, w in zip(range(1,9), [18,12,12,12,5,18,12,12]):
    ws1.column_dimensions[get_column_letter(c)].width = w

# ====================== SHEET 2: ALL CUSTOMERS ======================
ws2 = wb.create_sheet('全部客户数据')
ws2.sheet_properties.tabColor = '2B579A'

# Headers
col = 1
for group_name, fields in FIELD_GROUPS:
    for key, label in fields:
        cell = ws2.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        col += 1
ws2.row_dimensions[1].height = 30

# Data
for i, cust in enumerate(customers):
    row = i + 2
    col = 1
    alt_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid') if i % 2 == 0 else None
    for group_name, fields in FIELD_GROUPS:
        for key, label in fields:
            raw = cust.get(key)
            if key in ['displayCreateTime', 'displayLastFollowTime', 'recentlyFollowTime',
                       'quoteFirstDate', 'quoteLastDate', 'toHighseasTime']:
                val = ts_to_date(raw)
            elif key == 'status':
                val = {0: '活跃', 1: '公海', 4: '成交'}.get(raw, str(raw))
            elif isinstance(raw, list):
                val = ', '.join(str(x) for x in raw) if raw else ''
            elif isinstance(raw, (int, float)) and raw > 1000000000000:
                val = ts_to_date(raw)
            else:
                val = safe(raw)
            # Remove illegal Excel characters
            if isinstance(val, str):
                val = ''.join(c for c in val if c.isprintable() or c in '\n\r\t')
                val = val.strip()
            cell = ws2.cell(row=row, column=col, value=val)
            cell.font = Font(name='微软雅黑', size=9)
            cell.alignment = Alignment(vertical='center')
            cell.border = thin_border
            if alt_fill: cell.fill = alt_fill
            col += 1
    if (i + 1) % 500 == 0:
        print(f"  Writing row {i+1}/{total}...")

# Auto-width
for col_idx in range(1, len(ALL_FIELDS) + 1):
    max_len = len(str(ws2.cell(row=1, column=col_idx).value or ''))
    for sr in range(2, min(8, total + 2)):
        v = ws2.cell(row=sr, column=col_idx).value
        if v:
            vl = len(str(v))
            if any('\u4e00' <= ch <= '\u9fff' for ch in str(v)):
                vl *= 2
            max_len = max(max_len, vl)
    ws2.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

ws2.freeze_panes = 'C2'
ws2.auto_filter.ref = f'A1:{get_column_letter(len(ALL_FIELDS))}{total + 1}'

# ====================== SHEET 3: SALESMAN ======================
ws3 = wb.create_sheet('业务员统计')
ws3.sheet_properties.tabColor = '27AE60'

headers_3 = ['业务员', '客户数', '占比', '活跃', '公海', '已成交',
             '有邮箱', '有手机', '有报价', '总报价次数', '邮箱率', '手机率']

for i, h in enumerate(headers_3, 1):
    cell = ws3.cell(row=1, column=i, value=h)
    cell.font = header_font; cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center'); cell.border = thin_border

row = 2
for salesman, cnt in sorted(salesman_dist.items(), key=lambda x: -x[1]):
    sc = [c for c in customers if (c.get('displaySalesman') or '未知') == salesman]
    sc_total = len(sc)
    vals = [
        salesman, sc_total, f'{sc_total/total*100:.1f}%',
        sum(1 for c in sc if c.get('status') == 0),
        sum(1 for c in sc if c.get('status') == 1),
        sum(1 for c in sc if c.get('status') == 4),
        sum(1 for c in sc if c.get('contactEmail')),
        sum(1 for c in sc if c.get('contactMobile')),
        sum(1 for c in sc if (c.get('quoteCount') or 0) > 0),
        sum(c.get('quoteCount') or 0 for c in sc),
        f'{sum(1 for c in sc if c.get("contactEmail"))/sc_total*100:.0f}%' if sc_total > 0 else '0%',
        f'{sum(1 for c in sc if c.get("contactMobile"))/sc_total*100:.0f}%' if sc_total > 0 else '0%',
    ]
    for i, v in enumerate(vals, 1):
        cell = ws3.cell(row=row, column=i, value=v)
        cell.font = data_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1

for i, w in enumerate([18, 10, 8, 8, 8, 8, 10, 10, 10, 12, 8, 8], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = 'A2'

# ====================== SHEET 4: REGION ======================
ws4 = wb.create_sheet('国家地区统计')
ws4.sheet_properties.tabColor = 'E67E22'

headers_4 = ['国家/地区', '客户数', '占比', '有邮箱', '有手机',
             '潜在客户', '意向客户', '已成交', '总报价次数']

for i, h in enumerate(headers_4, 1):
    cell = ws4.cell(row=1, column=i, value=h)
    cell.font = header_font; cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center'); cell.border = thin_border

row = 2
for region, cnt in sorted(region_dist.items(), key=lambda x: -x[1]):
    rc = [c for c in customers if (c.get('displayRegion') or '未知') == region]
    r_cnt = len(rc)
    vals = [
        region, r_cnt, f'{r_cnt/total*100:.1f}%',
        sum(1 for c in rc if c.get('contactEmail')),
        sum(1 for c in rc if c.get('contactMobile')),
        sum(1 for c in rc if c.get('grade') == '潜在客户'),
        sum(1 for c in rc if c.get('grade') == '意向客户'),
        sum(1 for c in rc if c.get('grade') == '成交客户'),
        sum(c.get('quoteCount') or 0 for c in rc),
    ]
    for i, v in enumerate(vals, 1):
        cell = ws4.cell(row=row, column=i, value=v)
        cell.font = data_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1

for i, w in enumerate([35, 10, 8, 10, 10, 10, 10, 10, 12], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = 'A2'

# Save
output_path = r'C:\Users\Admin\Desktop\Working\BLIIOT_富通CRM_客户数据全量报表.xlsx'
wb.save(output_path)
print(f"\n✅ Done! Saved to: {output_path}")
print(f"   Sheet 1: 客户总览 (KPI dashboard)")
print(f"   Sheet 2: 全部客户数据 ({total} rows, {len(ALL_FIELDS)} cols)")
print(f"   Sheet 3: 业务员统计 ({len(salesman_dist)} salespeople)")
print(f"   Sheet 4: 国家地区统计 ({len(region_dist)} countries/regions)")