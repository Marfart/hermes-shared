"""BLIIOT CRM 终极版 — 跟进记录 + Excel生成 + 双向同步"""
import sqlite3, json, sys, os, argparse, subprocess
from datetime import datetime
from typing import Optional, List, Dict

# Paths
DB_DIR = os.path.join(os.environ.get('LOCALAPPDATA', r'C:\Users\Admin\AppData\Local'),
                       'hermes', 'memories', '脚本缓存', '富通CRM')
DB_PATH = os.path.join(DB_DIR, 'crm_followups.db')
CUSTOMER_DATA_PATH = os.path.join(DB_DIR, 'all_customers_raw.json')
EXCEL_PATH = r'C:\Users\Admin\Desktop\Working\BLIIOT_富通CRM_客户数据全量报表.xlsx'

FOLLOWUP_TYPES = ['跟进', '邮件', 'WhatsApp', '电话', '报价', '订单', '会议', '备注', '系统']

def get_db():
    os.makedirs(DB_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS followups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER NOT NULL,
            customer_name TEXT NOT NULL DEFAULT '',
            contact_name TEXT DEFAULT '',
            type TEXT NOT NULL DEFAULT '跟进',
            content TEXT NOT NULL,
            operator TEXT NOT NULL DEFAULT 'Kali Marfa',
            status TEXT DEFAULT 'active',
            created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime')),
            synced INTEGER NOT NULL DEFAULT 0,
            joinf_follow_id TEXT DEFAULT NULL,
            source TEXT DEFAULT 'manual'
        );
        CREATE INDEX IF NOT EXISTS idx_followups_customer ON followups(customer_id, created_at DESC);
        CREATE INDEX IF NOT EXISTS idx_followups_synced ON followups(synced);
        CREATE TABLE IF NOT EXISTS customer_cache (
            customer_id INTEGER PRIMARY KEY,
            customer_name TEXT NOT NULL DEFAULT '',
            contact_name TEXT DEFAULT '',
            contact_email TEXT DEFAULT '',
            contact_mobile TEXT DEFAULT '',
            source TEXT DEFAULT '',
            grade TEXT DEFAULT '',
            region TEXT DEFAULT '',
            updated_at TEXT DEFAULT (datetime('now', 'localtime'))
        );
        CREATE TABLE IF NOT EXISTS sync_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            followup_id INTEGER, customer_id INTEGER, action TEXT,
            status TEXT, message TEXT, created_at TEXT DEFAULT (datetime('now', 'localtime'))
        );
    """)
    conn.commit()
    conn.close()

def load_customers() -> List[Dict]:
    if not os.path.exists(CUSTOMER_DATA_PATH):
        print("⚠️ 客户数据文件未找到")
        return []
    with open(CUSTOMER_DATA_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)

def update_customer_cache():
    customers = load_customers()
    conn = get_db()
    conn.execute("DELETE FROM customer_cache")
    for c in customers:
        conn.execute("""INSERT OR REPLACE INTO customer_cache
            (customer_id, customer_name, contact_name, contact_email, contact_mobile, source, grade, region, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, datetime('now', 'localtime'))""",
            (c.get('id') or 0, c.get('name', '') or '', c.get('contactName', '') or '',
             c.get('contactEmail', '') or '', c.get('contactMobile', '') or '',
             c.get('source', '') or '', c.get('grade', '') or '',
             (c.get('displayRegion') or '')[:60]))
    conn.commit()
    conn.close()
    print(f"✅ 已缓存 {len(customers)} 个客户信息")

def add_followup(customer_id: int, content: str, type_: str = '跟进',
                 operator: str = 'Kali Marfa', source: str = 'manual',
                 customer_name: str = '', sync: bool = False, auto_sync: bool = False) -> Dict:
    """Add a follow-up record. If auto_sync=True, queues for background sync."""
    conn = get_db()
    if not customer_name:
        row = conn.execute("SELECT customer_name FROM customer_cache WHERE customer_id=?",
                          (customer_id,)).fetchone()
        if row:
            customer_name = row['customer_name']
        else:
            for c in load_customers():
                if c.get('id') == customer_id:
                    customer_name = c.get('name', '') or c.get('contactName', '')
                    break

    cursor = conn.execute("""INSERT INTO followups (customer_id, customer_name, type, content, operator, source)
        VALUES (?, ?, ?, ?, ?, ?)""", (customer_id, customer_name, type_, content, operator, source))
    followup_id = cursor.lastrowid
    conn.commit()

    result = {'id': followup_id, 'customer_id': customer_id, 'customer_name': customer_name,
              'type': type_, 'content': content,
              'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'synced': False}

    if sync:
        synced = _sync_one_to_joinf(followup_id, customer_id, content, type_, result['created_at'])
        result['synced'] = synced

    conn.close()
    return result

def _sync_one_to_joinf(followup_id, customer_id, content, type_='跟进', created_at=None):
    """Sync via CDP browser to JoinF CRM"""
    if not created_at:
        created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    script_dir = DB_DIR
    sync_script = os.path.join(script_dir, 'sync_to_joinf.cjs')
    record = {'id': followup_id, 'customer_id': customer_id, 'content': content,
              'type': type_, 'created_at': created_at}
    try:
        result = subprocess.run(['node', sync_script], input=json.dumps(record),
                              capture_output=True, text=True, timeout=30, cwd=script_dir)
        if result.returncode == 0:
            try:
                response = json.loads(result.stdout)
                if response.get('status') == 'complete':
                    conn = get_db()
                    conn.execute("UPDATE followups SET synced=1 WHERE id=?", (followup_id,))
                    conn.commit(); conn.close()
                    return True
            except: pass
        return False
    except: return False

def list_followups(customer_id: Optional[int] = None, limit: int = 50,
                   type_: Optional[str] = None) -> List[Dict]:
    conn = get_db()
    query = "SELECT * FROM followups WHERE 1=1"
    params = []
    if customer_id:
        query += " AND customer_id=?"
        params.append(customer_id)
    if type_:
        query += " AND type=?"
        params.append(type_)
    query += " ORDER BY created_at DESC LIMIT ?"
    params.append(limit)
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def search_followups(keyword: str) -> List[Dict]:
    conn = get_db()
    rows = conn.execute("""SELECT * FROM followups
        WHERE content LIKE ? OR customer_name LIKE ? OR contact_name LIKE ?
        ORDER BY created_at DESC LIMIT 50""",
        (f'%{keyword}%', f'%{keyword}%', f'%{keyword}%')).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_customer_followups_stats(customer_id: int) -> Dict:
    conn = get_db()
    row = conn.execute("""SELECT COUNT(*) as total_count,
        COUNT(CASE WHEN type='邮件' THEN 1 END) as email_count,
        COUNT(CASE WHEN type='WhatsApp' THEN 1 END) as whatsapp_count,
        COUNT(CASE WHEN type='报价' THEN 1 END) as quote_count,
        COUNT(CASE WHEN type='电话' THEN 1 END) as call_count,
        MAX(created_at) as last_followup, MIN(created_at) as first_followup
        FROM followups WHERE customer_id=?""", (customer_id,)).fetchone()
    conn.close()
    return dict(row) if row else {}

def get_stats() -> Dict:
    conn = get_db()
    stats = {}
    stats['total_followups'] = conn.execute("SELECT COUNT(*) FROM followups").fetchone()[0]
    stats['synced_count'] = conn.execute("SELECT COUNT(*) FROM followups WHERE synced=1").fetchone()[0]
    stats['pending_sync'] = conn.execute("SELECT COUNT(*) FROM followups WHERE synced=0").fetchone()[0]
    stats['active_customers'] = conn.execute("SELECT COUNT(DISTINCT customer_id) FROM followups").fetchone()[0]
    type_rows = conn.execute("SELECT type, COUNT(*) as cnt FROM followups GROUP BY type ORDER BY cnt DESC").fetchall()
    stats['by_type'] = {r['type']: r['cnt'] for r in type_rows}
    stats['today'] = conn.execute("""SELECT COUNT(*) FROM followups
        WHERE date(created_at) = date('now', 'localtime')""").fetchone()[0]
    conn.close()
    return stats

def generate_excel():
    """Generate the comprehensive final Excel with all data + follow-up integration."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import Counter

    customers = load_customers()
    total = len(customers)
    print(f"Loaded {total} customers")

    def ts_to_date(ts):
        if not ts or ts == 0: return ''
        try: return datetime.fromtimestamp(ts / 1000).strftime('%Y-%m-%d %H:%M')
        except: return ''

    def safe(v):
        if v is None: return ''
        if isinstance(v, (int, float)): return v
        s = str(v)
        return ''.join(c for c in s if c.isprintable() or c in '\n\r\t').strip()

    # Get follow-up data
    conn = get_db()
    followup_rows = conn.execute("""SELECT customer_id, COUNT(*) as cnt,
        MAX(created_at) as last_time, GROUP_CONCAT(type || ': ' || substr(content, 1, 80), ' || ') as recent
        FROM followups WHERE status='active' GROUP BY customer_id""").fetchall()
    followup_map = {r['customer_id']: dict(r) for r in followup_rows}

    # All follow-ups for the detail sheet
    all_followups = conn.execute("""SELECT f.*, c.region, c.contact_email, c.contact_mobile
        FROM followups f LEFT JOIN customer_cache c ON f.customer_id = c.customer_id
        ORDER BY f.created_at DESC""").fetchall()
    conn.close()

    FIELD_GROUPS = [
        ("基本信息", [
            ('code', '客户代码'), ('name', '客户名称'), ('shortName', '客户简称'),
            ('displayType', '客户类型'), ('source', '客户来源'), ('grade', '客户等级'),
            ('industryType', '行业类型'), ('mainProduct', '主营产品'),
        ]),
        ("联系方式", [
            ('contactName', '联系人姓名'), ('contactEmail', '联系人邮箱'),
            ('contactMobile', '联系人手机号'), ('telephone', '固定电话'),
            ('webSite', '企业网站'), ('address', '联系地址'),
        ]),
        ("地理位置与社交", [
            ('displayRegion', '国家/地区'), ('timeZone', '时区'),
            ('linkedinAccount', 'LinkedIn'), ('faceBookCmpMain', 'Facebook'),
        ]),
        ("销售信息", [
            ('displaySalesman', '业务员'), ('creator', '创建人'),
            ('activity', '最近活动'), ('flowStep', '跟进阶段'), ('quoteCount', '报价次数'),
        ]),
        ("时间信息", [
            ('displayCreateTime', '创建日期'), ('displayLastFollowTime', '最近联系时间'),
            ('recentlyFollowTime', '最近跟进时间'), ('quoteFirstDate', '首次报价日期'),
            ('quoteLastDate', '最近报价日期'),
        ]),
        ("财务信息", [
            ('orderAmountUsd', '成交金额(USD)'), ('receiptSumAmountUsd', '回款金额(USD)'),
            ('settleOverdueDay', '收汇逾期天数'), ('employeesCount', '员工人数'),
            ('description', '备注'),
        ]),
        ("💬 本地跟进 (CRM工具)", [
            ('_local_followup_count', '本地跟进次数'),
            ('_local_last_followup', '最近跟进时间'),
            ('_local_recent_followups', '近期跟进摘要'),
        ]),
    ]

    ALL_FIELDS = [(key, label) for group_name, fields in FIELD_GROUPS for key, label in fields]

    # Stats
    grade_dist = Counter(cust.get('grade', '') or '未设置' for cust in customers)
    source_dist = Counter(cust.get('source', '') or '未知' for cust in customers)
    region_dist = Counter(cust.get('displayRegion', '') or '未知' for cust in customers)
    salesman_dist = Counter(cust.get('displaySalesman', '') or '未知' for cust in customers)

    # Colors
    DARK_BLUE, MEDIUM_BLUE, LIGHT_BLUE, WHITE = '1E3A5F', '2B579A', 'D6E4F0', 'FFFFFF'
    header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
    header_font = Font(name='微软雅黑', bold=True, color=WHITE, size=10)
    section_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
    section_font = Font(name='微软雅黑', bold=True, color=DARK_BLUE, size=10)
    data_font = Font(name='微软雅黑', size=10)
    thin_border = Border(left=Side(style='thin',color='D0D0D0'), right=Side(style='thin',color='D0D0D0'),
                         top=Side(style='thin',color='D0D0D0'), bottom=Side(style='thin',color='D0D0D0'))

    wb = Workbook()

    # ===== SHEET 1: DASHBOARD =====
    ws1 = wb.active
    ws1.title = '客户总览'
    ws1.sheet_properties.tabColor = '1E3A5F'

    ws1.merge_cells('A1:H1')
    ws1['A1'] = 'BLIIOT 富通CRM 客户数据全量报表'
    ws1['A1'].font = Font(name='微软雅黑', bold=True, size=16, color=DARK_BLUE)
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 40
    ws1.merge_cells('A2:H2')
    ws1['A2'] = f'导出：{datetime.now().strftime("%Y-%m-%d %H:%M")}  |  客户：{total}  |  字段：{len(ALL_FIELDS)}  |  跟进：{len(all_followups)}条'
    ws1['A2'].font = Font(name='微软雅黑', size=9, color='888888')
    ws1['A2'].alignment = Alignment(horizontal='center')

    kpi_data = [
        ('总客户', total, '1E3A5F'),
        ('已跟进', len(followup_map), '27AE60'),
        ('跟进记录', len(all_followups), 'E67E22'),
        ('有邮箱', sum(1 for c in customers if c.get('contactEmail')), '8E44AD'),
        ('有手机', sum(1 for c in customers if c.get('contactMobile')), '2980B9'),
        ('有报价', sum(1 for c in customers if (c.get('quoteCount') or 0) > 0), 'C0392B'),
    ]
    row = 4
    for i, (label, val, color) in enumerate(kpi_data):
        col = i * 2 + 1
        ws1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        cell = ws1.cell(row=row, column=col, value=val)
        cell.font = Font(name='微软雅黑', bold=True, size=22, color=WHITE)
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.fill = fill; cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = thin_border
        ws1.cell(row=row, column=col+1).border = thin_border; ws1.cell(row=row, column=col+1).fill = fill
        ws1.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
        lc = ws1.cell(row=row+1, column=col, value=label)
        lc.font = Font(name='微软雅黑', size=10, color=WHITE); lc.fill = fill; lc.alignment = Alignment(horizontal='center', vertical='center'); lc.border = thin_border
        ws1.cell(row=row+1, column=col+1).border = thin_border; ws1.cell(row=row+1, column=col+1).fill = fill
        ws1.row_dimensions[row].height = 50; ws1.row_dimensions[row+1].height = 25

    # Grade distribution
    row = 8
    ws1.merge_cells(f'A{row}:D{row}')
    ws1[f'A{row}'] = '📊 客户等级分布'
    ws1[f'A{row}'].font = section_font; ws1[f'A{row}'].fill = section_fill
    for c in range(1,5): ws1.cell(row=row, column=c).fill = section_fill; ws1.cell(row=row, column=c).border = thin_border
    row += 1
    for i, h in enumerate(['等级', '数量', '占比', ''], 1):
        cell = ws1.cell(row=row, column=i, value=h); cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
    for g, cnt in sorted(grade_dist.items(), key=lambda x: -x[1]):
        row += 1; ws1.cell(row=row, column=1, value=g).font = data_font; ws1.cell(row=row, column=1).border = thin_border
        ws1.cell(row=row, column=2, value=cnt).font = data_font; ws1.cell(row=row, column=2).alignment = Alignment(horizontal='center'); ws1.cell(row=row, column=2).border = thin_border
        ws1.cell(row=row, column=3, value=f'{cnt/total*100:.1f}%').font = data_font; ws1.cell(row=row, column=3).alignment = Alignment(horizontal='center'); ws1.cell(row=row, column=3).border = thin_border

    # Source distribution
    col_s = 6
    ws1.merge_cells(start_row=8, start_column=col_s, end_row=8, end_column=col_s+2)
    ws1.cell(row=8, column=col_s, value='📊 客户来源 Top 15').font = section_font
    for c in range(col_s, col_s+3): ws1.cell(row=8, column=c).fill = section_fill; ws1.cell(row=8, column=c).border = thin_border
    r9 = 9
    for i, h in enumerate(['来源', '数量', '占比'], col_s):
        cell = ws1.cell(row=r9, column=i, value=h); cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
    for s, cnt in sorted(source_dist.items(), key=lambda x: -x[1])[:15]:
        r9 += 1; ws1.cell(row=r9, column=col_s, value=s).font = data_font; ws1.cell(row=r9, column=col_s).border = thin_border
        ws1.cell(row=r9, column=col_s+1, value=cnt).font = data_font; ws1.cell(row=r9, column=col_s+1).alignment = Alignment(horizontal='center'); ws1.cell(row=r9, column=col_s+1).border = thin_border
        ws1.cell(row=r9, column=col_s+2, value=f'{cnt/total*100:.1f}%').font = data_font; ws1.cell(row=r9, column=col_s+2).alignment = Alignment(horizontal='center'); ws1.cell(row=r9, column=col_s+2).border = thin_border

    # Country
    row_ct = r9 + 3
    ws1.merge_cells(f'A{row_ct}:H{row_ct}')
    ws1[f'A{row_ct}'] = '🌍 国家/地区 Top 20'
    ws1[f'A{row_ct}'].font = section_font; ws1[f'A{row_ct}'].fill = section_fill
    for c in range(1,9): ws1.cell(row=row_ct, column=c).fill = section_fill; ws1.cell(row=row_ct, column=c).border = thin_border
    row_ct += 1
    for i, h in enumerate(['国家/地区', '数量', '占比', ''], 1):
        cell = ws1.cell(row=row_ct, column=i, value=h); cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
    for reg, cnt in sorted(region_dist.items(), key=lambda x: -x[1])[:20]:
        row_ct += 1; ws1.cell(row=row_ct, column=1, value=reg).font = data_font; ws1.cell(row=row_ct, column=1).border = thin_border
        ws1.cell(row=row_ct, column=2, value=cnt).font = data_font; ws1.cell(row=row_ct, column=2).alignment = Alignment(horizontal='center'); ws1.cell(row=row_ct, column=2).border = thin_border
        ws1.cell(row=row_ct, column=3, value=f'{cnt/total*100:.1f}%').font = data_font; ws1.cell(row=row_ct, column=3).alignment = Alignment(horizontal='center'); ws1.cell(row=row_ct, column=3).border = thin_border

    for c, w in zip(range(1,9), [18,12,12,12,5,18,12,12]):
        ws1.column_dimensions[get_column_letter(c)].width = w

    # ===== SHEET 2: ALL CUSTOMERS (WITH FOLLOW-UP DATA) =====
    ws2 = wb.create_sheet('全部客户数据')
    ws2.sheet_properties.tabColor = '2B579A'

    col = 1
    for group_name, fields in FIELD_GROUPS:
        for key, label in fields:
            cell = ws2.cell(row=1, column=col, value=label)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            col += 1
    ws2.row_dimensions[1].height = 30

    for i, cust in enumerate(customers):
        row = i + 2
        col = 1
        cid = cust.get('id')
        fup = followup_map.get(cid, {})
        alt_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid') if i % 2 == 0 else None
        for group_name, fields in FIELD_GROUPS:
            for key, label in fields:
                if key.startswith('_local_'):
                    if key == '_local_followup_count':
                        val = fup.get('cnt', 0) or 0
                    elif key == '_local_last_followup':
                        val = fup.get('last_time', '') or ''
                    elif key == '_local_recent_followups':
                        val = (fup.get('recent', '') or '')[:200]
                    else:
                        val = ''
                else:
                    raw = cust.get(key)
                    if key in ['displayCreateTime', 'displayLastFollowTime', 'recentlyFollowTime',
                               'quoteFirstDate', 'quoteLastDate']:
                        val = ts_to_date(raw)
                    elif key == 'status':
                        val = {0: '活跃', 1: '公海', 4: '成交'}.get(raw, str(raw))
                    elif isinstance(raw, list):
                        val = ', '.join(str(x) for x in raw) if raw else ''
                    elif isinstance(raw, (int, float)) and raw > 1000000000000:
                        val = ts_to_date(raw)
                    else:
                        val = safe(raw)
                cell = ws2.cell(row=row, column=col, value=val)
                cell.font = Font(name='微软雅黑', size=9)
                cell.alignment = Alignment(vertical='center')
                cell.border = thin_border
                if alt_fill: cell.fill = alt_fill
                col += 1
        if (i + 1) % 500 == 0:
            print(f"  Writing row {i+1}/{total}...")

    for col_idx in range(1, len(ALL_FIELDS) + 1):
        max_len = len(str(ws2.cell(row=1, column=col_idx).value or ''))
        for sr in range(2, min(8, total + 2)):
            v = ws2.cell(row=sr, column=col_idx).value
            if v:
                vl = len(str(v))
                if any('\u4e00' <= ch <= '\u9fff' for ch in str(v)): vl *= 2
                max_len = max(max_len, vl)
        ws2.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws2.freeze_panes = 'C2'
    ws2.auto_filter.ref = f'A1:{get_column_letter(len(ALL_FIELDS))}{total + 1}'

    # Highlight the follow-up columns with a different color
    fup_start_col = len(ALL_FIELDS) - 2
    fup_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    for c in range(fup_start_col, len(ALL_FIELDS) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.fill = PatternFill(start_color='E67E22', end_color='E67E22', fill_type='solid')

    # ===== SHEET 3: SALESMAN =====
    ws3 = wb.create_sheet('业务员统计')
    ws3.sheet_properties.tabColor = '27AE60'
    headers_3 = ['业务员', '客户数', '占比', '已跟进', '跟进次数', '有邮箱', '有手机', '有报价', '邮箱率', '手机率']
    for i, h in enumerate(headers_3, 1):
        cell = ws3.cell(row=1, column=i, value=h); cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
    row = 2
    for salesman, cnt in sorted(salesman_dist.items(), key=lambda x: -x[1]):
        sc = [c for c in customers if (c.get('displaySalesman') or '未知') == salesman]
        sc_total = len(sc)
        sc_ids = [c.get('id') for c in sc if c.get('id')]
        tracked = sum(1 for cid in sc_ids if cid in followup_map)
        fup_cnt = sum(followup_map.get(cid, {}).get('cnt', 0) or 0 for cid in sc_ids)
        vals = [salesman, sc_total, f'{sc_total/total*100:.1f}%', tracked, fup_cnt,
                sum(1 for c in sc if c.get('contactEmail')),
                sum(1 for c in sc if c.get('contactMobile')),
                sum(1 for c in sc if (c.get('quoteCount') or 0) > 0),
                f'{sum(1 for c in sc if c.get("contactEmail"))/sc_total*100:.0f}%' if sc_total > 0 else '0%',
                f'{sum(1 for c in sc if c.get("contactMobile"))/sc_total*100:.0f}%' if sc_total > 0 else '0%']
        for i, v in enumerate(vals, 1):
            cell = ws3.cell(row=row, column=i, value=v); cell.font = data_font
            cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
        row += 1
    for i, w in enumerate([18, 10, 8, 10, 10, 10, 10, 10, 8, 8], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = 'A2'

    # ===== SHEET 4: REGION =====
    ws4 = wb.create_sheet('国家地区统计')
    ws4.sheet_properties.tabColor = 'E67E22'
    headers_4 = ['国家/地区', '客户数', '占比', '已跟进', '跟进次数', '有邮箱', '有手机', '总报价次数']
    for i, h in enumerate(headers_4, 1):
        cell = ws4.cell(row=1, column=i, value=h); cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
    row = 2
    for region, cnt in sorted(region_dist.items(), key=lambda x: -x[1]):
        rc = [c for c in customers if (c.get('displayRegion') or '未知') == region]
        r_cnt = len(rc)
        r_ids = [c.get('id') for c in rc if c.get('id')]
        tracked = sum(1 for cid in r_ids if cid in followup_map)
        fup_cnt = sum(followup_map.get(cid, {}).get('cnt', 0) or 0 for cid in r_ids)
        vals = [region, r_cnt, f'{r_cnt/total*100:.1f}%', tracked, fup_cnt,
                sum(1 for c in rc if c.get('contactEmail')),
                sum(1 for c in rc if c.get('contactMobile')),
                sum(c.get('quoteCount') or 0 for c in rc)]
        for i, v in enumerate(vals, 1):
            cell = ws4.cell(row=row, column=i, value=v); cell.font = data_font
            cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
        row += 1
    for i, w in enumerate([35, 10, 8, 10, 10, 10, 10, 12], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.freeze_panes = 'A2'

    # ===== SHEET 5: FOLLOW-UP DETAIL =====
    ws5 = wb.create_sheet('跟进记录明细')
    ws5.sheet_properties.tabColor = 'E74C3C'
    fup_headers = ['ID', '客户ID', '客户名称', '跟进类型', '跟进内容', '操作人',
                   '来源', '创建时间', '已同步', '国家/地区', '联系人邮箱', '联系人手机']
    for i, h in enumerate(fup_headers, 1):
        cell = ws5.cell(row=1, column=i, value=h)
        cell.font = Font(name='微软雅黑', bold=True, color=WHITE, size=10)
        cell.fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    for i, r in enumerate(all_followups):
        row = i + 2
        vals = [r['id'], r['customer_id'], r['customer_name'], r['type'], r['content'],
                r['operator'], r['source'], r['created_at'],
                '✅' if r['synced'] else '⏳', r['region'] or '', r['contact_email'] or '', r['contact_mobile'] or '']
        for j, v in enumerate(vals, 1):
            cell = ws5.cell(row=row, column=j, value=v)
            cell.font = Font(name='微软雅黑', size=9)
            cell.alignment = Alignment(vertical='center', wrap_text=True if j == 5 else False)
            cell.border = thin_border

    for i, w in enumerate([6, 10, 20, 10, 60, 15, 10, 18, 8, 20, 25, 18], 1):
        ws5.column_dimensions[get_column_letter(i)].width = w
    ws5.freeze_panes = 'A2'
    ws5.auto_filter.ref = f'A1:L{len(all_followups) + 1}'

    wb.save(EXCEL_PATH)
    print(f"\n✅ 完整报表已保存: {EXCEL_PATH}")
    print(f"   Sheet 1: 客户总览 (KPI)")
    print(f"   Sheet 2: 全部客户数据 ({total} 行, {len(ALL_FIELDS)} 列) 🆕含跟进统计")
    print(f"   Sheet 3: 业务员统计 ({len(salesman_dist)} 人)")
    print(f"   Sheet 4: 国家地区统计 ({len(region_dist)} 个国家)")
    print(f"   Sheet 5: 跟进记录明细 ({len(all_followups)} 条)")

# ==================== CLI ====================
def main():
    parser = argparse.ArgumentParser(description='BLIIOT CRM 跟进记录工具')
    sub = parser.add_subparsers(dest='command')

    sub.add_parser('init', help='初始化数据库')
    sub.add_parser('stats', help='查看统计')
    sub.add_parser('update-cache', help='更新客户缓存')
    sub.add_parser('generate-excel', help='生成完整Excel报表')

    p_add = sub.add_parser('add', help='添加跟进记录')
    p_add.add_argument('customer_id', type=int, help='客户ID')
    p_add.add_argument('content', help='跟进内容')
    p_add.add_argument('--type', default='跟进', choices=FOLLOWUP_TYPES, help='跟进类型')
    p_add.add_argument('--name', default='', help='客户名称')
    p_add.add_argument('--sync', action='store_true', help='同步到富通CRM')
    p_add.add_argument('--source', default='manual', help='来源(manual/email/whatsapp/system)')

    p_list = sub.add_parser('list', help='查看跟进记录')
    p_list.add_argument('--customer', type=int, help='客户ID过滤')
    p_list.add_argument('--type', help='跟进类型过滤')
    p_list.add_argument('--limit', type=int, default=20, help='显示条数')

    p_search = sub.add_parser('search', help='搜索跟进记录')
    p_search.add_argument('keyword', help='搜索关键词')

    p_sync = sub.add_parser('sync', help='同步待处理记录')
    p_sync.add_argument('--customer', type=int, help='指定客户ID')

    args = parser.parse_args()

    if args.command == 'init':
        init_db()
        update_customer_cache()
    elif args.command == 'stats':
        s = get_stats()
        print("📊 BLIIOT CRM 跟进统计\n")
        print(f"  总跟进记录: {s['total_followups']}")
        print(f"  已同步富通: {s['synced_count']}")
        print(f"  待同步: {s['pending_sync']}")
        print(f"  跟进中客户: {s['active_customers']}")
        print(f"  今日新增: {s['today']}")
        print(f"\n  按类型分布:")
        for t, c in s.get('by_type', {}).items():
            print(f"    {t}: {c}次")
    elif args.command == 'add':
        result = add_followup(args.customer_id, args.content, args.type,
                            customer_name=args.name, sync=args.sync, source=args.source)
        print(f"✅ 跟进记录已添加 [ID={result['id']}]")
        print(f"   客户: {result['customer_name'] or args.customer_id}")
        print(f"   类型: {result['type']}")
        print(f"   内容: {result['content'][:100]}")
        print(f"   时间: {result['created_at']}")
        if result.get('synced'):
            print("   ☁️ 已同步到富通CRM")
    elif args.command == 'list':
        records = list_followups(args.customer, args.limit, args.type)
        if not records: print("暂无跟进记录"); return
        print(f"共 {len(records)} 条跟进记录:\n")
        for r in records:
            sync_mark = '✅' if r['synced'] else '⏳'
            print(f"  [{r['id']}] {r['created_at']} | {r['type']} | 客户{r['customer_id']} {r['customer_name']} {sync_mark}")
            print(f"       {r['content'][:150]}")
            print()
    elif args.command == 'search':
        records = search_followups(args.keyword)
        if not records: print(f"未找到包含「{args.keyword}」的记录"); return
        print(f"找到 {len(records)} 条匹配记录:\n")
        for r in records[:10]:
            sync_mark = '✅' if r['synced'] else '⏳'
            print(f"  [{r['id']}] {r['created_at']} | 客户{r['customer_id']} {r['customer_name']} | {r['type']} {sync_mark}")
            print(f"       {r['content'][:200]}\n")
    elif args.command == 'sync':
        print("运行: node sync_all_pending.mjs (需要CDP连接)")
    elif args.command == 'generate-excel':
        generate_excel()
    elif args.command == 'update-cache':
        update_customer_cache()
    else:
        parser.print_help()

if __name__ == '__main__':
    main()