#!/usr/bin/env python3
"""
🐾 塔奇克马 WhatsApp 自动开发信发送脚本
========================================
功能：
1. 从 BLIIoT_Customer_Leads_v2.xlsx 读取30家客户
2. 根据客户类别自动选择对应的开发信模板
3. 用 Selenium + Chrome 打开 WhatsApp Web
4. 自动发送个性化开发信给每个客户

首次使用：需手动扫 WhatsApp Web 二维码（仅一次，session 会保存）
后续使用：全自动发送！

用法：
  python whatsapp_auto_sender.py                    # 全部发送
  python whatsapp_auto_sender.py --dry-run          # 仅预览，不真发
  python whatsapp_auto_sender.py --start 1 --end 5  # 只发第1-5个
  python whatsapp_auto_sender.py --single 1         # 只发第1个
"""

import os
import sys
import time
import json
import argparse
import re
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import openpyxl

# ============================================================
# CONFIG
# ============================================================
EXCEL_PATH = r"C:\Users\Admin\Desktop\Working\BLIIoT_Customer_Leads_v2.xlsx"
CHROME_USER_DATA = r"C:\Users\Admin\AppData\Local\Google\Chrome\User Data"
CHROME_PROFILE = "Default"  # Use default Chrome profile (has your login sessions)
WAIT_TIMEOUT = 30  # seconds to wait for WhatsApp to load
MESSAGE_DELAY = 5  # seconds between messages (avoid rate limiting)
SEND_DELAY = 3  # seconds after sending before going to next

# Script home directory
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SESSION_FILE = os.path.join(SCRIPT_DIR, "wa_session_ready.txt")

# ============================================================
# TEMPLATES - 按客户类别分配开发信
# ============================================================
# Each template takes: company_name, country, specific_product_hook
TEMPLATES = {
    "System Integrator": """Hi {name} team,

I'm reaching out from BLIIoT (bliiot.com) - an industrial IoT manufacturer based in Shenzhen, China.

We specialize in ARMxy edge controllers, remote I/O modules (Modbus/OPC UA/MQTT/BACnet/IEC104), IoT gateways, and 4G industrial routers.

Our products are trusted in SmartGrid projects across Africa, including transformer anti-theft monitoring systems in Zimbabwe. As a system integrator, our flexible ARMxy (4000+ configurations) and IOy modules could be a great fit for your projects in {country}.

Would you be interested in receiving our product catalog and pricing? Happy to share datasheets tailored to your needs.

Best regards,
Kali
BLIIoT | bliiot.com""",

    "Automation Company": """Hi {name} team,

I'm reaching out from BLIIoT - we manufacture industrial IoT hardware including ARMxy edge controllers, remote I/O modules, IoT gateways, and 4G industrial routers.

Our products are designed for industrial automation scenarios and are already deployed in SmartGrid projects across Africa. What sets us apart is our flexibility - ARMxy series offers 4000+ configuration combinations, and our IOy modules support Modbus/OPC UA/MQTT/BACnet/IEC104.

I believe there could be great synergy between our products and your automation solutions in {country}. Happy to share our catalog and see if anything fits your projects.

Best regards,
Kali
BLIIoT | bliiot.com""",

    "Industrial Automation": """Hi {name} team,

I'm writing from BLIIoT (bliiot.com) - an industrial IoT manufacturer based in Shenzhen, China.

We manufacture ARMxy edge controllers, remote I/O modules (IOy series), IoT gateways, and 4G industrial routers. Our products are designed for challenging industrial environments: -45~80°C wide temperature, electrical isolation, and support for Modbus/OPC UA/MQTT/BACnet/IEC104 protocols.

We're already serving SmartGrid projects in Africa and would love to explore how our solutions could support your industrial automation needs in {country}.

Would you be interested in seeing our product catalog?

Best regards,
Kali
BLIIoT | bliiot.com""",

    "Machine Builder": """Hi {name} team,

I'm from BLIIoT (bliiot.com) - we manufacture ARMxy industrial controllers that can replace traditional PLCs in your machinery.

Our EdgePLC series supports CODESYS/OpenPLC and EtherCAT for real-time control. The ARMxy BL450 comes with a 6TOPS NPU, enabling AI-powered visual inspection on the edge.

Our products are already deployed in SmartGrid projects across Africa and could be a game-changer for your machine building in {country}.

Interested in seeing our machine control solutions?

Best regards,
Kali
BLIIoT | bliiot.com""",

    "Building Automation": """Hi {name} team,

I'm from BLIIoT - we have BACnet-certified solutions for building automation.

Our BA116 HVAC gateway and BA190 BACnet IP I/O modules are specifically designed for smart building projects. We also offer Modbus/OPC UA/MQTT options for comprehensive building management.

Our solutions are already deployed in projects across Africa and could be a great addition to your building automation portfolio in {country}.

Can I share our building automation catalog?

Best regards,
Kali
BLIIoT | bliiot.com""",

    "Home/Building Automation": """Hi {name} team,

I'm from BLIIoT - we have BACnet-certified solutions for building and home automation.

Our BA116 HVAC gateway and BA190 BACnet IP I/O modules are designed for smart building projects. We also offer compact IoT gateways and controllers for home automation scenarios.

Our solutions are trusted in projects across Africa and could complement your offerings in {country}.

Would you like to see our catalog?

Best regards,
Kali
BLIIoT | bliiot.com""",

    "Electrical Engineering": """Hi {name} team,

I'm writing from BLIIoT (bliiot.com) - an industrial IoT manufacturer based in Shenzhen, China.

We manufacture remote I/O modules (Modbus/OPC UA/MQTT), IoT gateways, and 4G industrial routers. Our IOy series modules are designed for electrical engineering applications - reliable data acquisition and remote monitoring for industrial projects.

Our products are already deployed in SmartGrid projects across Africa. Would love to explore how they could support your electrical engineering work in {country}.

Happy to share our catalog and pricing!

Best regards,
Kali
BLIIoT | bliiot.com""",

    "Industrial Distribution": """Hi {name} team,

I'm from BLIIoT (bliiot.com) - we're looking for distribution partners in {country}!

We manufacture a complete IIoT product line:
- ARMxy edge controllers (4000+ configurations)
- IOy remote I/O modules (Modbus/OPC UA/MQTT/BACnet/IEC104)
- IoT gateways and 4G industrial routers
- Industrial switches and signal isolators

Our products are already trusted in SmartGrid projects across Africa. We offer competitive pricing, technical support, and marketing materials.

Interested in partnership opportunities?

Best regards,
Kali
BLIIoT | bliiot.com""",

    "Automation Giant": """Hi {name} team,

I'm reaching out from BLIIoT (bliiot.com) - an industrial IoT manufacturer based in Shenzhen, China.

Our IOy remote I/O modules can serve as cost-effective remote I/O expansion for your existing control systems. They support OPC UA (BL191) and MQTT (BL192) for seamless integration with higher-level systems.

We see our products as complementary to your existing automation portfolio, not competitive. We're already deployed in SmartGrid projects across Africa.

Would you be interested in learning more about how our IO modules could expand your solution capabilities in {country}?

Best regards,
Kali
BLIIoT | bliiot.com""",

    "Process Automation": """Hi {name} team,

I'm from BLIIoT (bliiot.com) - we manufacture industrial IoT hardware for process automation.

Our IOy series remote I/O modules support OPC UA (BL191, with X.509 encryption), MQTT (BL192), and Modbus TCP (BL190) - perfect for connecting to your SCADA/MES systems. Our ARMxy edge controllers offer flexible computing for process control applications.

Our products are already deployed in SmartGrid projects across Africa and could support your process automation needs in {country}.

Would you like to see our catalog?

Best regards,
Kali
BLIIoT | bliiot.com""",

    "Control Systems": """Hi {name} team,

I'm from BLIIoT (bliiot.com) - we manufacture ARMxy edge controllers and IO modules for control systems applications.

Our ARMxy BL350/BL360 (TI/NXP processors) are ideal for motion control, while the EdgePLC BL234 supports CODESYS for IEC 61131-3 programming. The IOy series provides flexible remote I/O expansion.

We're already deployed in SmartGrid projects across Africa. Interested in seeing how our products fit your control systems in {country}?

Best regards,
Kali
BLIIoT | bliiot.com""",

    "Engineering Solutions": """Hi {name} team,

I'm from BLIIoT (bliiot.com) - an industrial IoT manufacturer based in Shenzhen, China.

We manufacture 4G industrial routers (R40 series with dual SIM), IoT gateways (BL116), and remote I/O modules (IOy series) - perfect for engineering projects requiring reliable remote monitoring and data acquisition.

Our products are already deployed in SmartGrid projects across Africa. Would love to explore how they could support your engineering solutions in {country}.

Happy to share our catalog!

Best regards,
Kali
BLIIoT | bliiot.com""",

    "Software/Automation": """Hi {name} team,

I'm from BLIIoT (bliiot.com) - we manufacture IoT hardware that integrates seamlessly with software platforms.

Our IOy BL191 module supports OPC UA with X.509 certificate security (IEC 62443 compliant), and BL192 supports MQTT for cloud platform integration (AWS IoT, ThingsBoard, etc.). ARMxy controllers support Docker and Python for custom application development.

Since your team combines IT expertise with automation, I believe our modern-protocol products would be a great fit for your projects in {country}.

Would you like to see our developer-friendly catalog?

Best regards,
Kali
BLIIoT | bliiot.com""",

    "Telecom/SmartGrid": """Hi {name} team,

I'm from BLIIoT (bliiot.com) - we already have a great partnership going with your team in Zimbabwe!

I wanted to share our latest products that could further strengthen your SmartGrid capabilities:
- ARMxy BL450 - 8-core edge controller with 6TOPS NPU
- BE116 - IEC104/IEC61850 substation gateway
- BE190 - Modbus to IEC104 converter
- R40A/R40B - 4G edge control terminal with integrated I/O

Our transformer anti-theft monitoring system is already running successfully. Let's talk about expanding to more sites!

Best regards,
Kali
BLIIoT | bliiot.com""",

    "SmartGrid/IoT": """Hi {name} team,

Great to reconnect! I'm from BLIIoT - we've been working together on the SmartGrid transformer anti-theft monitoring project in Zimbabwe.

I wanted to share our latest products that could expand your monitoring capabilities:
- ARMxy BL450 edge controller (6TOPS NPU for advanced analytics)
- BE116 IEC104 substation gateway
- More IOy modules for expanded sensor coverage
- R40A 4G edge control terminal

Our partnership is going strong and I'm excited about the next phase of our SmartGrid project. Let's discuss expansion!

Best regards,
Kali
BLIIoT | bliiot.com""",

    "default": """Hi {name} team,

I'm reaching out from BLIIoT (bliiot.com) - an industrial IoT manufacturer based in Shenzhen, China.

We specialize in ARMxy edge controllers, remote I/O modules, IoT gateways, 4G industrial routers, and SmartGrid solutions. Our products are trusted in projects across Africa, including transformer anti-theft monitoring in Zimbabwe.

I believe our solutions could be valuable for your operations in {country}. Would you be interested in receiving our product catalog?

Happy to share datasheets and pricing!

Best regards,
Kali
BLIIoT | bliiot.com"""
}

# ============================================================
# LOGO (ASCII for terminal only)
# ============================================================
LOGO = r"""
  ╔══════════════════════════════════════════╗
  ║   🐾 塔奇克马 WhatsApp Auto Sender      ║
  ║   BLIIoT Customer Outreach v1.0          ║
  ╚══════════════════════════════════════════╝
"""


def load_leads(excel_path, start=1, end=None):
    """Load customer leads from Excel"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["BLIIoT High-Potential Leads"]
    
    leads = []
    for row in range(2, ws.max_row + 1):
        lead = {
            "index": row - 1,
            "name": ws.cell(row=row, column=2).value or "",
            "description": str(ws.cell(row=row, column=3).value or "")[:100],
            "why": str(ws.cell(row=row, column=4).value or "")[:100],
            "phone": str(ws.cell(row=row, column=5).value or "").strip(),
            "website": ws.cell(row=row, column=6).value or "",
            "linkedin": ws.cell(row=row, column=7).value or "",
            "has_whatsapp": ws.cell(row=row, column=8).value or "",
            "country": ws.cell(row=row, column=9).value or "",
            "category": ws.cell(row=row, column=10).value or "",
            "rating": ws.cell(row=row, column=11).value or "",
        }
        # Parse phone number to just digits for wa.me
        phone_digits = re.sub(r'[^0-9]', '', lead["phone"])
        lead["phone_digits"] = phone_digits
        leads.append(lead)
    
    wb.close()
    
    # Filter by range
    if end is None:
        end = len(leads)
    filtered = [l for l in leads if start <= l["index"] <= end]
    return filtered


def get_message_template(lead):
    """Get the right template for this lead's category"""
    category = lead.get("category", "")
    country = lead.get("country", "your market")
    name = lead.get("name", "there")
    
    # Find matching template
    template_key = "default"
    for key in TEMPLATES:
        if key.lower() in category.lower():
            template_key = key
            break
    
    template = TEMPLATES.get(template_key, TEMPLATES["default"])
    message = template.format(name=name, country=country)
    return message


def clean_phone(country_code, phone):
    """Clean phone number to international format"""
    digits = re.sub(r'[^0-9]', '', phone)
    if not digits:
        return None
    # If starts with 0, remove the leading 0
    if digits.startswith('0'):
        digits = digits[1:]
    # Ensure it has country code
    return digits


def setup_driver():
    """Setup Chrome driver with existing profile (for WhatsApp Web session persistence)"""
    chrome_options = Options()
    
    # Use the user's existing Chrome profile to keep WhatsApp Web session
    chrome_options.add_argument(f"--user-data-dir={CHROME_USER_DATA}")
    chrome_options.add_argument(f"--profile-directory={CHROME_PROFILE}")
    
    # Don't run headless - user needs to scan QR code first time
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1280,900")
    chrome_options.add_argument("--remote-debugging-port=9222")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
    
    # If session file exists, we can try headless (but WhatsApp Web doesn't really support headless)
    driver = webdriver.Chrome(options=chrome_options)
    return driver


def wait_for_qr_scan(driver):
    """Wait for user to scan QR code on WhatsApp Web"""
    print("\n  ⏳ Waiting for WhatsApp Web to load...")
    
    try:
        # Wait for either the chat panel (logged in) or QR code (not logged in)
        WebDriverWait(driver, WAIT_TIMEOUT).until(
            EC.presence_of_element_located((By.XPATH, 
                "//div[@contenteditable='true'] | //canvas[@aria-label='Scan this QR code']"))
        )
        
        # Check if QR code is showing
        try:
            qr = driver.find_element(By.XPATH, "//canvas[@aria-label='Scan this QR code']")
            if qr:
                print("  📱 QR CODE SHOWING - Please scan with your WhatsApp mobile app!")
                print("  ⏳ Waiting up to 60 seconds for scan...")
                # Wait for QR to disappear (user scanned)
                WebDriverWait(driver, 60).until(
                    EC.invisibility_of_element_located((By.XPATH, "//canvas[@aria-label='Scan this QR code']"))
                )
                print("  ✅ QR scanned! WhatsApp Web ready!")
                # Save session flag
                with open(SESSION_FILE, "w") as f:
                    f.write(f"Session ready: {datetime.now().isoformat()}")
        except TimeoutException:
            print("  ✅ Already logged in to WhatsApp Web!")
            with open(SESSION_FILE, "w") as f:
                f.write(f"Session ready: {datetime.now().isoformat()}")
            
    except TimeoutException:
        print("  ⚠️ WhatsApp Web didn't load in time. Continuing anyway...")
        time.sleep(10)


def send_message(driver, phone_digits, message, lead_name):
    """Send a WhatsApp message via web.whatsapp.com"""
    print(f"\n  📤 Sending to {lead_name} ({phone_digits})...")
    
    try:
        # Navigate to the chat
        wa_url = f"https://web.whatsapp.com/send?phone={phone_digits}"
        driver.get(wa_url)
        
        # Wait for the message input box to appear
        time.sleep(3)
        
        # Wait for the text input to be ready
        msg_box = WebDriverWait(driver, WAIT_TIMEOUT).until(
            EC.presence_of_element_located((By.XPATH, 
                "//div[@contenteditable='true' and @data-tab='10']"))
        )
        
        # Small pause to ensure the chat is loaded
        time.sleep(2)
        
        # Type the message
        msg_box.click()
        time.sleep(0.5)
        
        # Send message line by line (more reliable for long messages)
        lines = message.split('\n')
        for i, line in enumerate(lines):
            msg_box.send_keys(line)
            if i < len(lines) - 1:
                msg_box.send_keys(Keys.SHIFT + Keys.ENTER)
                time.sleep(0.05)
        
        time.sleep(1)
        
        # Send it!
        msg_box.send_keys(Keys.ENTER)
        time.sleep(2)
        
        print(f"  ✅ Message sent to {lead_name}!")
        return True
        
    except TimeoutException:
        print(f"  ❌ Timeout - Couldn't send to {lead_name} (number may not be on WhatsApp)")
        return False
    except Exception as e:
        print(f"  ❌ Error sending to {lead_name}: {str(e)[:80]}")
        return False


def print_preview(leads):
    """Preview mode - just show what would be sent"""
    print(f"\n{'='*60}")
    print(f"  📋 DRY RUN - Preview of {len(leads)} messages")
    print(f"{'='*60}")
    
    for i, lead in enumerate(leads):
        message = get_message_template(lead)
        print(f"\n{'─'*60}")
        print(f"  [{lead['index']}] {lead['name']}")
        print(f"  📱 {lead['phone']} | 🌍 {lead['country']} | 🏷️ {lead['category']} | ⭐{lead['rating']}")
        print(f"{'─'*60}")
        # Show first 200 chars of message
        preview = message[:300] + ("..." if len(message) > 300 else "")
        print(f"  📝 {preview}")
        print(f"  (Full: {len(message)} chars)")


def main():
    parser = argparse.ArgumentParser(description="🐾 Tachikoma WhatsApp Auto Sender")
    parser.add_argument("--dry-run", action="store_true", help="Preview only, don't send")
    parser.add_argument("--start", type=int, default=1, help="Start index (1-based)")
    parser.add_argument("--end", type=int, default=None, help="End index (1-based)")
    parser.add_argument("--single", type=int, default=None, help="Single lead index")
    parser.add_argument("--category", type=str, default=None, help="Filter by category")
    args = parser.parse_args()
    
    print(LOGO)
    print(f"  📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print()
    
    # Load leads
    leads = load_leads(EXCEL_PATH)
    
    # Filter by single
    if args.single:
        leads = [l for l in leads if l["index"] == args.single]
    
    # Filter by range
    if args.start or args.end:
        end = args.end if args.end else len(leads)
        leads = [l for l in leads if args.start <= l["index"] <= end]
    
    # Filter by category
    if args.category:
        leads = [l for l in leads if args.category.lower() in l["category"].lower()]
    
    # Remove leads without valid phone
    valid_leads = [l for l in leads if l["phone_digits"] and len(l["phone_digits"]) >= 10]
    invalid_leads = [l for l in leads if not l["phone_digits"] or len(l["phone_digits"]) < 10]
    
    print(f"  📊 Total leads: {len(leads)}")
    print(f"  ✅ Valid WhatsApp numbers: {len(valid_leads)}")
    print(f"  ❌ Invalid/missing numbers: {len(invalid_leads)}")
    
    if invalid_leads:
        print(f"\n  ⚠️  Skipping (no valid phone):")
        for l in invalid_leads:
            print(f"     [{l['index']}] {l['name']} - phone: {l['phone']}")
    
    if not valid_leads:
        print("\n  ❌ No valid leads to process!")
        return
    
    # DRY RUN
    if args.dry_run:
        print_preview(valid_leads)
        print(f"\n{'='*60}")
        print(f"  ✅ DRY RUN COMPLETE - {len(valid_leads)} messages previewed")
        print(f"{'='*60}")
        return
    
    # REAL SEND
    print(f"\n{'='*60}")
    print(f"  🚀 Starting WhatsApp Auto Send - {len(valid_leads)} leads")
    print(f"  ⏱️  Estimated time: ~{len(valid_leads) * (MESSAGE_DELAY + 15)} seconds")
    print(f"{'='*60}")
    print()
    
    # Confirm
    print("  ⚠️  This will open Chrome and send messages via WhatsApp Web!")
    print(f"  ⚠️  Sending {len(valid_leads)} messages.")
    if not os.path.exists(SESSION_FILE):
        print("  📱 FIRST RUN: You'll need to scan QR code once!")
    print()
    
    try:
        input("  Press Enter to continue, Ctrl+C to cancel...")
    except KeyboardInterrupt:
        print("\n  Cancelled.")
        return
    
    # Launch Chrome
    print("\n  🚀 Launching Chrome for WhatsApp Web...")
    driver = setup_driver()
    
    try:
        # Go to WhatsApp Web
        print("  🌐 Opening web.whatsapp.com...")
        driver.get("https://web.whatsapp.com")
        
        # Handle QR scan
        wait_for_qr_scan(driver)
        
        # Send messages
        success = 0
        failed = 0
        
        for i, lead in enumerate(valid_leads):
            print(f"\n{'─'*60}")
            print(f"  [{i+1}/{len(valid_leads)}] {lead['name']}")
            print(f"  📱 {lead['phone']} | 🌍 {lead['country']} | 🏷️ {lead['category']}")
            print(f"{'─'*60}")
            
            message = get_message_template(lead)
            
            result = send_message(driver, lead["phone_digits"], message, lead["name"])
            
            if result:
                success += 1
            else:
                failed += 1
            
            # Delay between messages to avoid rate limiting
            if i < len(valid_leads) - 1:
                print(f"  ⏳ Waiting {MESSAGE_DELAY}s before next message...")
                time.sleep(MESSAGE_DELAY)
        
        # Summary
        print(f"\n{'='*60}")
        print(f"  📊 SEND COMPLETE!")
        print(f"  ✅ Success: {success}")
        print(f"  ❌ Failed: {failed}")
        print(f"  🕐 Finished: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        print(f"{'='*60}")
        
    except KeyboardInterrupt:
        print("\n\n  ⛔ Interrupted by user")
    finally:
        # Keep browser open so user can see results
        print("\n  💡 Browser will close in 10 seconds...")
        time.sleep(10)
        driver.quit()
        print("  ✅ Done!")


if __name__ == "__main__":
    main()