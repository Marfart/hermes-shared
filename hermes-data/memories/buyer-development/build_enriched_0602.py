import openpyxl, json, sys

wb = openpyxl.load_workbook('C:/Users/Admin/Desktop/Working/BLIIoT_Customer_Leads_v2.xlsx')
ws = wb['BLIIoT High-Potential Leads']

def infer_needs(desc, need, category):
    t = (desc + ' ' + need + ' ' + category).lower()
    needs = []
    if any(k in t for k in ['scada', 'plc', 'dcs', 'protocol']): needs.append('PLC connectivity and industrial protocol conversion')
    if any(k in t for k in ['remote monitoring', 'remote']): needs.append('remote monitoring gateway deployment')
    if any(k in t for k in ['automation', 'integrator']): needs.append('automation gateway and edge integration')
    if any(k in t for k in ['iiot', 'iot', 'industry 4.0']): needs.append('IIoT platform connectivity and field data access')
    if any(k in t for k in ['energy', 'utility', 'bems', 'building']): needs.append('energy monitoring and building/utility management')
    if any(k in t for k in ['edge', 'ai']): needs.append('industrial edge computing hardware')
    if any(k in t for k in ['cloud']): needs.append('cloud-connected industrial data transmission')
    if not needs: needs.append('IIoT platform connectivity and field data access')
    return needs

def recommend_products(desc, need, category):
    t = (desc + ' ' + need + ' ' + category).lower()
    prods = ['BLIIOT gateways/routers (R40, BA/BE/BL gateway series)', 'ARMxy edge computers/controllers']
    if any(k in t for k in ['monitor', 'data', 'acquisition', 'io']): prods.append('Remote IO / RTU / data acquisition modules')
    return prods

enriched = []
excluded_names = []
for r in range(2, ws.max_row+1):
    name = ws.cell(r,2).value
    if not name: continue
    desc = ws.cell(r,3).value or ''
    need = ws.cell(r,4).value or ''
    phone_raw = ws.cell(r,5).value or ''
    phone_str = str(phone_raw).replace(' ', '').replace('-', '')
    website = ws.cell(r,6).value or ''
    linkedin = ws.cell(r,7).value or ''
    country = ws.cell(r,9).value or ''
    category = ws.cell(r,10).value or ''

    if 'EXISTING' in (category or '') or 'ZODSAT' in (name or '') or 'Powertel' in (name or ''):
        excluded_names.append(name)
        continue

    needs = infer_needs(desc, need, category)
    prods = recommend_products(desc, need, category)

    domain = ''
    if website:
        domain = website.replace('https://','').replace('http://','').replace('www.','').rstrip('/')

    item = {
        'name': name,
        'country': country,
        'domain': domain,
        'desc': desc,
        'emailCount': 0,
        'reasons': ['公司业务与 IIoT Integration 匹配;'],
        'social': [],
        'normalizedDomain': domain,
        'officialEmails': [],
        'linkedinUrls': [linkedin] if linkedin else [],
        'whatsappUrls': [f'https://wa.me/{phone_str}'],
        'whatsappNumbers': [phone_str],
        'fetchedUrls': [website] if website else [],
        'fitScore': 100,
        'inferredNeeds': needs,
        'recommendedProducts': prods
    }
    enriched.append(item)

# Verify phone numbers are not masked
masked = [e for e in enriched if '****' in e['whatsappNumbers'][0]]
if masked:
    print(f"WARNING: {len(masked)} leads have masked numbers!")
    for m in masked:
        print(f"  {m['name']}: {m['whatsappNumbers'][0]}")
    sys.exit(1)

outpath = 'C:/Users/Admin/AppData/Local/hermes/memories/buyer-development/iiot_search_enriched_2026-06-02.json'
with open(outpath, 'w', encoding='utf-8') as f:
    json.dump(enriched, f, indent=2, ensure_ascii=False)

print(f"OK: {len(enriched)} leads written (excluded {len(excluded_names)} existing)")
print(f"File: {outpath}")
print(f"Phone sample: {enriched[0]['whatsappNumbers'][0]}")
print(f"Total chars: {len(json.dumps(enriched))}")