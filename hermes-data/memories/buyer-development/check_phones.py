import openpyxl
wb = openpyxl.load_workbook('C:/Users/Admin/Desktop/Working/BLIIoT_Customer_Leads_v2.xlsx')
ws = wb['BLIIoT High-Potential Leads']

for r in range(2, min(ws.max_row+1, 6)):
    name = ws.cell(r,2).value
    phone_cell = ws.cell(r,5)
    print(f"{r-1}. {name} | phone_value={repr(phone_cell.value)} | type={type(phone_cell.value)}")