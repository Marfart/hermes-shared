import xlrd

# Check the X板和Y板基本信息表 file
fpath = r"C:\Users\Admin\Desktop\Working\1.报价文件\X板和Y板基本信息表（2025_04_23）.xlsx"
try:
    import openpyxl
    wb = openpyxl.load_workbook(fpath)
    print(f"Sheets: {wb.sheetnames}")
    for sn in wb.sheetnames:
        sh = wb[sn]
        print(f"\n=== Sheet: {sn} ({sh.max_row}x{sh.max_column}) ===")
        for r in range(1, min(sh.max_row+1, 80)):
            vals = []
            for c in range(1, sh.max_column+1):
                v = sh.cell(r, c).value
                vals.append(str(v) if v is not None else "")
            line = " | ".join(vals)
            if any(x in line.upper() for x in ["BL193", "Y11", "Y31", "样品", "SAMPLE", "MODEL", "PRICE"]):
                print(f"R{r}: {line}")
            elif r < 6:
                print(f"R{r}: {line}")
except ImportError:
    print("openpyxl not installed")
    # Try xlrd for xlsx
    try:
        wb = xlrd.open_workbook(fpath)
        print(f"Sheets: {wb.sheet_names()}")
    except:
        print("Can't read with xlrd either")
