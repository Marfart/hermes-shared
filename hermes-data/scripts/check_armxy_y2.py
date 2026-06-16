import openpyxl

fpath = r"C:\Users\Admin\Desktop\Working\1.报价文件\202512 ARMxy Series Embedded Computer Price List.xlsx"
wb = openpyxl.load_workbook(fpath)
print(f"Sheets: {wb.sheetnames}")
for sn in wb.sheetnames:
    sh = wb[sn]
    print(f"\n=== Sheet: {sn} ({sh.max_row}x{sh.max_column}) ===")
    for r in range(1, min(sh.max_row+1, 100)):
        vals = []
        for c in range(1, sh.max_column+1):
            v = sh.cell(r, c).value
            vals.append(str(v) if v is not None else "")
        line = " | ".join(vals)
        if any(x in line.upper() for x in ["Y11", "Y31", "BL193", "样品", "SAMPLE", "MODEL", "PRICE", "Y板"]):
            print(f"R{r}: {line}")
        elif r < 10:
            print(f"R{r}: {line}")
