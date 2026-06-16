import openpyxl

fpath = r"C:\Users\Admin\Desktop\Working\1.报价文件\202512 ARMxy Series Embedded Computer Price List.xlsx"
wb = openpyxl.load_workbook(fpath)
sh = wb['BL350']

print(f"=== BL350 Sheet ({sh.max_row}x{sh.max_column}) ===")

# Print section headers
for r in range(1, sh.max_row+1):
    vals = []
    for c in range(1, sh.max_column+1):
        v = sh.cell(r, c).value
        vals.append(str(v) if v is not None else "")
    line = " | ".join(vals)
    if any(x in line for x in ["Model List", "Series", "Model |", "X series", "Y series", "SOM", "naming"]):
        print(f"R{r}: {line}")

print("\n\n=== ARMxy BL350 Model List ===")
# Model list section
for r in range(1, sh.max_row+1):
    r13 = str(sh.cell(r, 1).value or "")
    if r13.startswith("BL35") or r13.startswith("BL34"):
        vals = [str(sh.cell(r, c).value or "") for c in range(1, sh.max_column+1)]
        print(f"R{r}: {' | '.join(vals)}")

print("\n\n=== SOM Model List ===")
for r in range(1, sh.max_row+1):
    r1 = str(sh.cell(r, 1).value or "")
    if r1.startswith("SOM"):
        vals = [str(sh.cell(r, c).value or "") for c in range(1, sh.max_column+1)]
        print(f"R{r}: {' | '.join(vals)}")

print("\n\n=== X Board Model List ===")
for r in range(1, sh.max_row+1):
    r1 = str(sh.cell(r, 1).value or "")
    if r1.startswith("X"):
        vals = [str(sh.cell(r, c).value or "") for c in range(1, sh.max_column+1)]
        print(f"R{r}: {' | '.join(vals)}")
