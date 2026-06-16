import openpyxl

fpath = r"C:\Users\Admin\Desktop\Working\1.报价文件\202512 ARMxy Series Embedded Computer Price List.xlsx"
wb = openpyxl.load_workbook(fpath)
sh = wb['BL350']

# Print ALL rows of BL350 sheet
print(f"=== BL350 FULL ({sh.max_row}x{sh.max_column}) ===")
for r in range(1, sh.max_row+1):
    vals = [str(sh.cell(r, c).value or "") for c in range(1, sh.max_column+1)]
    line = " | ".join(vals)
    if line.strip():
        print(f"R{r}: {line}")
